"""Regression tests for defects found during the UI and workflow review."""
import csv
import pathlib
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_DOWN
from io import BytesIO, StringIO

from pypdf import PdfReader
from conftest import finalize_group

from attendance import db
from attendance.calculator import calculate_payroll_month
from attendance.models import AttendanceOverride, AttendanceRecord, AuditLog, Employee, Holiday, LeaveLedger, PayrollMonth, PayrollResult, SalaryRecord, WeekOffRule
from attendance.parser import implausible_session_minutes, parse_punch_times, working_minutes_from_punches
from attendance.settings import MONTHLY_RULES, monthly_rule_rows
from attendance.reports import build_error_report_pdf, build_manual_override_report_pdf, error_report_csv, manual_override_report_csv, pdf_money, total_paid_days
from attendance.utils import display_month, is_valid_payroll_month, leave_days


def login(client):
    client.post("/login", data={"username": "admin", "password": "12345"})


def seed_month(month="2026-07"):
    db.session.add(PayrollMonth(month=month))
    db.session.add(Employee(id="5", name="Worker", salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
    db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
    db.session.commit()


# --- Malformed payroll month keys used to raise IndexError from calendar.month_name ---

def test_is_valid_payroll_month_accepts_only_year_month():
    assert is_valid_payroll_month("2026-07")
    assert is_valid_payroll_month("2026-12")
    assert not is_valid_payroll_month("2026-13")
    assert not is_valid_payroll_month("2026-00")
    assert not is_valid_payroll_month("not-a-month")
    assert not is_valid_payroll_month("")
    assert not is_valid_payroll_month(None)


def test_display_month_is_safe_for_malformed_input():
    assert display_month("2026-07") == "July 2026"
    assert display_month("2026-99") == "2026-99"
    assert display_month(None) == "Not started"


def test_malformed_month_returns_404_instead_of_500(client, app):
    login(client)
    for path in (
        "/payroll/2026-99",
        "/payroll/not-a-month",
        "/payroll/2026-99/employee/5",
        "/attendance/2026-99",
        "/reports/2026-99/payroll-summary.csv",
        "/reports/2026-99/final-report.pdf",
    ):
        assert client.get(path).status_code == 404, path


def test_invalid_month_is_rejected_when_creating_payroll(client, app):
    login(client)
    response = client.post("/payroll/new", data={"month": "2026-99"}, follow_redirects=True)
    assert b"valid payroll month" in response.data
    with app.app_context():
        assert PayrollMonth.query.count() == 0


# --- A reversed In/Out pair silently rolled past midnight and was paid as overtime ---

def test_reversed_punch_pair_is_flagged_instead_of_paid():
    punches = parse_punch_times("06:30 PM\n09:32 AM")
    # The rollover still happens so genuine night shifts keep working ...
    assert working_minutes_from_punches(punches) == 902
    # ... but the resulting 15h 02m session is reported as implausible.
    assert implausible_session_minutes(punches) == 902


def test_normal_and_overnight_shifts_are_not_flagged():
    assert implausible_session_minutes(parse_punch_times("09:32 AM\n06:30 PM")) == 0
    # A genuine night shift rolls past midnight but stays within the limit.
    assert implausible_session_minutes(parse_punch_times("10:00 PM\n06:00 AM")) == 0
    assert implausible_session_minutes(parse_punch_times("08:00 PM\n08:00 AM")) == 0
    assert implausible_session_minutes(parse_punch_times("09:30 AM\n01:00 PM\n02:00 PM\n06:30 PM")) == 0
    assert implausible_session_minutes([]) == 0
    assert implausible_session_minutes(["09:30 AM"]) == 0


def test_long_same_day_shift_is_still_paid():
    """Real case from the sample data: in 09:35, out 22:19 is 12h 44m of real work.

    It is longer than the limit but never rolled past midnight, so it is overtime,
    not a punch error, and must not be withheld from payroll.
    """
    punches = parse_punch_times("09:35 AM\n10:19 PM")
    assert working_minutes_from_punches(punches) == 764
    assert implausible_session_minutes(punches) == 0


def test_attendance_manager_marks_reversed_punches_as_needing_review(client, app):
    with app.app_context():
        seed_month()
        record = AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker", date=date(2026, 7, 1), day="Wednesday", parse_status="OK")
        db.session.add(record)
        db.session.commit()
        record_id = record.id

    login(client)
    client.post("/attendance/2026-07", data={"action": "save", f"punches_{record_id}": "06:30 PM\n09:32 AM"}, follow_redirects=True)
    with app.app_context():
        saved = db.session.get(AttendanceRecord, record_id)
        assert saved.parse_status == "NEEDS_REVIEW"
        assert "Punch out before punch in" in saved.warning


# --- Attendance grid column order depended on which employee came first ---

def test_attendance_grid_dates_are_sorted_even_with_gaps(app):
    from routes.attendance_manager import attendance_grid

    with app.app_context():
        seed_month()
        db.session.add(Employee(id="9", name="Later Joiner", salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("10000")))
        # Employee "5" sorts first but is missing 1 July, so first-seen ordering
        # would have appended that date after all of employee 5's dates.
        for day in (2, 3):
            db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker", date=date(2026, 7, day)))
        for day in (1, 2, 3):
            db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="9", employee_name="Later Joiner", date=date(2026, 7, day)))
        db.session.commit()

        dates, rows = attendance_grid("2026-07")
        assert dates == [date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 3)]
        assert dates == sorted(dates)
        assert len(rows) == 2


def test_attendance_grid_separates_missing_punch_days_from_week_offs(app):
    from routes.attendance_manager import attendance_grid

    with app.app_context():
        seed_month()
        # 1 July 2026 is a Wednesday (working day), 5 July 2026 is a Sunday (week off).
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker", date=date(2026, 7, 1)))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker", date=date(2026, 7, 5)))
        db.session.commit()

        _dates, rows = attendance_grid("2026-07")
        cells = rows[0]["cells"]
        assert cells[date(2026, 7, 1)]["missing"] is True
        assert cells[date(2026, 7, 5)]["missing"] is False
        assert cells[date(2026, 7, 5)]["week_off"] is True
        assert rows[0]["missing_count"] == 1
        assert rows[0]["needs_review"] is True


def test_attendance_filter_can_target_other_punch_issues(client, app):
    from routes.attendance_manager import attendance_grid

    with app.app_context():
        seed_month()
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 1), day="Wednesday",
                                        first_punch="09:30 AM", last_punch="03:30 AM",
                                        parse_status="NEEDS_REVIEW",
                                        warning="Punch out before punch in (18h 00m session)"))
        db.session.commit()

        _dates, rows = attendance_grid("2026-07")
        assert rows[0]["other_issue_count"] == 1
        assert rows[0]["has_other_issue"] is True
        assert rows[0]["needs_review"] is True

    login(client)
    page = client.get("/attendance/2026-07").data
    assert page.count(b'class="form-check-input issue-filter-choice"') == 4
    assert b'value="missing"> No punch days' in page
    assert b'value="odd"> Odd punch' in page
    assert b'value="other"> Other issues' in page
    assert b'data-has-other-issue="1"' in page
    assert b"other-issue-cell" in page


# --- Dashboard counted DAILY wage employees as unsupported ---

def test_dashboard_reports_daily_as_supported_wage_type(client, app):
    with app.app_context():
        seed_month()
        db.session.add(Employee(id="6", name="Day Worker", salary_type="Daily", normalized_salary_type="DAILY", salary=Decimal("500")))
        db.session.add(WeekOffRule(employee_id="6", confirmed_at=datetime.utcnow()))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker", salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="6", name="Day Worker", salary_type="Daily", normalized_salary_type="DAILY", salary=Decimal("500")))
        db.session.commit()

    login(client)
    page = client.get("/")
    assert b"1 monthly, 1 daily" in page.data
    assert b"unsupported" not in page.data


# --- Settings page listed rule keys that nothing read ---

def test_settings_rules_are_all_live_configuration():
    rows = monthly_rule_rows()
    assert {row["key"] for row in rows} == set(MONTHLY_RULES)
    assert all(row["label"] and row["detail"] for row in rows)
    for dead_key in ("SHIFT_START_MINUTES", "SHIFT_END_MINUTES", "GRACE_MINUTES", "LEAVE_EARN_DIVISOR"):
        assert dead_key not in MONTHLY_RULES


def test_settings_page_renders_rule_effects(client, app):
    login(client)
    page = client.get("/settings")
    assert b"Full-day grace threshold" in page.data
    assert b"8h 50m" in page.data
    assert b"SHIFT_START_MINUTES" not in page.data


# --- Leave accrual rate is configuration, not a magic number ---

def test_leave_earned_uses_configured_monthly_rate(app):
    from attendance.payroll_rules import calculate_monthly_leave_earned

    assert MONTHLY_RULES["LEAVE_EARNED_PER_MONTH"] == 2
    assert calculate_monthly_leave_earned(Decimal("31"), 31) == Decimal("2.0")
    assert calculate_monthly_leave_earned(Decimal("15.5"), 31) == Decimal("1.0")
    assert calculate_monthly_leave_earned(Decimal("0"), 31) == Decimal("0.0")
    assert calculate_monthly_leave_earned(Decimal("31"), 0) == Decimal("0.0")


# --- Flash messages could not be dismissed (Bootstrap markup, no Bootstrap JS) ---

def test_flash_messages_use_working_dismiss_hooks(client, app):
    login(client)
    response = client.post("/payroll/new", data={"month": ""}, follow_redirects=True)
    assert b"data-dismiss-alert" in response.data
    assert b"data-bs-dismiss" not in response.data


# --- The payroll month page now guides the user through the workflow ---

def test_payroll_month_page_shows_workflow_steps(client, app):
    with app.app_context():
        seed_month()

    login(client)
    page = client.get("/payroll/2026-07")
    for label in (b"Import attendance", b"Load wages", b"Review &amp; submit", b"Calculate payroll", b"Finalize"):
        assert label in page.data
    # First unfinished step is highlighted as the current one.
    assert b"workflow-step is-current" in page.data
    # Both recalculate buttons sit on the Calculate payroll step, and the
    # Run calculation panel they once lived in is gone.
    assert b">Recalculate</button>" in page.data
    assert b">Reset &amp; recalculate</button>" in page.data
    assert b'id="run-calculation"' not in page.data
    # The stale "Monthly only" claim is gone now that DAILY has a rule.
    assert b"Wage Type Monthly only" not in page.data


def test_payroll_workflow_steps_advance_with_progress(client, app):
    with app.app_context():
        seed_month()
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker", date=date(2026, 7, 1), day="Wednesday", first_punch="09:30 AM", last_punch="06:30 PM", raw_working_hours="9h 00m", actual_minutes=540, parse_status="OK"))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker", salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        month = db.session.get(PayrollMonth, "2026-07")
        month.attendance_submitted = True
        db.session.commit()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07")
    # Four steps done, only Finalize outstanding.
    assert page.data.count(b"workflow-step is-done") == 4
    assert page.data.count(b"workflow-step is-current") == 1


# --- Department and designation are editable master fields, sourced from attendance ---

def test_attendance_import_populates_department_and_designation(client, app, tmp_path):
    from attendance.parser import import_attendance_csv

    csv_path = tmp_path / "attendance.csv"
    csv_path.write_text(
        "Employee ID,Employee Name,Department,Designation,Date,Day,Shift,From,To,First Punch,Last Punch,Total Working Hours\n"
        "5,Worker,Electrical Production,Wireman,01-07-2026,Wednesday,Normal,,,09:30 AM,06:30 PM,9h 00m\n",
        encoding="utf-8",
    )
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()
        import_attendance_csv(str(csv_path), "2026-07", "admin")
        employee = db.session.get(Employee, "5")
        assert employee.department == "Electrical Production"
        assert employee.designation == "Wireman"


def test_add_employee_accepts_department_and_designation(client, app):
    login(client)
    response = client.post("/master", data={
        "employee_id": "77",
        "name": "Manual Entry",
        "department": "Design",
        "designation": "Design Engineer",
        "wage_type": "Monthly",
        "salary": "25000",
    }, follow_redirects=True)
    assert b"Employee master saved" in response.data
    with app.app_context():
        employee = db.session.get(Employee, "77")
        assert employee.department == "Design"
        assert employee.designation == "Design Engineer"
        audit = AuditLog.query.filter_by(action="Employee Master Created").one()
        assert "Department Design" in audit.detail
        assert "Designation Design Engineer" in audit.detail


def test_editing_employee_records_department_change_in_audit_log(client, app):
    with app.app_context():
        seed_month()
        employee = db.session.get(Employee, "5")
        employee.department = "Service"
        employee.designation = "Jr. Service Engineer"
        db.session.commit()

    login(client)
    client.post("/master", data={
        "employee_id": "5",
        "name": "Worker",
        "department": "Design",
        "designation": "Design Engineer",
        "wage_type": "Monthly",
        "salary": "30000",
        "master_controls_present": "1",
    }, follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.department == "Design"
        assert employee.designation == "Design Engineer"
        audit = AuditLog.query.filter_by(action="Employee Master Updated").first()
        assert "Department Service -> Design" in audit.detail
        assert "Designation Jr. Service Engineer -> Design Engineer" in audit.detail


def test_master_import_updates_new_columns_and_tolerates_old_format(client, app):
    from io import BytesIO

    with app.app_context():
        seed_month()
        employee = db.session.get(Employee, "5")
        employee.department = "Service"
        employee.designation = "Helper"
        db.session.commit()

    login(client)
    # New format carries both columns.
    client.post("/master/import", data={"employee_master_csv": (
        BytesIO(b"Employee ID,Name,Department,Designation,Wage Type,Salary\n5,Worker,Accounts,Accounts Executive,Monthly,31000\n"),
        "master.csv")}, content_type="multipart/form-data", follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.department == "Accounts"
        assert employee.designation == "Accounts Executive"
        assert Decimal(employee.salary) == Decimal("31000")

    # An older export without those columns must leave the stored values untouched.
    client.post("/master/import", data={"employee_master_csv": (
        BytesIO(b"Employee ID,Name,Wage Type,Salary\n5,Worker,Monthly,32000\n"),
        "old.csv")}, content_type="multipart/form-data", follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.department == "Accounts"
        assert employee.designation == "Accounts Executive"
        assert Decimal(employee.salary) == Decimal("32000")


# --- Monthly and Daily payroll are finalized independently ---

def seed_mixed_month(month="2026-07"):
    """One monthly and one daily employee with a full-day punch each."""
    db.session.add(PayrollMonth(month=month, attendance_submitted=True))
    for employee_id, name, wage, salary in (("5", "Monthly Worker", "Monthly", "30000"), ("6", "Daily Worker", "Daily", "600")):
        db.session.add(Employee(id=employee_id, name=name, salary_type=wage,
                                normalized_salary_type=wage.upper(), salary=Decimal(salary)))
        db.session.add(WeekOffRule(employee_id=employee_id, confirmed_at=datetime.utcnow()))
        db.session.add(SalaryRecord(payroll_month=month, employee_id=employee_id, name=name,
                                    salary_type=wage, normalized_salary_type=wage.upper(), salary=Decimal(salary)))
        db.session.add(AttendanceRecord(payroll_month=month, employee_id=employee_id, employee_name=name,
                                        date=date(2026, 7, 1), day="Wednesday", first_punch="09:30 AM",
                                        last_punch="06:30 PM", raw_working_hours="9h 00m",
                                        actual_minutes=540, parse_status="OK"))
    db.session.commit()


def test_finalizing_monthly_leaves_daily_open(client, app):
    from attendance.wage_groups import is_group_finalized

    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    response = client.post("/payroll/2026-07", data={
        "action": "finalize", "wage_group": "MONTHLY", "admin_password": "12345",
    }, follow_redirects=True)
    assert b"Monthly wage payroll finalized and locked" in response.data

    with app.app_context():
        month = db.session.get(PayrollMonth, "2026-07")
        assert is_group_finalized(month, "MONTHLY")
        assert not is_group_finalized(month, "DAILY")
        # The month as a whole is not finalized while Daily is still open.
        assert month.status == "DRAFT"


def test_month_status_finalizes_only_when_every_group_is_locked(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    for group in ("MONTHLY", "DAILY"):
        client.post("/payroll/2026-07", data={
            "action": "finalize", "wage_group": group, "admin_password": "12345",
        }, follow_redirects=True)
    with app.app_context():
        month = db.session.get(PayrollMonth, "2026-07")
        assert month.status == "FINALIZED"
        assert month.finalized_at is not None

    # Unlocking one group reopens the month.
    client.post("/payroll/2026-07", data={
        "action": "unlock", "wage_group": "DAILY", "admin_password": "12345",
    }, follow_redirects=True)
    with app.app_context():
        month = db.session.get(PayrollMonth, "2026-07")
        assert month.status == "DRAFT"
        assert month.finalized_at is None


def test_recalculation_never_touches_a_finalized_wage_group(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")
        monthly_before = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        frozen_salary = Decimal(monthly_before.final_salary)

    login(client)
    client.post("/payroll/2026-07", data={
        "action": "finalize", "wage_group": "MONTHLY", "admin_password": "12345",
    }, follow_redirects=True)

    with app.app_context():
        # Change something that would alter the monthly result if it were recalculated.
        salary = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        salary.salary = Decimal("90000")
        db.session.commit()
        calculate_payroll_month("2026-07")
        monthly_after = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(monthly_after.final_salary) == frozen_salary
        # Daily is still open and was recalculated.
        assert PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").count() == 1


def test_employee_edit_locks_follow_the_employees_wage_group(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    client.post("/payroll/2026-07", data={
        "action": "finalize", "wage_group": "MONTHLY", "admin_password": "12345",
    }, follow_redirects=True)

    monthly = client.post("/payroll/2026-07/employee/5", data={"action": "save", "adjustment": "500"}, follow_redirects=True)
    assert b"finalized and locked" in monthly.data
    daily = client.post("/payroll/2026-07/employee/6", data={"action": "save", "adjustment": "500"}, follow_redirects=True)
    assert b"Employee changes saved" in daily.data
    with app.app_context():
        assert Decimal(SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="5").one().adjustment) == Decimal("0")
        assert Decimal(SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="6").one().adjustment) == Decimal("500")


def test_employee_detail_links_to_that_employees_attendance_summary(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    for employee_id, included, excluded in (("5", "Monthly Worker", "Daily Worker"), ("6", "Daily Worker", "Monthly Worker")):
        page = client.get(f"/payroll/2026-07/employee/{employee_id}")
        assert page.status_code == 200
        assert b"Attendance summary" in page.data
        assert f"/reports/2026-07/employee/{employee_id}/attendance-summary.pdf".encode() in page.data

        response = client.get(f"/reports/2026-07/employee/{employee_id}/attendance-summary.pdf")
        assert response.status_code == 200
        text = pdf_text(response.data)
        assert included in text
        assert excluded not in text


def test_payroll_page_offers_separate_locks_and_a_wage_filter(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07")
    assert b"Finalize monthly" in page.data
    assert b"Finalize daily" in page.data
    assert b"Monthly wage employees" in page.data
    assert b"Daily wage employees" in page.data

    monthly_only = client.get("/payroll/2026-07?wage=monthly")
    assert b"Monthly wage employees" in monthly_only.data
    assert b"Daily wage employees" not in monthly_only.data


# --- Redundant panels removed; the stepper carries the actions itself ---

def test_workflow_step_loads_wages_without_the_removed_panel(client, app):
    """The Wage Master panel is gone, so step 2 must run the sync, not just link to master."""
    with app.app_context():
        seed_mixed_month()
        SalaryRecord.query.delete()
        db.session.commit()
        assert SalaryRecord.query.filter_by(payroll_month="2026-07").count() == 0

    login(client)
    page = client.get("/payroll/2026-07")
    # The duplicated panels are gone ...
    assert b"Import or re-import the attendance CSV/XLSX" not in page.data
    assert b"Load wage from master" not in page.data
    # ... and step 2 offers the action that panel used to own.
    assert b"Load wages" in page.data

    response = client.post("/payroll/2026-07", data={"action": "salary"}, follow_redirects=True)
    assert b"Wage data loaded from master" in response.data
    with app.app_context():
        assert SalaryRecord.query.filter_by(payroll_month="2026-07").count() == 2


def test_wage_step_action_stays_available_after_completion(client, app):
    """Salaries change in master, so reloading wages must not disappear once done."""
    with app.app_context():
        seed_mixed_month()

    login(client)
    page = client.get("/payroll/2026-07")
    assert b"Reload wages" in page.data

    with app.app_context():
        employee = db.session.get(Employee, "5")
        employee.salary = Decimal("45000")
        db.session.commit()
    client.post("/payroll/2026-07", data={"action": "salary"}, follow_redirects=True)
    with app.app_context():
        salary = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(salary.salary) == Decimal("45000")


def test_wage_step_action_is_hidden_once_a_group_is_finalized(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    client.post("/payroll/2026-07", data={
        "action": "finalize", "wage_group": "MONTHLY", "admin_password": "12345",
    }, follow_redirects=True)
    page = client.get("/payroll/2026-07")
    # The wage step has no actions left once the group is locked: the reload button
    # is gone and the Employee Master link was removed to keep the row aligned.
    assert b"Reload wages" not in page.data
    assert step_actions(page.data, "Load wages") == []


# --- Monthly salary breakup and compliance flags ---

def add_employee(client, **overrides):
    data = {
        "employee_id": "80", "name": "Breakup Worker", "wage_type": "Monthly",
        "salary": "30000", "master_controls_present": "1",
    }
    data.update(overrides)
    return client.post("/master", data=data, follow_redirects=True)


def test_salary_breakup_must_add_up_to_salary(client, app):
    login(client)
    bad = add_employee(client, basic_salary="15000", hra="6000", allowance="5000")
    assert b"must add up to the salary" in bad.data
    assert b"Difference 4000.00" in bad.data
    with app.app_context():
        assert db.session.get(Employee, "80") is None

    good = add_employee(client, basic_salary="15000", hra="6000", allowance="9000")
    assert b"Employee master saved" in good.data
    with app.app_context():
        employee = db.session.get(Employee, "80")
        assert Decimal(employee.basic_salary) == Decimal("15000")
        assert Decimal(employee.hra) == Decimal("6000")
        assert Decimal(employee.allowance) == Decimal("9000")
        total = employee.basic_salary + employee.hra + employee.allowance
        assert Decimal(total) == Decimal(employee.salary)


def test_breakup_may_be_left_empty(client, app):
    """An all-zero breakup means "not captured yet" and must stay valid."""
    login(client)
    response = add_employee(client, basic_salary="0", hra="0", allowance="0")
    assert b"Employee master saved" in response.data
    with app.app_context():
        assert Decimal(db.session.get(Employee, "80").basic_salary) == Decimal("0")


def test_compliance_flags_save_for_monthly(client, app):
    login(client)
    add_employee(client, pf_enabled="on", esic_enabled="on")
    with app.app_context():
        employee = db.session.get(Employee, "80")
        assert employee.pf_enabled is True
        assert employee.esic_enabled is True
        audit = AuditLog.query.filter_by(action="Employee Master Created").one()
        assert "PF Yes" in audit.detail and "ESIC Yes" in audit.detail


def test_daily_wage_never_stores_breakup_or_compliance(client, app):
    """The fields are monthly-only, so a daily row must not keep values for them."""
    login(client)
    client.post("/master", data={
        "employee_id": "81", "name": "Day Worker", "wage_type": "Daily", "salary": "600",
        "master_controls_present": "1",
        "basic_salary": "400", "hra": "100", "allowance": "100",
        "pf_enabled": "on", "esic_enabled": "on",
    }, follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "81")
        assert Decimal(employee.basic_salary) == Decimal("0")
        assert Decimal(employee.hra) == Decimal("0")
        assert employee.pf_enabled is False
        assert employee.esic_enabled is False


# --- Import identity lock ---

def test_import_cannot_rename_an_existing_employee(client, app):
    from io import BytesIO

    with app.app_context():
        db.session.add(Employee(id="1", name="Manish Hirani", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    blocked = client.post("/master/import", data={"employee_master_csv": (
        BytesIO(b"Employee ID,Name,Wage Type,Salary\n1,Manish C Hirani,Monthly,32000\n"), "master.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"already named" in blocked.data
    assert b"Edit the name on the employee page instead" in blocked.data
    with app.app_context():
        employee = db.session.get(Employee, "1")
        assert employee.name == "Manish Hirani"
        # The whole import is rejected, so the salary change does not land either.
        assert Decimal(employee.salary) == Decimal("30000")

    # Matching name is accepted and the wage fields update.
    allowed = client.post("/master/import", data={"employee_master_csv": (
        BytesIO(b"Employee ID,Name,Wage Type,Salary\n1,Manish Hirani,Monthly,32000\n"), "master.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"Employee master imported" in allowed.data
    with app.app_context():
        assert Decimal(db.session.get(Employee, "1").salary) == Decimal("32000")

    # The name is still editable from the employee page.
    client.post("/master", data={"employee_id": "1", "name": "Manish C Hirani",
                                 "wage_type": "Monthly", "salary": "32000"}, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Employee, "1").name == "Manish C Hirani"


def test_import_round_trips_breakup_and_compliance(client, app):
    from io import BytesIO

    with app.app_context():
        db.session.add(Employee(id="1", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    ok = client.post("/master/import", data={"employee_master_csv": (BytesIO(
        b"Employee ID,Name,Wage Type,Salary,Basic,HRA,Allowance,TDS,PF,ESIC\n"
        b"1,Worker,Monthly,30000,15000,6000,9000,2500,Yes,No\n"), "master.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"Employee master imported" in ok.data
    with app.app_context():
        employee = db.session.get(Employee, "1")
        assert Decimal(employee.basic_salary) == Decimal("15000")
        assert employee.pf_enabled is True
        assert employee.esic_enabled is False

    export = client.get("/master/export.csv")
    assert b"Basic,HRA,Allowance,TDS,PF,ESIC" in export.data
    assert b"15000.00,6000.00,9000.00,2500.00,Yes,No" in export.data

    bad = client.post("/master/import", data={"employee_master_csv": (BytesIO(
        b"Employee ID,Name,Wage Type,Salary,Basic,HRA,Allowance\n"
        b"1,Worker,Monthly,30000,15000,6000,12000\n"), "master.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"must add up to the salary" in bad.data


def test_import_rejects_breakup_on_a_daily_employee(client, app):
    from io import BytesIO

    with app.app_context():
        db.session.add(Employee(id="2", name="Day Worker", salary_type="Daily",
                                normalized_salary_type="DAILY", salary=Decimal("600")))
        db.session.commit()

    login(client)
    response = client.post("/master/import", data={"employee_master_csv": (BytesIO(
        b"Employee ID,Name,Wage Type,Salary,Basic\n2,Day Worker,Daily,600,400\n"), "master.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"only applies to monthly wage employees" in response.data
    with app.app_context():
        assert Decimal(db.session.get(Employee, "2").basic_salary) == Decimal("0")


def test_older_export_without_new_columns_still_imports(client, app):
    from io import BytesIO

    with app.app_context():
        db.session.add(Employee(id="1", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000"),
                                basic_salary=Decimal("15000"), hra=Decimal("6000"),
                                allowance=Decimal("9000"), pf_enabled=True))
        db.session.commit()

    login(client)
    response = client.post("/master/import", data={"employee_master_csv": (
        BytesIO(b"Employee ID,Name,Wage Type,Salary\n1,Worker,Monthly,30000\n"), "old.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"Employee master imported" in response.data
    with app.app_context():
        employee = db.session.get(Employee, "1")
        # Columns absent from the file leave the stored values untouched.
        assert Decimal(employee.basic_salary) == Decimal("15000")
        assert employee.pf_enabled is True


# --- "Ignore OT" replaced an inverted "OT eligible" flag ---

def test_ignore_ot_suppresses_payable_overtime(app):
    """Ticking Ignore OT must zero payable OT while still reporting raw OT."""
    with app.app_context():
        seed_month()
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 1), day="Wednesday", first_punch="09:30 AM",
                                        last_punch="08:30 PM", raw_working_hours="11h 00m",
                                        actual_minutes=660, parse_status="OK"))
        db.session.commit()

        calculate_payroll_month("2026-07")
        paid = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert paid.ot_minutes > 0
        assert paid.payable_ot_minutes > 0
        assert Decimal(paid.ot_amount) > 0

        employee = db.session.get(Employee, "5")
        employee.ot_ignored = True
        db.session.commit()
        calculate_payroll_month("2026-07")
        ignored = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert ignored.ot_minutes > 0, "raw OT is still reported"
        assert ignored.payable_ot_minutes == 0
        assert Decimal(ignored.ot_amount) == Decimal("0")


def test_ignore_less_hours_suppresses_shortage_deduction(app):
    with app.app_context():
        seed_month()
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 1), day="Wednesday", first_punch="09:30 AM",
                                        last_punch="05:00 PM", raw_working_hours="7h 30m",
                                        actual_minutes=450, parse_status="OK"))
        db.session.commit()

        calculate_payroll_month("2026-07")
        deducted = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert deducted.less_hours_minutes > 0
        assert Decimal(deducted.less_hours_deduction) > 0

        employee = db.session.get(Employee, "5")
        employee.less_hours_ignored = True
        db.session.commit()
        calculate_payroll_month("2026-07")
        ignored = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert ignored.less_hours_minutes == 0
        assert Decimal(ignored.less_hours_deduction) == Decimal("0")


def test_new_employees_default_to_ignoring_nothing(client, app):
    login(client)
    client.post("/master", data={
        "employee_id": "95", "name": "Default Worker", "wage_type": "Monthly",
        "salary": "20000", "master_controls_present": "1",
    }, follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "95")
        assert employee.ot_ignored is False
        assert employee.less_hours_ignored is False


def test_ignore_flags_save_and_are_audited(client, app):
    login(client)
    client.post("/master", data={
        "employee_id": "96", "name": "Exception Worker", "wage_type": "Monthly",
        "salary": "20000", "master_controls_present": "1",
        "ot_ignored": "on", "less_hours_ignored": "on",
    }, follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "96")
        assert employee.ot_ignored is True
        assert employee.less_hours_ignored is True
        audit = AuditLog.query.filter_by(action="Employee Master Created").one()
        assert "Ignore OT Yes" in audit.detail
        assert "Ignore Less Hours Yes" in audit.detail


def test_legacy_ot_enabled_column_migrates_inverted(tmp_path):
    """An existing DB with the old ot_enabled column must flip cleanly to ot_ignored."""
    import sqlite3
    from app import create_app

    db_path = tmp_path / "legacy.db"
    connection = sqlite3.connect(db_path)
    connection.executescript(
        """
        CREATE TABLE employee (
            id VARCHAR(64) PRIMARY KEY,
            name VARCHAR(160) NOT NULL,
            department VARCHAR(160),
            designation VARCHAR(160),
            ot_enabled BOOLEAN NOT NULL DEFAULT 1,
            less_hours_exempt BOOLEAN NOT NULL DEFAULT 0,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        INSERT INTO employee (id, name, ot_enabled, less_hours_exempt) VALUES
            ('1', 'OT Paid', 1, 0),
            ('2', 'OT Blocked', 0, 1);
        """
    )
    connection.commit()
    connection.close()

    app = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///" + str(db_path), "TESTING": True})
    with app.app_context():
        # ot_enabled=1 ("OT paid") becomes ot_ignored=0, and vice versa.
        assert db.session.get(Employee, "1").ot_ignored is False
        assert db.session.get(Employee, "1").less_hours_ignored is False
        assert db.session.get(Employee, "2").ot_ignored is True
        assert db.session.get(Employee, "2").less_hours_ignored is True


# --- Workflow steps are gated on the steps before them ---

def step_actions(html, label):
    """The buttons/links inside one workflow step, with their disabled state."""
    import re
    blocks = re.findall(rb'<li class="workflow-step[^>]*>(.*?)</li>', html, re.S)
    for block in blocks:
        if b"<strong>" + label.encode() + b"</strong>" in block:
            return [
                (re.sub(rb"<[^>]+>", b"", m.group(2)).strip().decode(), b"disabled" in m.group(1))
                for m in re.finditer(rb"<(?:button|a)([^>]*)>(.*?)</(?:button|a)>", block, re.S)
            ]
    return []


def test_later_steps_are_disabled_until_attendance_is_imported(client, app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.commit()

    login(client)
    page = client.get("/payroll/2026-07").data

    # Step 1 has no prerequisite, so it stays actionable.
    assert step_actions(page, "Import attendance") == [("Import attendance", False)]
    # Everything after it is inert until attendance lands.
    for label in ("Load wages", "Review &amp; submit", "Finalize"):
        actions = step_actions(page, label)
        assert actions, label
        assert all(disabled for _text, disabled in actions), f"{label} should be disabled: {actions}"
    # Calculate payroll carries both recalculate buttons, disabled until attendance lands.
    calculate = step_actions(page, "Calculate payroll")
    assert [text for text, _ in calculate] == ["Recalculate", "Reset &amp; recalculate"]
    assert all(disabled for _text, disabled in calculate)
    assert b"disabled-link" in page
    assert b"Complete the earlier steps first" in page


def test_step_two_unlocks_once_attendance_is_imported(client, app):
    with app.app_context():
        seed_month()
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 1), day="Wednesday", first_punch="09:30 AM",
                                        last_punch="06:30 PM", raw_working_hours="9h 00m",
                                        actual_minutes=540, parse_status="OK"))
        db.session.commit()

    login(client)
    page = client.get("/payroll/2026-07").data
    assert all(not disabled for _text, disabled in step_actions(page, "Load wages"))
    # Step 3 still waits on step 2.
    assert all(disabled for _text, disabled in step_actions(page, "Review &amp; submit"))


def test_completed_steps_stay_actionable(client, app):
    """A done step keeps its action so wages can be reloaded after a master change."""
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07").data
    for label in ("Import attendance", "Load wages", "Review &amp; submit", "Finalize"):
        actions = step_actions(page, label)
        assert actions, label
        assert all(not disabled for _text, disabled in actions), f"{label} should be enabled: {actions}"
    # Both recalculate buttons live on the step that names the action.
    calculate = step_actions(page, "Calculate payroll")
    assert [text for text, _ in calculate] == ["Recalculate", "Reset &amp; recalculate"]
    assert all(not disabled for _text, disabled in calculate)


def test_header_carries_only_the_page_level_action(client, app):
    """Recalculating belongs to its workflow step, not to three red header buttons."""
    with app.app_context():
        seed_mixed_month()

    login(client)
    page = client.get("/payroll/2026-07").data
    header = page.split(b'<div class="page-head">')[1].split(b"</div>\n</div>")[0]
    assert b"Delete payroll" in header
    for moved in (b"Recalculate", b"Reset", b"Attendance Manager"):
        assert moved not in header, moved


def test_finalize_step_cta_is_named_finalize(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07").data
    assert ("Finalize", False) in step_actions(page, "Finalize")
    assert b"Go to locks" not in page


def test_each_step_carries_only_its_own_actions(client, app):
    """Every stage owns its buttons, so the stepper is the whole flow."""
    with app.app_context():
        seed_mixed_month()

    login(client)
    page = client.get("/payroll/2026-07").data
    for label in ("Import attendance", "Load wages", "Review &amp; submit", "Finalize"):
        actions = step_actions(page, label)
        assert len(actions) == 1, f"{label} has {len(actions)} actions: {actions}"
    # Calculate is the one stage with two, because resetting is a distinct action.
    assert len(step_actions(page, "Calculate payroll")) == 2
    # Step 1 keeps its entry point into Attendance Manager.
    assert step_actions(page, "Import attendance")[0][0] == "Review attendance"
    # Step 2 keeps the action that does the work, not the link to master.
    assert step_actions(page, "Load wages")[0][0] == "Reload wages"
    # Step 4 owns the recalculate pair.
    assert [text for text, _ in step_actions(page, "Calculate payroll")] == [
        "Recalculate", "Reset &amp; recalculate"]


# --- Branded error pages ---

def test_unknown_url_renders_the_branded_404_page(client):
    response = client.get("/no-such-page")
    assert response.status_code == 404
    assert b"Page not found" in response.data
    assert b"error-card" in response.data
    assert b"Go to dashboard" in response.data
    # Not the default Werkzeug body.
    assert b"nabla" not in response.data.lower()


def test_malformed_month_uses_the_404_page_too(client, app):
    login(client)
    response = client.get("/payroll/2026-99")
    assert response.status_code == 404
    assert b"Page not found" in response.data
    assert b"YYYY-MM" in response.data


# --- Login hardening ---

def test_repeated_failures_lock_the_account(client, app):
    from attendance.models import User

    for attempt in range(4):
        response = client.post("/login", data={"username": "admin", "password": "wrong"}, follow_redirects=True)
        assert b"Invalid username or password" in response.data, attempt

    locked = client.post("/login", data={"username": "admin", "password": "wrong"}, follow_redirects=True)
    assert b"Too many failed attempts" in locked.data

    # The correct password is refused while the lockout stands.
    blocked = client.post("/login", data={"username": "admin", "password": "12345"}, follow_redirects=True)
    assert b"Too many failed attempts" in blocked.data
    assert b"Dashboard" not in blocked.data

    with app.app_context():
        user = User.query.filter_by(username="admin").one()
        assert user.locked_until is not None
        assert AuditLog.query.filter_by(action="User Login Locked").count() == 1
        assert AuditLog.query.filter_by(action="User Login Failed").count() == 4


def test_lockout_expires_and_a_good_password_clears_the_counter(client, app):
    from attendance.models import User

    for _ in range(5):
        client.post("/login", data={"username": "admin", "password": "wrong"})

    with app.app_context():
        user = User.query.filter_by(username="admin").one()
        user.locked_until = datetime.utcnow() - timedelta(seconds=1)
        db.session.commit()

    response = client.post("/login", data={"username": "admin", "password": "12345"}, follow_redirects=True)
    assert b"Dashboard" in response.data
    with app.app_context():
        user = User.query.filter_by(username="admin").one()
        assert user.locked_until is None
        assert user.failed_login_count == 0


def test_unknown_username_is_not_distinguishable_and_is_logged(client, app):
    response = client.post("/login", data={"username": "ghost", "password": "whatever"}, follow_redirects=True)
    assert b"Invalid username or password" in response.data
    assert b"not found" not in response.data.lower()
    with app.app_context():
        audit = AuditLog.query.filter_by(action="User Login Failed").one()
        assert audit.actor == "ghost"


def test_stale_takeover_confirmation_is_rejected(client, app):
    first = app.test_client()
    second = app.test_client()
    first.post("/login", data={"username": "admin", "password": "12345"})

    blocked = second.post("/login", data={"username": "admin", "password": "12345"})
    assert b"already signed in" in blocked.data

    # Let the window that the password check authorised go stale.
    with second.session_transaction() as sess:
        sess["pending_login_granted_at"] = datetime.utcnow().timestamp() - 3600
    stale = second.post("/login", data={"action": "force_login"}, follow_redirects=True)
    assert b"Login confirmation expired" in stale.data
    assert b"Dashboard" not in stale.data


def test_password_policy_rejects_weak_choices(client, app):
    login(client)
    for new_password, expected in [
        ("short1", b"at least 10 characters"),
        ("alllettersonly", b"one letter and one number"),
        ("12345", b"at least 10 characters"),
    ]:
        response = client.post("/settings/security", data={
            "current_password": "12345", "new_password": new_password, "confirm_password": new_password,
        }, follow_redirects=True)
        assert expected in response.data, new_password


def test_session_cookie_is_hardened(client, app):
    assert app.config["SESSION_COOKIE_HTTPONLY"] is True
    assert app.config["SESSION_COOKIE_SAMESITE"] == "Lax"
    response = client.post("/login", data={"username": "admin", "password": "12345"})
    cookie = response.headers.get("Set-Cookie", "")
    assert "HttpOnly" in cookie
    assert "SameSite=Lax" in cookie


# --- Conveyance Allowance merged into Allowance ---

def test_conveyance_value_is_folded_into_allowance_on_upgrade(tmp_path):
    """An existing breakup must still add up after the column is removed."""
    import sqlite3
    from app import create_app

    db_path = tmp_path / "legacy.db"
    connection = sqlite3.connect(db_path)
    connection.executescript(
        """
        CREATE TABLE employee (
            id VARCHAR(64) PRIMARY KEY,
            name VARCHAR(160) NOT NULL,
            department VARCHAR(160),
            designation VARCHAR(160),
            salary NUMERIC(12,2) NOT NULL DEFAULT 0,
            basic_salary NUMERIC(12,2) NOT NULL DEFAULT 0,
            hra NUMERIC(12,2) NOT NULL DEFAULT 0,
            allowance NUMERIC(12,2) NOT NULL DEFAULT 0,
            conveyance_allowance NUMERIC(12,2) NOT NULL DEFAULT 0,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        INSERT INTO employee (id, name, salary, basic_salary, hra, allowance, conveyance_allowance)
        VALUES ('1', 'Worker', 30000, 15000, 6000, 6000, 3000);
        """
    )
    connection.commit()
    connection.close()

    app = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///" + str(db_path), "TESTING": True})
    with app.app_context():
        employee = db.session.get(Employee, "1")
        assert Decimal(employee.allowance) == Decimal("9000")  # 6000 + 3000
        total = employee.basic_salary + employee.hra + employee.allowance
        assert Decimal(total) == Decimal(employee.salary)


def test_import_rejects_a_file_that_still_carries_conveyance(client, app):
    from io import BytesIO

    with app.app_context():
        db.session.add(Employee(id="1", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    response = client.post("/master/import", data={"employee_master_csv": (BytesIO(
        b"Employee ID,Name,Wage Type,Salary,Basic,HRA,Allowance,Conveyance Allowance\n"
        b"1,Worker,Monthly,30000,15000,6000,6000,3000\n"), "old.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"merged into Allowance" in response.data
    with app.app_context():
        assert Decimal(db.session.get(Employee, "1").allowance) == Decimal("0")


def test_import_accepts_the_old_basic_salary_column_name(client, app):
    from io import BytesIO

    with app.app_context():
        db.session.add(Employee(id="1", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    response = client.post("/master/import", data={"employee_master_csv": (BytesIO(
        b"Employee ID,Name,Wage Type,Salary,Basic Salary,HRA,Allowance\n"
        b"1,Worker,Monthly,30000,15000,6000,9000\n"), "old.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"Employee master imported" in response.data
    with app.app_context():
        assert Decimal(db.session.get(Employee, "1").basic_salary) == Decimal("15000")


# --- Exports carry a self-documenting sample template ---

def test_master_export_leads_with_sample_rows(client, app):
    with app.app_context():
        db.session.add(Employee(id="9", name="Real Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000"),
                                ot_ignored=True, less_hours_ignored=False))
        db.session.commit()

    login(client)
    body = client.get("/master/export.csv").data.decode()
    lines = [line for line in body.splitlines() if line.strip()]
    assert lines[0].endswith("Ignore OT,Ignore Less Hours,Ignore Monthly Bonus,Week Off Pattern,Status,Last Working Day")
    assert lines[1].startswith("EXAMPLE-MONTHLY,Example Monthly Employee,Accounts,Accounts Executive,Monthly,50000,35000,10000,5000,2500,Yes,No,Yes,No")
    assert lines[2].startswith("EXAMPLE-DAILY,Example Daily Employee,Mechanical Production,Helper,Daily,5000,0,0,0,,No,No,Yes,No")
    # The sample IDs cannot be mistaken for an employee number, so a reader never
    # reads the file as holding two rows for the same person.
    assert not any(line.split(",")[0].isdigit() for line in lines[1:3])
    # The sample breakup demonstrates the rule it documents.
    assert Decimal("35000") + Decimal("10000") + Decimal("5000") == Decimal("50000")
    # Real data follows the samples.
    assert any(line.startswith("9,Real Worker") for line in lines)
    assert "Yes,No" in [line for line in lines if line.startswith("9,")][0]


def test_exported_master_reimports_without_touching_sample_ids(client, app):
    """The illustrative rows must be skipped on the way back in, not applied."""
    with app.app_context():
        db.session.add(Employee(id="1", name="Manish C Hirani", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("68200")))
        db.session.add(Employee(id="2", name="Bijal T Patel", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("22500")))
        db.session.commit()

    login(client)
    exported = client.get("/master/export.csv").data
    assert b"Example Monthly Employee" in exported

    response = client.post("/master/import", data={
        "employee_master_csv": (BytesIO(exported), "employee_master.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"Employee master imported" in response.data
    assert b"already named" not in response.data
    with app.app_context():
        # Real records are untouched and never renamed to the sample names.
        assert db.session.get(Employee, "1").name == "Manish C Hirani"
        assert Decimal(db.session.get(Employee, "1").salary) == Decimal("68200")
        assert db.session.get(Employee, "2").name == "Bijal T Patel"


def test_master_import_reads_the_ignore_flags(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    response = client.post("/master/import", data={"employee_master_csv": (BytesIO(
        b"Employee ID,Name,Wage Type,Salary,Ignore OT,Ignore Less Hours\n"
        b"5,Worker,Monthly,30000,Yes,Yes\n"), "master.csv")},
        content_type="multipart/form-data", follow_redirects=True)
    assert b"Employee master imported" in response.data
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.ot_ignored is True
        assert employee.less_hours_ignored is True
        audit = AuditLog.query.filter_by(action="Employee Master Bulk Updated").one()
        assert "Ignore OT No -> Yes" in audit.detail


def test_leave_balance_export_and_reimport_skip_samples(client, app):
    with app.app_context():
        db.session.add(Employee(id="1", name="Manish C Hirani", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    exported = client.get("/leave-balances/export.csv").data
    lines = [line for line in exported.decode().splitlines() if line.strip()]
    assert lines[1].startswith("EXAMPLE-MONTHLY,Example Monthly Employee,12.5")
    assert lines[2].startswith("EXAMPLE-DAILY,Example Daily Employee,0")

    response = client.post("/leave-balances/import", data={
        "leave_balance_csv": (BytesIO(exported), "leave_balances.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"not found" not in response.data
    with app.app_context():
        # The illustrative rows must not become employees.
        assert db.session.get(Employee, "EXAMPLE-DAILY") is None
        assert db.session.get(Employee, "EXAMPLE-MONTHLY") is None


# --- Dashboard analytics ---

def test_dashboard_shows_analytics_sections(client, app):
    with app.app_context():
        seed_mixed_month()
        employee = db.session.get(Employee, "5")
        employee.department = "Accounts"
        db.session.commit()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/").data
    for marker in (b"Payable trend", b"Cost by department", b"Deductions",
                   b"Attendance clean", b"Leave liability", b"trend-chart", b"share-track"):
        assert marker in page, marker
    assert b"Accounts" in page


def test_dashboard_compares_against_the_previous_month(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")
        # An earlier, cheaper month to compare against.
        db.session.add(PayrollMonth(month="2026-06"))
        db.session.add(SalaryRecord(payroll_month="2026-06", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("10000")))
        db.session.add(PayrollResult(payroll_month="2026-06", employee_id="5", payroll_rule_type="MONTHLY",
                                     calculation_status="Calculated", final_salary=Decimal("10000"),
                                     total_deduction=Decimal("0")))
        db.session.commit()

    login(client)
    page = client.get("/?month=2026-07").data
    assert b"Compared with June 2026" in page
    assert b"delta-" in page


def test_dashboard_handles_a_month_with_no_data(client, app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.commit()

    login(client)
    response = client.get("/")
    assert response.status_code == 200
    assert b"No calculated months yet" in response.data or b"trend-chart" in response.data
    assert b"Load wages to see department costs" in response.data
    assert b"No deductions this month" in response.data


def test_report_totals_match_the_payroll_figures_not_rounded_row_sums(app):
    """KPI totals must equal what payroll deducted, not the sum of displayed rows.

    Each day's shortage is rounded to 2dp for display; adding those up drifts from
    the once-rounded total that is actually deducted.
    """
    from attendance.reports import build_less_hours_report_pdf

    with app.app_context():
        seed_month()
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("30000")))
        # A salary that makes the quarter-hour rate a repeating decimal, so per-day
        # rounding and once-rounding genuinely differ across several days.
        for day in (1, 2, 3, 6, 7, 8, 9, 10):
            db.session.add(AttendanceRecord(
                payroll_month="2026-07", employee_id="5", employee_name="Worker",
                date=date(2026, 7, day), day="Weekday", first_punch="09:30 AM",
                last_punch="05:15 PM", raw_working_hours="7h 45m",
                actual_minutes=465, parse_status="OK"))
        db.session.commit()
        calculate_payroll_month("2026-07")

        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        stored_total = Decimal(result.less_hours_deduction)
        per_day_sum = sum(
            (Decimal(str(item.get("shortage_deduction") or "0")) for item in result.detail_json or []),
            Decimal("0"),
        )

        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(build_less_hours_report_pdf("2026-07"))).pages)
        assert f"{stored_total:,.2f}" in text, f"report should show the payroll total {stored_total}"
        if per_day_sum != stored_total:
            assert f"{per_day_sum:,.2f}" not in text, "report must not show the drifted row sum"


# --- Employment status dropdown ---

def test_status_dropdown_offers_all_four_states(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    for path in ("/master", "/master/5"):
        page = client.get(path).data
        assert b'name="employment_status"' in page, path
        for value in (b'value="ACTIVE"', b'value="INACTIVE"', b'value="LEFT"', b'value="TERMINATED"'):
            assert value in page, f"{value} missing from {path}"


def test_status_change_saves_and_is_audited(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    client.post("/master", data={
        "employee_id": "5", "name": "Worker", "wage_type": "Monthly", "salary": "30000",
        "master_controls_present": "1", "employment_status": "INACTIVE",
    }, follow_redirects=True)
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.employment_status == "INACTIVE"
        assert employee.inactive_at is not None
        audit = AuditLog.query.filter_by(action="Employee Master Updated").first()
        assert "Status ACTIVE -> INACTIVE" in audit.detail


def test_status_only_save_works_when_other_fields_are_disabled(client, app):
    """A non-active employee posts only the status; the rest must not be wiped."""
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", department="Design", designation="Engineer",
                                salary_type="Monthly", normalized_salary_type="MONTHLY",
                                salary=Decimal("30000"), employment_status="LEFT"))
        db.session.commit()

    login(client)
    response = client.post("/master/5", data={
        "employee_id": "5", "master_controls_present": "1", "employment_status": "ACTIVE",
    }, follow_redirects=True)
    assert b"Employee master updated" in response.data
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.employment_status == "ACTIVE"
        assert employee.inactive_at is None
        # Stored values survived a status-only submit.
        assert employee.name == "Worker"
        assert Decimal(employee.salary) == Decimal("30000")
        assert employee.normalized_salary_type == "MONTHLY"


def test_inactive_employee_is_excluded_from_payroll(client, app):
    with app.app_context():
        seed_mixed_month()
        calculate_payroll_month("2026-07")
        assert PayrollResult.query.filter_by(payroll_month="2026-07").count() == 2

        db.session.get(Employee, "5").employment_status = "INACTIVE"
        db.session.commit()
        calculate_payroll_month("2026-07")
        remaining = [r.employee_id for r in PayrollResult.query.filter_by(payroll_month="2026-07").all()]
        assert "5" not in remaining


def test_unknown_status_is_rejected(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    response = client.post("/master", data={
        "employee_id": "5", "name": "Worker", "wage_type": "Monthly", "salary": "30000",
        "master_controls_present": "1", "employment_status": "RETIRED",
    }, follow_redirects=True)
    assert b"Unknown employment status" in response.data
    with app.app_context():
        assert db.session.get(Employee, "5").employment_status == "ACTIVE"


# --- Sticky table columns must be opaque or content scrolls through them ---

def test_sticky_columns_use_opaque_backgrounds():
    css = pathlib.Path("static/css/app.css").read_text()
    for token in ("--sticky-head-bg", "--sticky-hover-bg"):
        assert token in css
    # The translucent tinted fills must not be used on a sticky cell.
    sticky_rules = [
        ".sticky-id-name-table thead th:first-child",
        ".attendance-manager-table thead .sticky-col",
    ]
    for rule in sticky_rules:
        index = css.index(rule)
        block = css[index:index + 260]
        assert "--smartfill-table-head" not in block, rule


# --- Absences are settled against the leave balance automatically ---

def seed_leave_month(month="2026-07", opening=Decimal("2")):
    """One monthly employee with a stated opening leave balance."""
    from attendance.models import LeaveLedger
    db.session.add(PayrollMonth(month=month))
    db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                            normalized_salary_type="MONTHLY", salary=Decimal("31000")))
    db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
    db.session.add(SalaryRecord(payroll_month=month, employee_id="5", name="Worker",
                                salary_type="Monthly", normalized_salary_type="MONTHLY",
                                salary=Decimal("31000")))
    # Opening balance comes from a prior finalized month's carry-forward.
    db.session.add(PayrollResult(payroll_month="2026-06", employee_id="5",
                                 payroll_rule_type="MONTHLY", calculation_status="Calculated",
                                 closing_leave=opening, final_salary=Decimal("31000")))
    db.session.commit()


def add_july_attendance(absent_days):
    """Every July day punched 09:30-18:30 except Sundays and `absent_days`."""
    for day in range(1, 32):
        when = date(2026, 7, day)
        punched = when.weekday() != 6 and day not in absent_days
        db.session.add(AttendanceRecord(
            payroll_month="2026-07", employee_id="5", employee_name="Worker",
            date=when, day=when.strftime("%A"),
            first_punch="09:30 AM" if punched else "",
            last_punch="06:30 PM" if punched else "",
            raw_working_hours="9h 00m" if punched else "",
            actual_minutes=540 if punched else None,
            parse_status="OK" if punched else "NEEDS_REVIEW",
            warning="" if punched else "Missing punch and working hours",
        ))
    db.session.commit()


def test_absences_consume_leave_then_fall_to_lop(app):
    """The worked example: opening 2, absent on 3, 10, 14 and 22."""
    with app.app_context():
        seed_leave_month(opening=Decimal("2"))
        add_july_attendance({3, 10, 14, 22})
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()

        by_date = {row["date"]: row for row in result.detail_json}
        # Opening balance of 2 covers the first two absences.
        assert by_date["2026-07-03"]["attendance_status"] == "Paid Leave"
        assert by_date["2026-07-10"]["attendance_status"] == "Paid Leave"
        # Earned leave then covers the third in full ...
        assert by_date["2026-07-14"]["attendance_status"] == "Paid Leave"
        # ... and the fourth by half, with the other half unpaid.
        assert by_date["2026-07-22"]["attendance_status"] == "Half-Day Paid Leave / Half-Day LOP"

        # The accrual is pro-rated by the days that ended up paid - 30.5 of 31, since
        # half of 22 July went unpaid - rather than by the count before the leave it
        # granted was spent. Two decimals keep what a single decimal truncated away.
        assert Decimal(result.leave_earned) == Decimal("1.96")
        assert Decimal(result.leave_used) == Decimal("3.5")
        assert Decimal(result.closing_leave) == Decimal("0.46")
        assert Decimal(result.lop_days) == Decimal("0.5")


def test_exact_attendance_threshold_gets_full_monthly_leave(app):
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance({3, 10})
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row for row in result.detail_json}
        # 25 attended days + 4 week offs = 29 attendance-credit days. In a 31-day
        # month that reaches days_in_month - 2, so the month earns the full 2 leaves.
        assert by_date["2026-07-03"]["attendance_status"] == "Paid Leave"
        assert by_date["2026-07-10"]["attendance_status"] == "Paid Leave"
        assert Decimal(result.leave_earned) == Decimal("2")
        assert Decimal(result.leave_used) == Decimal("2")
        assert Decimal(result.lop_days) == Decimal("0")


def test_no_punch_day_is_an_absence_not_a_review_item(app):
    """167 no-punch days used to sit in Needs Review, neither paid nor deducted."""
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance({3})
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        statuses = {row["attendance_status"] for row in result.detail_json}
        assert "Needs Review" not in statuses
        assert result.calculation_status == "Calculated"


def test_punch_error_still_needs_review(app):
    """A genuine punch anomaly must not be silently absorbed as an absence."""
    with app.app_context():
        seed_leave_month(opening=Decimal("5"))
        add_july_attendance(set())
        record = AttendanceRecord.query.filter_by(payroll_month="2026-07", employee_id="5", date=date(2026, 7, 2)).one()
        record.first_punch = "09:30 AM"
        record.last_punch = ""
        record.raw_working_hours = ""
        record.actual_minutes = None
        record.parse_status = "NEEDS_REVIEW"
        record.warning = "Odd punch count"
        db.session.commit()

        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row for row in result.detail_json}
        assert by_date["2026-07-02"]["attendance_status"] == "Needs Review"
        assert result.calculation_status == "Needs Review"


# --- Leave balance only moves once the month is finalized ---

def test_draft_month_does_not_move_the_leave_balance(client, app):
    from attendance.leave_balances import stored_leave_balance

    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(PayrollMonth(month="2026-06", status="FINALIZED"))
        db.session.add(PayrollResult(payroll_month="2026-06", employee_id="5",
                                     payroll_rule_type="MONTHLY", calculation_status="Calculated",
                                     closing_leave=Decimal("3.0"), final_salary=Decimal("30000")))
        db.session.add(PayrollMonth(month="2026-07", status="DRAFT"))
        db.session.add(PayrollResult(payroll_month="2026-07", employee_id="5",
                                     payroll_rule_type="MONTHLY", calculation_status="Calculated",
                                     closing_leave=Decimal("4.8"), final_salary=Decimal("30000")))
        db.session.commit()

        # July is still a draft, so the balance stays at June's carry-forward.
        assert stored_leave_balance("5") == Decimal("3.0")

        db.session.get(PayrollMonth, "2026-07").status = "FINALIZED"
        db.session.commit()
        assert stored_leave_balance("5") == Decimal("4.8")


def test_leave_balance_page_flags_the_pending_month(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(PayrollMonth(month="2026-07", status="DRAFT"))
        db.session.add(PayrollResult(payroll_month="2026-07", employee_id="5",
                                     payroll_rule_type="MONTHLY", calculation_status="Calculated",
                                     closing_leave=Decimal("1.8"), final_salary=Decimal("30000")))
        db.session.commit()

    login(client)
    page = client.get("/leave-balances").data
    assert b"is not finalized" in page
    assert b"pending" in page


# --- Daily wage no-punch days ---

def test_daily_wage_missing_punch_is_absent(app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="6", name="Day Worker", salary_type="Daily",
                                normalized_salary_type="DAILY", salary=Decimal("600")))
        db.session.add(WeekOffRule(employee_id="6", confirmed_at=datetime.utcnow()))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="6", name="Day Worker",
                                    salary_type="Daily", normalized_salary_type="DAILY", salary=Decimal("600")))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="6", employee_name="Day Worker",
                                        date=date(2026, 7, 1), day="Wednesday", parse_status="NEEDS_REVIEW",
                                        warning="Missing punch and working hours"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="6", employee_name="Day Worker",
                                        date=date(2026, 7, 2), day="Thursday", first_punch="09:30 AM",
                                        last_punch="06:30 PM", raw_working_hours="9h 00m",
                                        actual_minutes=540, parse_status="OK"))
        db.session.commit()

        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        by_date = {row["date"]: row["attendance_status"] for row in result.detail_json}
        assert by_date["2026-07-01"] == "Absent / Attendance Missing"
        assert by_date["2026-07-02"] == "Full Day Present"
        # Daily wage has no leave, so the absent day is simply not paid.
        assert result.calculation_status == "Calculated"
        assert Decimal(result.paid_working_days) == Decimal("1")


def test_error_report_excludes_week_off_attendance_warnings(app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 5), day="Sunday", parse_status="NEEDS_REVIEW",
                                        warning="Missing punch and working hours"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 6), day="Monday", parse_status="NEEDS_REVIEW",
                                        warning="Missing punch and working hours"))
        db.session.commit()

        csv_text = error_report_csv("2026-07")
        assert "2026-07-05" not in csv_text
        assert "2026-07-06" in csv_text

        pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(build_error_report_pdf("2026-07"))).pages)
        assert "2026-07-05" not in pdf_text
        assert "2026-07-06" in pdf_text


def test_error_report_aggregates_rows_by_issue_priority(app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        for employee_id, name in (("5", "Missing Staff"), ("6", "Odd Staff"), ("7", "Other Staff")):
            db.session.add(Employee(id=employee_id, name=name, salary_type="Monthly",
                                    normalized_salary_type="MONTHLY", salary=Decimal("30000")))
            db.session.add(WeekOffRule(employee_id=employee_id, confirmed_at=datetime.utcnow()))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Missing Staff",
                                        date=date(2026, 7, 6), day="Monday", parse_status="NEEDS_REVIEW",
                                        warning="Missing punch and working hours"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="6", employee_name="Odd Staff",
                                        date=date(2026, 7, 3), day="Friday", first_punch="09:30 AM",
                                        parse_status="NEEDS_REVIEW", warning="Odd punch count"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="6", employee_name="Odd Staff",
                                        date=date(2026, 7, 4), day="Saturday", first_punch="06:30 PM",
                                        parse_status="NEEDS_REVIEW", warning="Odd punch count"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="7", employee_name="Other Staff",
                                        date=date(2026, 7, 7), day="Tuesday",
                                        parse_status="NEEDS_REVIEW",
                                        warning="Punch out before punch in (18h 00m session)"))
        db.session.commit()

        rows = list(csv.DictReader(StringIO(error_report_csv("2026-07"))))
        assert [row["Issue"] for row in rows] == [
            "Odd punch count",
            "Odd punch count",
            "Missing punch and working hours",
            "Punch out before punch in (18h 00m session)",
        ]
        assert [row["Issue Count"] for row in rows] == ["2", "2", "1", "1"]

        pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(build_error_report_pdf("2026-07"))).pages)
        assert pdf_text.index("Odd punch count") < pdf_text.index("Missing punch and working hours")
        assert "Punch out before punch in" in pdf_text


def test_manual_override_report_compares_override_with_imported_attendance(app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("30000")))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Worker",
                                        date=date(2026, 7, 6), day="Monday",
                                        first_punch="09:30 AM", last_punch="06:30 PM",
                                        raw_working_hours="9h 00m", actual_minutes=540,
                                        parse_status="OK"))
        db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="5",
                                          date=date(2026, 7, 6), manual_status="Paid Leave",
                                          notes="Approved by manager"))
        db.session.commit()

        csv_text = manual_override_report_csv("2026-07")
        assert "Worker" in csv_text
        assert "Full Day Present" in csv_text
        assert "Paid Leave" in csv_text
        assert "Approved by manager" in csv_text

        pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(build_manual_override_report_pdf("2026-07"))).pages)
        assert "Manual Override Report" in pdf_text
        assert "Full Day Present" in pdf_text
        assert "Paid Leave" in pdf_text
        assert "Approved by manager" in pdf_text


# --- Day notes only appear when they explain an exception ---

def test_present_days_carry_no_explanatory_note(client, app):
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance({3})
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data
    assert b"Actual duration meets" not in page
    # A day already resolved into leave or LOP no longer shows the raw import warning.
    assert b"Missing punch and working hours" not in page


# --- Bulk import can add new employees ---

def import_master(client, body):
    from io import BytesIO
    return client.post("/master/import", data={
        "employee_master_csv": (BytesIO(body), "master.csv")
    }, content_type="multipart/form-data", follow_redirects=True)


def test_import_adds_new_employees_alongside_existing_ones(client, app):
    with app.app_context():
        db.session.add(Employee(id="1", name="Manish C Hirani", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("68200")))
        db.session.add(Employee(id="2", name="Bijal T Patel", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("22500")))
        db.session.commit()

    login(client)
    response = import_master(client, (
        b"Employee ID,Name,Department,Designation,Wage Type,Salary\n"
        b"1,Manish C Hirani,Design,Design Manager,Monthly,70000\n"
        b"2,Bijal T Patel,Purchase,Purchase Engineer,Monthly,22500\n"
        b"3,New Starter,Service,Technician,Monthly,31000\n"
        b"4,Day Starter,Production,Helper,Daily,650\n"
    ))
    assert b"2 employee(s) added, 2 updated" in response.data

    with app.app_context():
        new_monthly = db.session.get(Employee, "3")
        assert new_monthly.name == "New Starter"
        assert new_monthly.normalized_salary_type == "MONTHLY"
        assert Decimal(new_monthly.salary) == Decimal("31000")
        assert new_monthly.department == "Service"
        assert new_monthly.employment_status == "ACTIVE"
        # A new employee needs the same defaults the Add Employee form creates.
        assert WeekOffRule.query.filter_by(employee_id="3").count() == 1
        assert LeaveLedger.query.filter_by(employee_id="3", transaction_type="OPENING").count() == 1

        assert db.session.get(Employee, "4").normalized_salary_type == "DAILY"
        # The existing employee was updated, not duplicated.
        assert Decimal(db.session.get(Employee, "1").salary) == Decimal("70000")
        assert Employee.query.count() == 4
        assert AuditLog.query.filter_by(action="Employee Master Bulk Created").count() == 2


def test_new_employee_needs_a_name_and_wage_type(client, app):
    login(client)
    missing_name = import_master(client, b"Employee ID,Name,Wage Type,Salary\n9,,Monthly,30000\n")
    assert b"Name is required to add new Employee ID 9" in missing_name.data

    missing_type = import_master(client, b"Employee ID,Name,Wage Type,Salary\n9,New Person,,30000\n")
    assert b"Wage type is required to add new Employee ID 9" in missing_type.data

    with app.app_context():
        assert db.session.get(Employee, "9") is None


def test_existing_employee_still_cannot_be_renamed_by_import(client, app):
    with app.app_context():
        db.session.add(Employee(id="1", name="Manish C Hirani", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("68200")))
        db.session.commit()

    login(client)
    blocked = import_master(client, (
        b"Employee ID,Name,Wage Type,Salary\n"
        b"1,Manish Hirani,Monthly,68200\n"
        b"3,Brand New,Monthly,25000\n"
    ))
    assert b"already named" in blocked.data
    with app.app_context():
        assert db.session.get(Employee, "1").name == "Manish C Hirani"
        # The whole import is rejected, so the new employee is not created either.
        assert db.session.get(Employee, "3") is None


def test_new_employee_breakup_must_reconcile(client, app):
    login(client)
    bad = import_master(client, (
        b"Employee ID,Name,Wage Type,Salary,Basic,HRA,Allowance\n"
        b"7,Breakup Starter,Monthly,30000,15000,6000,5000\n"
    ))
    assert b"must add up to the salary" in bad.data
    with app.app_context():
        assert db.session.get(Employee, "7") is None

    good = import_master(client, (
        b"Employee ID,Name,Wage Type,Salary,Basic,HRA,Allowance,TDS,PF,ESIC,Ignore OT,Ignore Less Hours\n"
        b"7,Breakup Starter,Monthly,30000,15000,6000,9000,1200,Yes,No,Yes,No\n"
    ))
    assert b"1 employee(s) added" in good.data
    with app.app_context():
        employee = db.session.get(Employee, "7")
        assert Decimal(employee.basic_salary) == Decimal("15000")
        assert employee.pf_enabled is True
        assert employee.ot_ignored is True
        assert employee.less_hours_ignored is False


# --- A split day must not read as one continuous session ---

def test_split_day_shows_each_punch_pair():
    from attendance.reports import punch_sessions

    # Real case: in 09:34, out 10:55, back 17:43, out 18:30 is 2h 08m, not 8h 56m.
    assert punch_sessions(["09:34 AM", "10:55 AM", "05:43 PM", "06:30 PM"]) == [
        "09:34 AM - 10:55 AM", "05:43 PM - 06:30 PM",
    ]
    assert punch_sessions(["09:30 AM", "06:30 PM"]) == ["09:30 AM - 06:30 PM"]
    # An odd punch count is shown as incomplete rather than silently paired.
    assert punch_sessions(["09:30 AM"]) == ["09:30 AM - ?"]
    assert punch_sessions([]) == []
    # Falls back to first/last when no punch list was stored.
    assert punch_sessions([], "09:30 AM", "06:30 PM") == ["09:30 AM - 06:30 PM"]


def test_wage_sync_reports_which_employees_were_skipped(client, app):
    from attendance.master import sync_salary_records_from_master

    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="1", name="Paid Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(Employee(id="2", name="Zero Salary", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("0")))
        db.session.add(Employee(id="3", name="No Wage Type", salary=Decimal("500")))
        db.session.add(Employee(id="4", name="Gone", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("20000"),
                                employment_status="LEFT"))
        db.session.commit()

        created, updated, skipped = sync_salary_records_from_master("2026-07", "admin")
        db.session.commit()
        assert created == 1 and updated == 0
        assert len(skipped) == 3
        joined = " | ".join(skipped)
        assert "2 - Zero Salary: salary is zero" in joined
        assert "3 - No Wage Type: no wage type set" in joined
        assert "4 - Gone: status is Left" in joined
        audit = AuditLog.query.filter_by(action="Wage Master Loaded").one()
        assert "Skipped:" in audit.detail

    login(client)
    response = client.post("/payroll/2026-07", data={"action": "salary"}, follow_redirects=True)
    # The count alone hid who was missing from payroll; each one is now named.
    assert b"3 skipped" in response.data
    assert b"Zero Salary: salary is zero" in response.data
    assert b"Gone: status is Left" in response.data


# --- "Worked On-Site" override marks a full present day ---

def seed_override_month(status_day, override_status):
    """One monthly employee, one working day carrying the given override."""
    from attendance.models import AttendanceOverride
    db.session.add(PayrollMonth(month="2026-07"))
    db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                            normalized_salary_type="MONTHLY", salary=Decimal("31000")))
    db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
    db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                salary_type="Monthly", normalized_salary_type="MONTHLY",
                                salary=Decimal("31000")))
    for day in range(1, 32):
        when = date(2026, 7, day)
        punched = when.weekday() != 6 and day != status_day
        db.session.add(AttendanceRecord(
            payroll_month="2026-07", employee_id="5", employee_name="Worker",
            date=when, day=when.strftime("%A"),
            first_punch="09:30 AM" if punched else "",
            last_punch="06:30 PM" if punched else "",
            raw_working_hours="9h 00m" if punched else "",
            actual_minutes=540 if punched else None,
            parse_status="OK" if punched else "NEEDS_REVIEW",
        ))
    db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="5",
                                      date=date(2026, 7, status_day), manual_status=override_status))
    db.session.commit()


def test_worked_on_site_counts_as_a_full_present_day(app):
    with app.app_context():
        seed_override_month(3, "Worked On-Site")
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row for row in result.detail_json}
        assert by_date["2026-07-03"]["attendance_status"] == "Worked On-Site"
        assert by_date["2026-07-03"]["paid_day_value"] == "1"
        # Paid, not deducted, and not consuming leave.
        assert Decimal(result.lop_days) == Decimal("0")
        assert Decimal(result.leave_used) == Decimal("0")
        assert Decimal(result.paid_working_days) == Decimal("27")


def test_work_from_home_override_is_also_paid(app):
    """It was offered as an override but never handled, so the day vanished."""
    with app.app_context():
        seed_override_month(3, "Work From Home")
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(result.paid_working_days) == Decimal("27")
        assert Decimal(result.lop_days) == Decimal("0")


def test_worked_on_site_is_offered_and_shown(client, app):
    with app.app_context():
        seed_override_month(3, "Worked On-Site")
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data
    assert b"Worked On-Site" in page
    assert b"tone-offsite" in page

    # The status shows on the attendance calendar, which is on the summary report.
    pdf = client.get("/reports/2026-07/attendance-summary-monthly.pdf")
    assert pdf.status_code == 200
    from io import BytesIO
    text = "\n".join(p.extract_text() or "" for p in PdfReader(BytesIO(pdf.data)).pages)
    assert "Worked On-Site" in text


# --- Daily wage attendance bonus (notice of 08/12/2023) ---

def seed_daily_bonus_month(minutes_by_day, month="2026-07", rate="600"):
    """A daily wage employee with one attendance row per entry in minutes_by_day.

    July 2026 starts on a Wednesday; Sundays are the default week off. Passing None
    for a day's minutes means no punches at all, i.e. a full absence.
    """
    db.session.add(PayrollMonth(month=month))
    db.session.add(Employee(id="6", name="Day Worker", salary_type="Daily",
                            normalized_salary_type="DAILY", salary=Decimal(rate)))
    db.session.add(WeekOffRule(employee_id="6", confirmed_at=datetime.utcnow()))
    db.session.add(SalaryRecord(payroll_month=month, employee_id="6", name="Day Worker",
                                salary_type="Daily", normalized_salary_type="DAILY", salary=Decimal(rate)))
    for day, minutes in sorted(minutes_by_day.items()):
        when = date(2026, 7, day)
        if minutes is None:
            db.session.add(AttendanceRecord(payroll_month=month, employee_id="6", employee_name="Day Worker",
                                            date=when, day=when.strftime("%A"), parse_status="NEEDS_REVIEW",
                                            warning="Missing punch and working hours"))
            continue
        db.session.add(AttendanceRecord(payroll_month=month, employee_id="6", employee_name="Day Worker",
                                        date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                        last_punch="06:30 PM", raw_working_hours=f"{minutes // 60}h {minutes % 60:02d}m",
                                        actual_minutes=minutes, parse_status="OK"))
    db.session.commit()


def daily_bonus_result(minutes_by_day, month="2026-07", rate="600"):
    seed_daily_bonus_month(minutes_by_day, month=month, rate=rate)
    calculate_payroll_month(month)
    return PayrollResult.query.filter_by(payroll_month=month, employee_id="6").one()


def test_full_attendance_earns_the_ten_percent_bonus(app):
    with app.app_context():
        # 1-4 July 2026 are Wed-Sat, all worked to the full nine hours.
        result = daily_bonus_result({1: 540, 2: 540, 3: 540, 4: 540})
        assert result.absence_minutes == 0
        assert Decimal(result.attendance_bonus_percent) == Decimal("10")
        # Four days at 600 is 2400 earned, so the bonus is 240.
        assert Decimal(result.attendance_bonus_amount) == Decimal("240.00")
        assert Decimal(result.final_salary) == Decimal("2640.00")


def test_days_inside_the_full_day_grace_carry_no_absence(app):
    """8h55m is a full day with no short hours, so it is not absence either."""
    with app.app_context():
        result = daily_bonus_result({1: 535, 2: 533, 3: 540, 4: 545})
        assert result.absence_minutes == 0
        assert Decimal(result.paid_working_days) == Decimal("4")
        assert Decimal(result.less_hours_minutes) == 0
        assert Decimal(result.attendance_bonus_percent) == Decimal("10")


def test_short_day_below_the_grace_counts_as_absence(app):
    with app.app_context():
        # 8h 00m is below the 8h 50m grace, so the whole hour short counts.
        result = daily_bonus_result({1: 480, 2: 540, 3: 540, 4: 540})
        assert result.absence_minutes == 60
        assert Decimal(result.attendance_bonus_percent) == Decimal("5")


def test_absence_within_the_allowance_earns_five_percent(app):
    with app.app_context():
        # One full absence is nine hours; 2h 30m short on another makes 11h 30m.
        result = daily_bonus_result({1: None, 2: 390, 3: 540, 4: 540})
        assert result.absence_minutes == 540 + 150
        assert Decimal(result.attendance_bonus_percent) == Decimal("5")


def test_absence_beyond_the_allowance_earns_nothing(app):
    with app.app_context():
        result = daily_bonus_result({1: None, 2: None, 3: 540, 4: 540})
        assert result.absence_minutes == 1080
        assert Decimal(result.attendance_bonus_percent) == Decimal("0")
        assert Decimal(result.attendance_bonus_amount) == Decimal("0.00")
        # Two absent days out of four, so only two days are payable, with no bonus.
        assert Decimal(result.final_salary) == Decimal("1200.00")


def test_absence_allowance_boundary_is_inclusive(app):
    with app.app_context():
        # Exactly one and a half working days: one absent day plus 4h 30m elsewhere.
        result = daily_bonus_result({1: None, 2: 270, 3: 540, 4: 540})
        assert result.absence_minutes == 810
        assert Decimal(result.attendance_bonus_percent) == Decimal("5")


def test_one_minute_past_the_allowance_earns_nothing(app):
    with app.app_context():
        result = daily_bonus_result({1: None, 2: 269, 3: 540, 4: 540})
        assert result.absence_minutes == 811
        assert Decimal(result.attendance_bonus_percent) == Decimal("0")


def test_week_offs_and_holidays_are_not_absence(app):
    with app.app_context():
        db.session.add(Holiday(date=date(2026, 7, 3), name="Test Holiday"))
        db.session.commit()
        # 5 July 2026 is a Sunday, the default week off; 3 July is now a holiday.
        result = daily_bonus_result({1: 540, 2: 540, 3: 0, 4: 540, 5: 0})
        assert result.holidays == 1
        assert result.week_offs == 1
        assert result.absence_minutes == 0
        assert Decimal(result.attendance_bonus_percent) == Decimal("10")


def test_worked_on_site_override_creates_no_absence(app):
    with app.app_context():
        seed_daily_bonus_month({1: 540, 2: 540, 3: None, 4: 540})
        absent = AttendanceRecord.query.filter_by(employee_id="6", date=date(2026, 7, 3)).one()
        db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="6", date=absent.date,
                                          manual_status="Worked On-Site"))
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        assert result.absence_minutes == 0
        assert Decimal(result.attendance_bonus_percent) == Decimal("10")


def test_monthly_wage_earns_no_attendance_bonus(app):
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(result.attendance_bonus_percent or 0) == Decimal("0")
        assert Decimal(result.attendance_bonus_amount or 0) == Decimal("0")


def test_attendance_bonus_is_shown_on_the_page_and_the_pdf(client, app):
    with app.app_context():
        daily_bonus_result({1: 540, 2: 540, 3: 540, 4: 540})

    login(client)
    page = client.get("/payroll/2026-07/employee/6").data
    assert b"Attendance bonus" in page
    assert b"Absence this month" in page

    pdf = client.get("/reports/2026-07/employee/6.pdf")
    assert pdf.status_code == 200
    text = "\n".join(p.extract_text() or "" for p in PdfReader(BytesIO(pdf.data)).pages)
    # The daily sheet states the band only, with no amount anywhere.
    assert "Bonus" in text
    assert "10%" in text
    assert "of earned wage" not in text
    assert "Absence This Month" in text


# --- Split punch days under 3 hours need a human, not automatic LOP ---

def add_punch_day(day, punches, minutes, month="2026-07", employee_id="5"):
    when = date(2026, 7, day)
    db.session.add(AttendanceRecord(
        payroll_month=month, employee_id=employee_id, employee_name="Worker", date=when,
        day=when.strftime("%A"), punches_json=punches, first_punch=punches[0] if punches else "",
        last_punch=punches[-1] if punches else "", raw_working_hours=f"{minutes // 60}h {minutes % 60:02d}m",
        actual_minutes=minutes, parse_status="OK"))


def test_split_punches_under_three_hours_need_review_instead_of_lop(app):
    """Rakesh's 21 July: 09:34-10:55 plus 17:43-18:30 is 2h 08m across two pairs."""
    with app.app_context():
        seed_month()
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("30000")))
        add_punch_day(1, ["09:34 AM", "10:55 AM", "05:43 PM", "06:30 PM"], 128)
        # A single short pair is a genuinely short day and stays loss of pay.
        add_punch_day(2, ["09:34 AM", "11:42 AM"], 128)
        db.session.commit()

        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row for row in result.detail_json}
        assert by_date["2026-07-01"]["attendance_status"] == "Needs Review"
        assert "Multiple punch pairs" in by_date["2026-07-01"]["explanation"]
        assert by_date["2026-07-02"]["attendance_status"] == "Full Day LOP"
        # A day awaiting review is not silently deducted as loss of pay.
        assert Decimal(result.lop_days) == Decimal("1")
        assert result.calculation_status == "Needs Review"


def test_split_punches_reaching_half_day_are_still_paid(app):
    """The review rule only catches days that would otherwise be unpayable."""
    with app.app_context():
        seed_month()
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("30000")))
        add_punch_day(1, ["09:30 AM", "12:30 PM", "02:00 PM", "03:00 PM"], 240)
        db.session.commit()

        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row["attendance_status"] for row in result.detail_json}
        assert by_date["2026-07-01"] == "Half Day Present"


def test_daily_wage_split_punches_under_three_hours_need_review(app):
    with app.app_context():
        seed_daily_bonus_month({})
        add_punch_day(1, ["09:34 AM", "10:55 AM", "05:43 PM", "06:30 PM"], 128, employee_id="6")
        add_punch_day(2, ["09:34 AM", "11:42 AM"], 128, employee_id="6")
        db.session.commit()

        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        by_date = {row["date"]: row["attendance_status"] for row in result.detail_json}
        assert by_date["2026-07-01"] == "Needs Review"
        assert by_date["2026-07-02"] == "Absent / Attendance Missing"


def test_review_reason_is_shown_on_the_employee_page(client, app):
    with app.app_context():
        seed_month()
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("30000")))
        add_punch_day(1, ["09:34 AM", "10:55 AM", "05:43 PM", "06:30 PM"], 128)
        db.session.commit()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data
    assert b"Needs Review" in page
    # The day parses cleanly, so the rule's own explanation is the only reason text.
    assert b"Multiple punch pairs" in page


# --- Per-employee opt-out from the daily wage attendance bonus ---

def test_ignoring_the_bonus_removes_it_from_a_perfect_month(app):
    with app.app_context():
        seed_daily_bonus_month({1: 540, 2: 540, 3: 540, 4: 540})
        db.session.get(Employee, "6").bonus_ignored = True
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        # Absence is still measured, so the page can show what the month would have earned.
        assert result.absence_minutes == 0
        assert Decimal(result.attendance_bonus_percent) == Decimal("0")
        assert Decimal(result.attendance_bonus_amount) == Decimal("0.00")
        assert Decimal(result.final_salary) == Decimal("2400.00")


def test_bonus_opt_out_is_daily_only_on_the_employee_form(client, app):
    login(client)
    for wage_type, expected in (("Daily", True), ("Monthly", False)):
        client.post("/master", data={
            "employee_id": "9", "name": "Opt Out", "wage_type": wage_type, "salary": "500",
            "master_controls_present": "1", "bonus_ignored": "on",
        }, follow_redirects=True)
        with app.app_context():
            assert bool(db.session.get(Employee, "9").bonus_ignored) is expected
            # A monthly record must not be left carrying a daily-only flag.
            db.session.query(Employee).filter_by(id="9").delete()
            db.session.commit()


def test_bonus_flag_round_trips_through_export_and_import(client, app):
    with app.app_context():
        db.session.add(Employee(id="6", name="Day Worker", salary_type="Daily",
                                normalized_salary_type="DAILY", salary=Decimal("600"),
                                bonus_ignored=True))
        db.session.add(Employee(id="5", name="Month Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    export = client.get("/master/export.csv").data.decode()
    assert "Ignore Monthly Bonus" in export.splitlines()[0]
    rows = {line.split(",")[0]: line for line in export.splitlines()}
    # The bonus flag sits just before the week off pattern and the status columns.
    assert ",Yes,," in rows["6"] and rows["6"].endswith("ACTIVE,")
    # Monthly employees leave the column blank, the same way daily leaves Basic blank.
    assert ",No,,,ACTIVE," in rows["5"] or ",,,ACTIVE," in rows["5"]

    # Re-importing an untouched export is a no-op, not an error.
    assert b"imported" in import_master(client, export.encode()).data
    with app.app_context():
        assert db.session.get(Employee, "6").bonus_ignored is True

    import_master(client, (
        "Employee ID,Name,Wage Type,Salary,Ignore Monthly Bonus\n"
        "6,Day Worker,Daily,600,No\n"
    ).encode())
    with app.app_context():
        assert db.session.get(Employee, "6").bonus_ignored is False


def test_import_rejects_the_bonus_flag_on_a_monthly_employee(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Month Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    body = (
        "Employee ID,Name,Wage Type,Salary,Ignore Monthly Bonus\n"
        "5,Month Worker,Monthly,30000,Yes\n"
    ).encode()
    page = import_master(client, body)
    assert b"only applies to daily wage employees" in page.data
    with app.app_context():
        assert bool(db.session.get(Employee, "5").bonus_ignored) is False


def test_excluded_employee_reads_as_excluded_not_unearned(client, app):
    with app.app_context():
        seed_daily_bonus_month({1: 540, 2: 540, 3: 540, 4: 540})
        db.session.get(Employee, "6").bonus_ignored = True
        db.session.commit()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/6").data
    assert b"Excluded" in page
    assert b"excluded from the attendance bonus in Employee Master" in page

    # The page explains why the bonus is zero; the worker's own sheet states the
    # band only, so an excluded employee reads as NIL there.
    pdf = client.get("/reports/2026-07/employee/6.pdf")
    text = "\n".join(p.extract_text() or "" for p in PdfReader(BytesIO(pdf.data)).pages)
    assert "NIL" in text
    assert "Excluded in Employee Master" not in text
    assert "Not earned this month" not in text


# --- Leave is tracked to two decimals, not one ---

def test_leave_accrual_keeps_two_decimals(app):
    from attendance.payroll_rules import calculate_monthly_leave_earned
    # 23 eligible days in a 31-day month accrues (23/31)*2 = 1.4838...
    # A single decimal truncated that to 1.4, losing nearly a tenth of a day a month.
    assert calculate_monthly_leave_earned(Decimal("23"), 31) == Decimal("1.48")
    assert calculate_monthly_leave_earned(Decimal("1"), 31) == Decimal("0.06")
    # Accrual is still truncated, never rounded up, so it cannot overshoot.
    assert calculate_monthly_leave_earned(Decimal("29"), 31) == Decimal("1.87")


def test_leave_columns_round_trip_two_decimals(app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Worker"))
        db.session.add(PayrollResult(
            payroll_month="2026-07", employee_id="5", payroll_rule_type="MONTHLY",
            calculation_status="Calculated", opening_leave=Decimal("2.75"),
            leave_earned=Decimal("1.87"), leave_used=Decimal("0.25"),
            closing_leave=Decimal("4.37"), final_salary=Decimal("30000")))
        db.session.commit()
        db.session.expire_all()
        stored = PayrollResult.query.filter_by(employee_id="5").one()
        assert stored.opening_leave == Decimal("2.75")
        assert stored.leave_earned == Decimal("1.87")
        assert stored.leave_used == Decimal("0.25")
        assert stored.closing_leave == Decimal("4.37")


def test_manual_leave_balance_keeps_two_decimals(app):
    from attendance.leave_balances import parse_leave_balance
    assert parse_leave_balance("2.75") == Decimal("2.75")
    assert parse_leave_balance("2.759") == Decimal("2.75")
    assert parse_leave_balance("3") == Decimal("3.00")


# --- Sandwich leave never reaches across a month boundary ---

def seed_boundary_month(month, year, month_number, absent_days, opening=Decimal("0")):
    """A monthly employee with Sundays off and the given days absent."""
    import calendar as cal
    db.session.add(PayrollMonth(month=month))
    db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                            normalized_salary_type="MONTHLY", salary=Decimal("30000")))
    db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
    db.session.add(SalaryRecord(payroll_month=month, employee_id="5", name="Worker",
                                salary_type="Monthly", normalized_salary_type="MONTHLY",
                                salary=Decimal("30000")))
    if opening:
        previous = f"{year - 1}-12" if month_number == 1 else f"{year}-{month_number - 1:02d}"
        db.session.add(PayrollResult(payroll_month=previous, employee_id="5", payroll_rule_type="MONTHLY",
                                     calculation_status="Calculated", closing_leave=opening,
                                     final_salary=Decimal("30000")))
    for day in range(1, cal.monthrange(year, month_number)[1] + 1):
        when = date(year, month_number, day)
        if when.weekday() == 6 or day in absent_days:
            db.session.add(AttendanceRecord(payroll_month=month, employee_id="5", employee_name="Worker",
                                            date=when, day=when.strftime("%A"), parse_status="NEEDS_REVIEW",
                                            warning="Missing punch and working hours"))
        else:
            db.session.add(AttendanceRecord(payroll_month=month, employee_id="5", employee_name="Worker",
                                            date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                            last_punch="06:30 PM", raw_working_hours="9h 00m",
                                            actual_minutes=540, parse_status="OK"))
    db.session.commit()
    calculate_payroll_month(month)
    result = PayrollResult.query.filter_by(payroll_month=month, employee_id="5").one()
    return {row["date"]: row["attendance_status"] for row in result.detail_json}, result


def test_week_off_on_the_first_of_the_month_stays_a_paid_week_off(app):
    """1 Nov 2026 is a Sunday. The day before it is in October, so no sandwich."""
    with app.app_context():
        by_date, result = seed_boundary_month("2026-11", 2026, 11, absent_days={2})
        # Paid as a week off, whatever October looked like.
        assert by_date["2026-11-01"] == "Week Off"
        assert "Sandwich Leave" not in by_date.values()
        assert result.week_offs == 5


def test_week_off_on_the_last_day_of_the_month_stays_a_paid_week_off(app):
    """31 May 2026 is a Sunday. The day after it is in June, so no sandwich."""
    with app.app_context():
        by_date, result = seed_boundary_month("2026-05", 2026, 5, absent_days={29, 30})
        assert by_date["2026-05-31"] == "Week Off"
        assert "Sandwich Leave" not in by_date.values()


def test_sandwich_still_applies_to_week_offs_inside_the_month(app):
    """The boundary rule must not disable the policy for ordinary mid-month weeks."""
    with app.app_context():
        # 8 Nov 2026 is a Sunday; absent on the Saturday before and Monday after.
        # The opening balance covers all three days, so the sandwich day is charged
        # to leave rather than being withdrawn for want of a balance.
        by_date, result = seed_boundary_month("2026-11", 2026, 11, absent_days={7, 9}, opening=Decimal("3"))
        assert by_date["2026-11-08"] == "Sandwich Leave"
        assert Decimal(result.leave_used) > 0


def test_sandwich_leave_with_no_balance_behind_it_is_shown_as_loss_of_pay(app):
    """Leave the employee does not hold must not read as leave on any screen.

    The day was always charged as loss of pay in the totals, but it used to keep the
    "Sandwich Leave" label and the full figure went to the leave ledger, so the
    ledger did not balance and the detail page contradicted the paid-day count.
    """
    with app.app_context():
        by_date, result = seed_boundary_month("2026-11", 2026, 11, absent_days={7, 9})

        # The balance stretches to 1.5 days, so 7 Nov is covered in full and 8 Nov
        # only halfway; 9 Nov has nothing behind it at all.
        assert by_date["2026-11-09"] == "Full Day LOP"
        assert by_date["2026-11-08"] == "Half-Day Paid Leave / Half-Day LOP"
        # Leave used never exceeds what was available, and what is left carries over.
        available = Decimal(result.opening_leave) + Decimal(result.leave_earned)
        assert Decimal(result.leave_used) <= available
        assert Decimal(result.closing_leave) == available - Decimal(result.leave_used)
        # The paid-day identity holds: nothing is counted as paid that was not paid.
        assert (Decimal(result.paid_working_days) + Decimal(result.week_offs)
                + Decimal(result.paid_leaves) + Decimal(result.holidays or 0)) == total_paid_days(result)
        assert Decimal(result.leave_used) == Decimal(result.paid_leaves)


# --- Final salary report split into per-wage-group attendance summaries ---

def seed_two_wage_groups(month="2026-07"):
    db.session.add(PayrollMonth(month=month))
    for eid, name, wage, rate, dept, desig in (
        ("5", "Month Worker", "Monthly", "30000", "Design", "Design Manager"),
        ("6", "Day Worker", "Daily", "600", "Stores", "Helper"),
    ):
        normalized = wage.upper()
        db.session.add(Employee(id=eid, name=name, salary_type=wage, normalized_salary_type=normalized,
                                salary=Decimal(rate), department=dept, designation=desig))
        db.session.add(WeekOffRule(employee_id=eid, confirmed_at=datetime.utcnow()))
        db.session.add(SalaryRecord(payroll_month=month, employee_id=eid, name=name,
                                    salary_type=wage, normalized_salary_type=normalized, salary=Decimal(rate)))
        for day in (1, 2):
            when = date(2026, 7, day)
            db.session.add(AttendanceRecord(payroll_month=month, employee_id=eid, employee_name=name,
                                            date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                            last_punch="06:30 PM", raw_working_hours="9h 00m",
                                            actual_minutes=540, parse_status="OK"))
    db.session.commit()
    calculate_payroll_month(month)


def pdf_text(data):
    return "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(data)).pages)


def slip_text(client, app, employee_id="5", month="2026-07"):
    """Read a salary slip. Slips only leave the app once the wage group is signed
    off, so the month is finalized first."""
    finalize_group(app, month, "MONTHLY")
    return pdf_text(client.get(f"/reports/{month}/employee/{employee_id}.pdf").data)


def test_monthly_summary_drops_pay_figures_but_keeps_the_logo(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    response = client.get("/reports/2026-07/attendance-summary-monthly.pdf")
    assert response.status_code == 200
    text = pdf_text(response.data)
    assert "Attendance Summary" in text
    assert "Month Worker" in text
    # Daily wage employees belong to the other report.
    assert "Day Worker" not in text
    # Every pay figure from the salary slip is gone.
    for removed in ("Salary Slip", "Payable Salary", "In Words", "Base Salary", "Wage Type"):
        assert removed not in text, removed
    # The attendance and leave picture stays.
    for kept in ("Paid Working Days", "Week Offs", "Leave Earned This Month", "LOP Days"):
        assert kept in text, kept


def test_daily_summary_carries_no_company_branding_anywhere(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    response = client.get("/reports/2026-07/summary-daily-wage.pdf")
    assert response.status_code == 200
    # Not in the text, not in the raw bytes, not in the PDF metadata, not the filename.
    assert b"SMARTfill" not in response.data
    assert "smartfill" not in response.headers["Content-Disposition"].lower()
    reader = PdfReader(BytesIO(response.data))
    assert "SMARTfill" not in str(reader.metadata)
    text = pdf_text(response.data)
    assert "Day Worker" in text
    assert "Month Worker" not in text
    # No employee number, no designation, no department on a daily sheet.
    for removed in ("Salary Slip", "Payable Salary", "Daily Wage", "Helper", "Stores", "Days in Month"):
        assert removed not in text, removed
    for kept in ("Payable Days", "Absence This Month", "Bonus"):
        assert kept in text, kept


def test_department_wise_report_lists_monthly_before_daily(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    response = client.get("/reports/2026-07/department-wise.pdf")
    assert response.status_code == 200
    text = pdf_text(response.data)
    assert text.index("Monthly Wage Employees") < text.index("Daily Wage Employees")
    assert text.index("Design") < text.index("Stores")
    assert "Month Worker" in text and "Day Worker" in text
    assert "employee(s)" in text and "Total" in text


def test_department_wise_report_carries_attendance_not_pay(client, app):
    """It goes to department managers, so no pay figure may appear on it."""
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    text = pdf_text(client.get("/reports/2026-07/department-wise.pdf").data)
    for banned in ("Payable", "Base Salary", "Deduction", "Salary", "30000", "30,000"):
        assert banned not in text, banned
    # The attendance and leave figures a manager tracks. Column headings wrap
    # inside their column, so match the words rather than the whole heading.
    for kept in ("Days", "Working", "Paid", "LOP", "Leave", "Used", "Earned", "CF"):
        assert kept in text, kept
    # Daily wage employees have no leave, so they are measured on absence instead.
    assert "Absence" in text
    # Each employee's month calendar is included.
    assert "SUN" in text and "SAT" in text


def test_every_department_starts_on_a_new_page(client, app):
    """A department's sheet can be torn off without another department's rows on it."""
    with app.app_context():
        seed_two_wage_groups()
        # A second monthly department, so the monthly group has two to separate.
        db.session.add(Employee(id="7", name="Store Keeper", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("20000"),
                                department="Stores", designation="Storekeeper"))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="7", name="Store Keeper",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("20000")))
        db.session.add(PayrollResult(payroll_month="2026-07", employee_id="7", payroll_rule_type="MONTHLY",
                                     calculation_status="Calculated", final_salary=Decimal("20000")))
        db.session.commit()

    login(client)
    pages = [page.extract_text() or "" for page in
             PdfReader(BytesIO(client.get("/reports/2026-07/department-wise.pdf").data)).pages]
    headings = [
        [name for name in ("Design —", "Stores —") if name in text]
        for text in pages
    ]
    # No page carries the heading of two different departments.
    assert all(len(found) <= 1 for found in headings), headings
    design = next(i for i, found in enumerate(headings) if "Design —" in found)
    stores = next(i for i, found in enumerate(headings) if "Stores —" in found)
    assert design != stores


def test_department_wise_wage_groups_start_on_their_own_page(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    pages = [page.extract_text() or "" for page in
             PdfReader(BytesIO(client.get("/reports/2026-07/department-wise.pdf").data)).pages]
    monthly = next(i for i, t in enumerate(pages) if "Monthly Wage Employees" in t)
    daily = next(i for i, t in enumerate(pages) if "Daily Wage Employees" in t)
    assert daily > monthly
    assert "Monthly Wage Employees" not in pages[daily]


def test_new_reports_are_listed_on_the_reports_page(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/reports/").data
    assert b"Attendance Summary for Monthly" in page
    assert b"Summary for Daily Wage Group" in page
    assert b"Department Wise Attendance" in page


def test_summary_reports_handle_a_wage_group_with_no_employees(client, app):
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.commit()

    login(client)
    for path in ("attendance-summary-monthly.pdf", "summary-daily-wage.pdf", "department-wise.pdf"):
        response = client.get(f"/reports/2026-07/{path}")
        assert response.status_code == 200, path


# --- Report refinements from the marked-up PDFs ---

def test_daily_summary_title_is_summary_and_bonus_is_just_a_percentage(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    text = pdf_text(client.get("/reports/2026-07/summary-daily-wage.pdf").data)
    assert "Summary" in text
    assert "Attendance Summary" not in text
    # The bonus reads as a bare percentage; the rate it applies to is not restated.
    assert "10%" in text
    assert "of earned wage" not in text


def test_monthly_summary_groups_days_then_leave_then_attendance_time(client, app):
    with app.app_context():
        seed_two_wage_groups()
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        result.less_hours_minutes = 60
        result.less_hours_deduction = Decimal("672.22")
        details = list(result.detail_json or [])
        details[0] = {**details[0], "shortage_minutes": 165}
        result.detail_json = details
        db.session.commit()

    login(client)
    text = pdf_text(client.get("/reports/2026-07/attendance-summary-monthly.pdf").data)
    order = [text.index(label) for label in (
        "Paid Working Days", "Total Paid Days", "Week Offs", "LOP Days",
        "Leave Balance", "Leave Carry Forwarded", "Less Hours", "Adjustment",
    )]
    assert order == sorted(order), "day counts, then leave, then attendance time"
    assert "1h 00m" in text
    assert "2h 45m short" in text
    assert "Less Hours Deduction" not in text
    assert pdf_money(Decimal("672.22")) not in text
    # Overtime and adjustment share the final row.
    assert text.index("Over Time") < text.index("Adjustment")


def test_attendance_summary_status_badges_keep_status_readable(client, app):
    with app.app_context():
        seed_two_wage_groups()
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        result.calculation_status = "Needs Review"
        db.session.commit()

    login(client)
    response = client.get("/reports/2026-07/attendance-summary-monthly.pdf")
    assert response.status_code == 200
    text = pdf_text(response.data)
    assert "Status" in text
    assert "Needs Review" in text


def test_monthly_total_paid_days_include_holidays(app):
    from attendance.reports import total_paid_days

    result = PayrollResult(
        payroll_rule_type="MONTHLY",
        paid_working_days=Decimal("22.00"),
        week_offs=5,
        paid_leaves=Decimal("2.00"),
        holidays=2,
    )

    assert total_paid_days(result) == Decimal("31.00")


def test_payroll_summary_splits_wage_groups_onto_their_own_pages(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    reader = PdfReader(BytesIO(client.get("/reports/2026-07/payroll-summary.pdf").data))
    assert len(reader.pages) == 2
    first, second = (page.extract_text() or "" for page in reader.pages)
    assert "Payroll Summary - Monthly Wage" in first
    assert "Month Worker" in first and "Day Worker" not in first
    assert "Payroll Summary - Daily Wage" in second
    assert "Day Worker" in second and "Month Worker" not in second


def test_payroll_summary_lists_employees_in_numeric_id_order(app):
    from attendance.reports import payroll_summary_rows
    with app.app_context():
        seed_two_wage_groups()
        for eid in ("2", "10", "11"):
            db.session.add(Employee(id=eid, name=f"Worker {eid}", salary_type="Monthly",
                                    normalized_salary_type="MONTHLY", salary=Decimal("30000")))
            db.session.add(SalaryRecord(payroll_month="2026-07", employee_id=eid, name=f"Worker {eid}",
                                        salary_type="Monthly", normalized_salary_type="MONTHLY",
                                        salary=Decimal("30000")))
            db.session.add(PayrollResult(payroll_month="2026-07", employee_id=eid, payroll_rule_type="MONTHLY",
                                         calculation_status="Calculated", final_salary=Decimal("30000")))
        db.session.commit()
        rows, _payable, _deduction = payroll_summary_rows("2026-07", "MONTHLY")
        ids = [row[0] for row in rows]
        # A string sort would put 10 and 11 before 2.
        assert ids == ["2", "5", "10", "11"]


def test_payroll_month_table_shows_paid_ot_and_less_hours_for_both_groups(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    html = client.get("/payroll/2026-07").data.decode()
    for label in ("Monthly wage employees", "Daily wage employees"):
        section = html[html.index(label):html.index(label) + 1400]
        assert "<th>Paid OT</th>" in section, label
        assert "<th>Less Hours Deduction</th>" in section, label


def test_daily_calculation_detail_column_is_named_attendance_bonus(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/payroll/2026-07/employee/6").data
    assert b"<th>Attendance bonus</th>" in page
    assert b"Bonus absence" not in page


def test_long_designation_wraps_instead_of_overrunning_its_column(client, app):
    """A plain string cell does not wrap in ReportLab, so it ran into the next column."""
    with app.app_context():
        seed_two_wage_groups()
        employee = db.session.get(Employee, "5")
        employee.designation = "Head of Business Development and Strategic Partnerships"
        db.session.commit()

    login(client)
    text = pdf_text(client.get("/reports/2026-07/department-wise.pdf").data)
    # Wrapping splits the designation across lines, so the words survive but the
    # single-line run does not. The day count that used to be overlapped is intact.
    assert "Head of" in text
    assert "Development" in text
    assert "Month Worker" in text


# --- PF and ESIC statutory contributions ---

def test_pf_is_capped_at_the_statutory_wage_ceiling(app):
    from attendance.statutory import pf_contributions
    # Basic well above the ceiling contributes on the ceiling only.
    pf = pf_contributions(Decimal("43614.52"))
    assert pf["wage"] == Decimal("15000.00")
    assert pf["employee"] == Decimal("1800")
    assert pf["employer"] == Decimal("1800")
    # Pension is 8.33% of the ceiling; the fund takes the balance of the 12%.
    assert pf["pension"] == Decimal("1250")
    assert pf["fund"] == Decimal("550")
    assert pf["pension"] + pf["fund"] == pf["employer"]
    assert pf["edli"] == Decimal("75")
    assert pf["admin"] == Decimal("75")


def test_pf_follows_earned_basic_below_the_ceiling(app):
    from attendance.statutory import pf_contributions
    # 25.5 of 31 days on a basic of 15,000 earns 12,338.71.
    pf = pf_contributions(Decimal("12338.71"))
    assert pf["wage"] == Decimal("12338.71")
    assert pf["employee"] == Decimal("1481")


def test_esi_stops_above_the_wage_ceiling(app):
    from attendance.statutory import esi_contributions
    covered = esi_contributions(Decimal("18500"), Decimal("18000"))
    assert covered["covered"] is True
    # Rounded up to the next rupee, as ESIC requires.
    assert covered["employee"] == Decimal("139")
    assert covered["employer"] == Decimal("602")
    outside = esi_contributions(Decimal("22000"), Decimal("22000"))
    assert outside["covered"] is False
    assert outside["employee"] == Decimal("0")
    assert outside["employer"] == Decimal("0")


def test_overtime_counts_for_esi_but_does_not_end_coverage(app):
    """Overtime is part of ESI wages but must not push someone out of the scheme."""
    from attendance.statutory import esi_contributions
    # Wage 20,800 plus 500 overtime exceeds the ceiling in total, but eligibility is
    # judged on the 20,800.
    result = esi_contributions(Decimal("21300"), Decimal("20800"))
    assert result["covered"] is True
    assert result["wage"] == Decimal("21300.00")


def seed_statutory_employee(pf=True, esic=False, salary="30000", basic="19500", hra="10500"):
    db.session.add(PayrollMonth(month="2026-07"))
    db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                            normalized_salary_type="MONTHLY", salary=Decimal(salary),
                            basic_salary=Decimal(basic), hra=Decimal(hra), allowance=Decimal("0"),
                            pf_enabled=pf, esic_enabled=esic))
    db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
    db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Worker",
                                salary_type="Monthly", normalized_salary_type="MONTHLY",
                                salary=Decimal(salary)))
    add_july_attendance(set())
    db.session.commit()
    calculate_payroll_month("2026-07")
    return PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()


def test_pf_employee_share_reduces_the_payable_salary(app):
    with app.app_context():
        result = seed_statutory_employee(pf=True)
        assert Decimal(result.pf_employee) == Decimal("1800.00")
        # The employer share is a company cost and must not touch take-home pay.
        assert Decimal(result.pf_employer) == Decimal("1800.00")
        assert Decimal(result.total_deduction) >= Decimal("1800.00")
        assert Decimal(result.final_salary) == (
            Decimal("30000") - Decimal(result.total_deduction) + Decimal(result.total_addition))


def test_no_statutory_deduction_when_the_flags_are_off(app):
    with app.app_context():
        result = seed_statutory_employee(pf=False, esic=False)
        assert Decimal(result.pf_employee) == 0
        assert Decimal(result.pf_employer) == 0
        assert Decimal(result.esi_employee) == 0
        # Professional tax still applies; only PF and ESI are switched off.
        assert Decimal(result.professional_tax) == Decimal("200.00")
        assert Decimal(result.final_salary) == Decimal("29800.00")


def test_esi_applies_when_the_employee_is_under_the_ceiling(app):
    with app.app_context():
        result = seed_statutory_employee(pf=False, esic=True, salary="18000", basic="11700", hra="6300")
        assert Decimal(result.esi_wage) > 0
        assert Decimal(result.esi_employee) > 0
        assert Decimal(result.esi_employer) > Decimal(result.esi_employee)


def test_daily_wage_never_carries_statutory_contributions(app):
    with app.app_context():
        result = daily_bonus_result({1: 540, 2: 540, 3: 540, 4: 540})
        assert Decimal(result.pf_employee or 0) == 0
        assert Decimal(result.esi_employee or 0) == 0


def test_statutory_contributions_are_shown_on_the_employee_page(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data
    # PF and ESIC live in the compliance panel, not in the metric cards.
    assert b"PF (employee)" not in page
    assert b"Payroll compliance" in page
    assert b"1800.00" in page


# --- Gujarat professional tax and the salary register ---

def test_professional_tax_follows_the_gujarat_slab(app):
    from attendance.statutory import professional_tax
    assert professional_tax(Decimal("30000")) == Decimal("200")
    assert professional_tax(Decimal("12000.01")) == Decimal("200")
    # At or below the threshold there is no tax.
    assert professional_tax(Decimal("12000")) == Decimal("0")
    assert professional_tax(Decimal("0")) == Decimal("0")


def test_professional_tax_is_deducted_from_the_payable_salary(app):
    with app.app_context():
        result = seed_statutory_employee(pf=False, esic=False)
        assert Decimal(result.professional_tax) == Decimal("200.00")
        assert Decimal(result.final_salary) == Decimal("29800.00")


def test_professional_tax_uses_the_earned_wage_not_the_contracted_one(app):
    with app.app_context():
        # Heavy loss of pay drops the earned wage under the slab, so no tax is due.
        seed_leave_month(opening=Decimal("0"))
        db.session.query(Employee).filter_by(id="5").update({"salary": Decimal("13000")})
        db.session.query(SalaryRecord).filter_by(employee_id="5").update({"salary": Decimal("13000")})
        add_july_attendance(set(range(1, 26)))
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(result.lop_days) > 0
        assert Decimal(result.professional_tax) == Decimal("0.00")


def test_salary_register_matches_the_payroll_sheet_layout(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)
    finalize_group(app, "2026-07", "MONTHLY")

    login(client)
    response = client.get("/reports/2026-07/salary-sheet.pdf")
    assert response.status_code == 200
    text = pdf_text(response.data)
    for heading in ("Sr.NO", "NAME", "Attendance", "WO", "BASIC", "HRA", "ALLOWANCE",
                    "PAID BASIC", "PAID HRA", "LATE", "LOAN", "ESI", "PF", "NET SALARY"):
        assert heading in text, heading


def test_salary_register_downloads_as_xlsx(client, app):
    import openpyxl
    with app.app_context():
        seed_statutory_employee(pf=True)
    finalize_group(app, "2026-07", "MONTHLY")

    login(client)
    response = client.get("/reports/2026-07/salary-sheet.xlsx")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["Content-Type"]
    assert response.headers["Content-Disposition"].endswith(".xlsx")

    sheet = openpyxl.load_workbook(BytesIO(response.data)).active
    assert sheet.cell(1, 1).value == "Salary Sheet - July 2026"
    assert sheet.cell(2, 2).value == "Days of Month"
    assert sheet.cell(2, 3).value == 31
    assert [sheet.cell(3, c).value for c in (1, 2, 3)] == ["Sr.NO", "ID", "NAME"]
    # Money lands as numbers, not text, so the sheet stays usable for further work.
    basic = sheet.cell(4, 8)
    assert isinstance(basic.value, (int, float))
    assert basic.number_format == "#,##0.00"
    # The rate cells sit directly above the columns they drive.
    assert sheet.cell(2, 18).value == 0.0075
    assert sheet.cell(2, 19).value == 0.12
    assert sheet.cell(3, 18).value == "ESI"
    assert sheet.cell(3, 19).value == "PF"


def test_salary_register_totals_the_money_columns(app):
    from attendance.reports import salary_register_rows, salary_register_totals
    with app.app_context():
        seed_statutory_employee(pf=True)
        rows = salary_register_rows("2026-07")
        totals = salary_register_totals(rows)
        assert len(rows) == 1
        # Net salary is the last column and must total the rows above it.
        assert totals[20] == Decimal(rows[0][20])
        assert totals[18] == Decimal("1800.00")
        assert totals[19] == Decimal("200.00")


def test_salary_register_excludes_daily_wage_employees(app):
    from attendance.reports import salary_register_rows
    with app.app_context():
        seed_two_wage_groups()
        rows = salary_register_rows("2026-07")
        assert [row[2] for row in rows] == ["Month Worker"]


def test_payroll_compliance_panel_shows_both_sides_of_each_contribution(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    assert "Payroll compliance" in page
    figures = dict(re.findall(r"<dt>([^<]+)</dt><dd>([^<]+)</dd>", page))
    assert figures["Employee PF"] == "1800.00"
    assert figures["Employer PF"] == "1800.00"
    # The employer 12% splits into pension and fund, and both are shown.
    assert figures["Pension (EPS)"] == "1250.00"
    assert figures["Fund (EPF)"] == "550.00"
    assert figures["EDLI"] == "75.00"
    assert figures["Admin charges"] == "75.00"
    assert figures["Professional tax"] == "200.00"
    # Deducted 1800 PF + 200 PT; company pays 1800 + 75 EDLI + 75 admin.
    assert figures["Deducted from employee"] == "2000.00"
    assert figures["Paid by company"] == "1950.00"


def test_uncovered_contributions_read_as_zero_not_a_bare_digit(client, app):
    """Decimal("0.00") is falsy, so a plain `or 0` printed a bare 0."""
    with app.app_context():
        seed_statutory_employee(pf=True, esic=False)

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    figures = dict(re.findall(r"<dt>([^<]+)</dt><dd>([^<]+)</dd>", page))
    assert figures["Employee ESIC"] == "0.00"
    assert figures["Employer ESIC"] == "0.00"
    assert "Not applicable" in page


def test_daily_wage_has_no_compliance_panel(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    assert b"Payroll compliance" not in client.get("/payroll/2026-07/employee/6").data
    assert b"Payroll compliance" in client.get("/payroll/2026-07/employee/5").data


# --- Leave is redeemed in half-day steps ---

def test_leave_is_redeemable_only_in_half_day_steps(app):
    from attendance.payroll_rules import redeemable_leave
    # The two cases from the rule: 1.38 redeems 1 day, 1.92 redeems 1.5.
    assert redeemable_leave(Decimal("1.38")) == Decimal("1")
    assert redeemable_leave(Decimal("1.92")) == Decimal("1.5")
    assert redeemable_leave(Decimal("0.80")) == Decimal("0.5")
    # A fraction under half a day cannot offset anything.
    assert redeemable_leave(Decimal("0.49")) == Decimal("0")
    assert redeemable_leave(Decimal("0.50")) == Decimal("0.5")
    assert redeemable_leave(Decimal("3.75")) == Decimal("3.5")
    assert redeemable_leave(Decimal("0")) == Decimal("0")
    assert redeemable_leave(Decimal("-1")) == Decimal("0")


def absent_rows(count=6):
    from attendance.payroll_rules import ABSENT_STATUS
    return [(None, None, {"status": ABSENT_STATUS, "paid_day": Decimal("0"),
                          "leave_used": Decimal("0"), "explanation": ""}) for _ in range(count)]


def test_a_balance_of_1_38_covers_exactly_one_absence(app):
    from attendance.payroll_rules import apply_leave_balance
    rows = absent_rows()
    assert apply_leave_balance(rows, Decimal("1.38")) == Decimal("1")
    statuses = [row[2]["status"] for row in rows]
    assert statuses[0] == "Paid Leave"
    # The 0.38 remainder cannot split a second day.
    assert statuses[1] == "Absent / Attendance Missing"


def test_a_balance_of_1_92_covers_a_day_and_a_half(app):
    from attendance.payroll_rules import apply_leave_balance, HALF_LEAVE_HALF_LOP_STATUS
    rows = absent_rows()
    assert apply_leave_balance(rows, Decimal("1.92")) == Decimal("1.5")
    statuses = [row[2]["status"] for row in rows]
    assert statuses[0] == "Paid Leave"
    assert statuses[1] == HALF_LEAVE_HALF_LOP_STATUS
    assert statuses[2] == "Absent / Attendance Missing"


def test_leftover_opening_leave_joins_the_earned_pool(app):
    """Leave is settled in two passes; a fraction must not be stranded in one."""
    with app.app_context():
        # An opening of 0.4 redeems nothing on its own, and so does an earning of 0.4,
        # but together they make 0.8, which redeems half a day.
        seed_leave_month(opening=Decimal("0.4"))
        add_july_attendance({3})
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(result.leave_earned) > Decimal("0.4")
        assert Decimal(result.leave_used) >= Decimal("0.5")
        by_date = {row["date"]: row["attendance_status"] for row in result.detail_json}
        assert by_date["2026-07-03"] != "Full Day LOP"


def test_carried_balance_shows_what_is_redeemable_as_leave(client, app):
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        carried = Decimal(result.closing_leave)

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    assert "Leave carried forward" in page
    assert "Redeemable as leave:" in page
    from attendance.payroll_rules import redeemable_leave
    assert str(redeemable_leave(carried)) in page


def test_encashment_days_keep_two_decimals(app):
    from routes.payroll import parse_leave_encashment_days
    # A remainder such as 0.38 can be encashed in full; one decimal lost 0.08 of it.
    assert parse_leave_encashment_days("0.38") == Decimal("0.38")
    assert parse_leave_encashment_days("1.929") == Decimal("1.92")
    assert parse_leave_encashment_days("") == Decimal("0")


# --- Employee page: one action bar, navigation out of the form ---

def test_employee_page_has_one_pair_of_submit_buttons(client, app):
    """The panel and the sticky bar each carried Save and Recalculate: four buttons
    for two actions."""
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    submits = re.findall(r'name="action" value="(\w+)">([^<]+)</button>', page)
    assert submits == [("save", "Save only"), ("recalculate", "Save &amp; recalculate")]


def test_navigation_links_sit_outside_the_form(client, app):
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    head, _, rest = page.partition('<form method="post" id="employeeDetailForm">')
    assert "Back to payroll" in head and "Back to payroll" not in rest
    # The month is still a draft, so the slip control is present but disabled.
    assert "Slip after finalize" in head and "Slip after finalize" not in rest
    assert "Open PDF" not in page


def test_finalized_month_shows_no_submit_buttons_but_keeps_navigation(client, app):
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        calculate_payroll_month("2026-07")
        month = db.session.get(PayrollMonth, "2026-07")
        month.status = "FINALIZED"
        month.monthly_finalized_at = datetime.utcnow()
        month.daily_finalized_at = datetime.utcnow()
        db.session.commit()

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    assert 'name="action" value="save"' not in page
    assert 'name="action" value="recalculate"' not in page
    assert "Back to payroll" in page and "Open PDF" in page


def test_adjustment_cards_match_the_wage_type(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    def cards(employee_id):
        page = client.get(f"/payroll/2026-07/employee/{employee_id}").data.decode()
        return [f.strip() for f in re.findall(
            r'<div class="adjust-field[^"]*">\s*<(?:label|span)[^>]*>(?:\s*<input[^>]*>\s*<span>)?([^<]+)', page)]

    # Leave encashment is monthly only; the attendance bonus is daily only.
    assert "Encash leave" in cards("5")
    assert "Attendance bonus" not in cards("5")
    assert "Attendance bonus" in cards("6")
    assert "Encash leave" not in cards("6")


def attendance_upload_csv(rows):
    headers = [
        "Employee ID", "Employee Name", "Department", "Designation", "Date", "Day",
        "Shift", "From", "To", "First Punch", "Last Punch", "Total Working Hours",
    ]
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode()


def test_employee_attendance_reimport_only_replaces_that_employee(client, app):
    with app.app_context():
        seed_two_wage_groups()
        db.session.add(AttendanceOverride(
            payroll_month="2026-07", employee_id="5", date=date(2026, 7, 1),
            manual_status="Paid Leave", notes="old correction",
        ))
        db.session.add(AttendanceOverride(
            payroll_month="2026-07", employee_id="6", date=date(2026, 7, 1),
            manual_status="Paid Leave", notes="keep this",
        ))
        db.session.commit()
        old_other = AttendanceRecord.query.filter_by(
            payroll_month="2026-07", employee_id="6", date=date(2026, 7, 1)
        ).one()
        old_other_last_punch = old_other.last_punch

    login(client)
    upload = attendance_upload_csv([
        {
            "Employee ID": "5", "Employee Name": "Month Worker", "Department": "Design",
            "Designation": "Design Manager", "Date": "03-07-2026", "Day": "Friday",
            "Shift": "Normal Shift", "From": "", "To": "", "First Punch": "09:45 AM",
            "Last Punch": "06:45 PM", "Total Working Hours": "9h 00m",
        },
        {
            "Employee ID": "6", "Employee Name": "Day Worker", "Department": "Stores",
            "Designation": "Helper", "Date": "01-07-2026", "Day": "Wednesday",
            "Shift": "Normal Shift", "From": "", "To": "", "First Punch": "10:00 AM",
            "Last Punch": "07:00 PM", "Total Working Hours": "9h 00m",
        },
    ])
    response = client.post(
        "/payroll/2026-07/employee/5",
        data={
            "action": "reimport_attendance",
            "employee_attendance_file": (BytesIO(upload), "corrected.csv"),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert b"Payroll recalculated for this employee only" in response.data

    with app.app_context():
        employee_records = AttendanceRecord.query.filter_by(payroll_month="2026-07", employee_id="5").all()
        assert [(record.date, record.first_punch, record.last_punch) for record in employee_records] == [
            (date(2026, 7, 3), "09:45 AM", "06:45 PM")
        ]
        other_record = AttendanceRecord.query.filter_by(
            payroll_month="2026-07", employee_id="6", date=date(2026, 7, 1)
        ).one()
        assert other_record.last_punch == old_other_last_punch
        assert AttendanceOverride.query.filter_by(payroll_month="2026-07", employee_id="5").count() == 0
        assert AttendanceOverride.query.filter_by(payroll_month="2026-07", employee_id="6").count() == 1
        assert PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").count() == 1
        assert PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").count() == 1
        audit = AuditLog.query.filter_by(action="Employee Attendance Re-imported").one()
        assert "Employee ID 5" in audit.detail


def test_employee_page_has_reimport_form(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    assert "Re-import this employee attendance" in page
    assert 'name="employee_attendance_file"' in page
    assert 'accept=".csv,.xlsx"' in page
    assert "/payroll/2026-07/employee/5/attendance.csv" in page


def test_employee_attendance_download_exports_only_that_employee(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    response = client.get("/payroll/2026-07/employee/5/attendance.csv")
    assert response.status_code == 200
    assert response.headers["Content-Disposition"] == "attachment; filename=smartfill-employee-5-attendance-2026-07.csv"
    rows = list(csv.DictReader(StringIO(response.data.decode())))
    assert len(rows) == 2
    assert {row["Employee ID"] for row in rows} == {"5"}
    assert rows[0]["Employee Name"] == "Month Worker"
    assert rows[0]["Date"] == "2026-07-01"
    assert rows[0]["First Punch"] == "09:30 AM"
    assert rows[0]["Last Punch"] == "06:30 PM"
    assert rows[0]["Total Working Hours"] == "9h 00m"


# --- Wage filtering is one control, not a button per card ---

def test_wage_filter_is_a_single_segmented_control(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/payroll/2026-07").data.decode()
    options = re.findall(r'wage-filter-option[^>]*>([^<]+)<', page)
    assert options == ["All (2)", "Monthly (1)", "Daily (1)"]
    # The per-card "View X only" buttons and the "Show all wage types" alert are gone.
    assert "View monthly only" not in page
    assert "Show all wage types" not in page


def test_wage_filter_marks_the_current_view(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    for wage, active in (("", "All"), ("monthly", "Monthly"), ("daily", "Daily")):
        page = client.get(f"/payroll/2026-07?wage={wage}").data.decode()
        current = re.findall(r'wage-filter-option is-active"[^>]*>([^ ]+)', page)
        assert current == [active], (wage, current)


def test_filtered_recalculate_keeps_the_wage_group(client, app):
    """The step's buttons must post the filter, or a filtered view would recalculate
    every wage group."""
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/payroll/2026-07?wage=daily").data.decode()
    step = re.search(r'<li class="workflow-step[^>]*>(?:(?!</li>).)*?Calculate payroll.*?</li>', page, re.S).group(0)
    assert step.count('name="wage_group" value="DAILY"') == 2
    # Unfiltered, the step posts no wage group, so every open group recalculates.
    page = client.get("/payroll/2026-07").data.decode()
    steps = re.search(r'<ol class="workflow-steps".*?</ol>', page, re.S).group(0)
    assert 'name="wage_group"' not in steps


# --- Page flow: one action location per page, navigation out of forms ---

def visible_buttons(page):
    """Buttons the user actually sees: dialog contents are hidden until opened."""
    body = re.sub(r"<dialog.*?</dialog>", "", page, flags=re.S)
    found = re.findall(r'class="btn[^"]*"[^>]*>\s*([^<]{1,40}?)\s*<', body)
    return [b.strip() for b in found if b.strip() and b.strip() != "Logout"]


def test_month_picker_explains_itself_and_lists_recent_months(client, app):
    with app.app_context():
        seed_month()

    login(client)
    page = client.get("/payroll/new").data.decode()
    assert visible_buttons(page) == ["Open payroll month"]
    # Says what pressing it does, rather than "Continue".
    assert "A month that does not exist yet is created" in page
    # Existing months are reachable without retyping them.
    assert "Recent months" in page
    assert "July 2026" in page


def test_week_offs_has_one_save_and_a_search(client, app):
    with app.app_context():
        seed_month()

    login(client)
    page = client.get("/weekoffs").data.decode()
    assert visible_buttons(page) == ["Save week offs"]
    assert 'id="weekoffSearch"' in page
    # The save sits with the rows it applies to, not above a long table.
    assert "sticky-actions" in page


def test_employee_list_has_three_actions_and_quiet_row_links(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/master").data.decode()
    assert visible_buttons(page) == ["Export", "Import", "Add employee"]
    assert 'id="employeeSearch"' in page
    # Marking someone left is a link, not a red button on every row.
    assert 'class="btn btn-sm btn-outline-danger disable-employee-button"' not in page
    assert "link-button is-danger disable-employee-button" in page


def test_attendance_manager_submits_from_one_sticky_bar(client, app):
    with app.app_context():
        seed_month()

    login(client)
    page = client.get("/attendance/2026-07").data.decode()
    submits = re.findall(r'data-attendance-action="(\w+)"[^>]*>\s*([^<]+?)\s*</button>', page)
    assert submits == [("save", "Save only"), ("submit", "Save &amp; calculate payroll")]
    # They live in the sticky bar at the end of the form, not in the card header.
    _, _, after_table = page.partition("table-footer")
    assert "Save &amp; calculate payroll" in after_table
    assert "module-card-actions" not in page


# --- Salary slips are a monthly wage document only ---

def test_salary_slips_cover_monthly_wage_only(client, app):
    with app.app_context():
        seed_two_wage_groups()
    finalize_group(app, "2026-07", "MONTHLY")

    login(client)
    response = client.get("/reports/2026-07/final-report.pdf")
    assert response.status_code == 200
    assert "salary-slips" in response.headers["Content-Disposition"]
    text = pdf_text(response.data)
    assert "Pay Slip" in text
    assert "EARNINGS (INR)" in text
    assert "Month Worker" in text
    # Daily wage employees are not issued a salary slip.
    assert "Day Worker" not in text


def test_a_daily_employee_pdf_is_an_attendance_summary_not_a_slip(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    response = client.get("/reports/2026-07/employee/6.pdf")
    assert response.status_code == 200
    text = pdf_text(response.data)
    assert "Pay Slip" not in text
    assert "Day Worker" in text
    # Same regulatory rule as the bulk daily report: no branding anywhere,
    # including the filename.
    assert b"SMARTfill" not in response.data
    assert "smartfill" not in response.headers["Content-Disposition"].lower()


def test_a_monthly_employee_pdf_is_still_a_salary_slip(client, app):
    with app.app_context():
        seed_two_wage_groups()
    finalize_group(app, "2026-07", "MONTHLY")

    login(client)
    response = client.get("/reports/2026-07/employee/5.pdf")
    assert response.status_code == 200
    assert "Pay Slip" in pdf_text(response.data)
    assert "salary-slip" in response.headers["Content-Disposition"]


def test_reports_page_names_the_slip_report_by_wage_group(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    page = client.get("/reports/").data
    assert b"Salary Slips (Monthly)" in page
    assert b"Final Salary Report" not in page


# --- Salary slip in the classic payslip layout ---

def test_salary_slip_has_the_payslip_structure(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    text = slip_text(client, app)
    for heading in ("Pay Slip:", "Payable days:", "Loss of pay days:",
                    "Employee Name", "Employee Code", "Department", "Designation",
                    "EARNINGS (INR)", "DEDUCTIONS (INR)", "Actual Amount", "Paid Amount",
                    "Sub Total", "Others", "Gross Pay", "Gross Deductions", "Net Pay",
                    "In words:", "PF & ESIC Contributions",
                    "PF contribution by Employer", "ESIC contribution by Employer",
                    "PF Admin & EDLI Charges", "ESIC contribution by Employee",
                    "PF contribution by Employee",
                    "Total contribution to PF account",
                    "LEAVE SUMMARY", "Balance from last month", "Earned this month",
                    "Used this month", "Carry forward",
                    "system generated payslip"):
        assert heading in text, heading
    # Only the three components survive from the sample slip.
    assert "Basic" in text and "House Rent Allowance" in text and "Conveyance Allowance" in text
    for dropped in ("Kit Allowance", "Medical Allowance", "Special Allowance", "Travel Allowance"):
        assert dropped not in text, dropped
    # Bank, PF number, UAN and PAN are deliberately absent.
    for dropped in ("Bank A/C", "UAN No", "PAN No", "Business Unit", "Cost Center", "Date of Join"):
        assert dropped not in text, dropped


def test_slip_net_pay_reconciles_with_the_payroll_result(app):
    from attendance.reports import slip_earning_rows, slip_other_earnings, slip_other_deductions
    with app.app_context():
        result = seed_statutory_employee(pf=True)
        salary = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        employee = db.session.get(Employee, "5")
        paid = sum((row[2] for row in slip_earning_rows(employee, salary, result)), Decimal("0"))
        other_in = sum((amount for _l, amount in slip_other_earnings(result)), Decimal("0"))
        statutory = (Decimal(result.professional_tax) + Decimal(result.pf_employee)
                     + Decimal(result.esi_employee))
        other_out = sum((amount for _l, amount in slip_other_deductions(result)), Decimal("0"))
        net = paid + other_in - statutory - other_out
        assert net.quantize(Decimal("0.01")) == Decimal(result.final_salary)


def test_paid_amounts_follow_loss_of_pay_only(app):
    """Short hours is its own deduction line, so it must not also shrink the
    paid component amounts."""
    from attendance.reports import slip_paid_ratio
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set(range(1, 8)))
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        salary = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        expected = (Decimal(salary.salary) - Decimal(result.lop_deduction)) / Decimal(salary.salary)
        assert slip_paid_ratio(salary, result) == expected


def test_slip_shows_the_whole_salary_when_no_breakup_is_captured(app):
    from attendance.reports import slip_earning_rows
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        salary = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        employee = db.session.get(Employee, "5")
        assert Decimal(employee.basic_salary or 0) == 0
        rows = slip_earning_rows(employee, salary, result)
        # Otherwise the slip would show a gross of zero against a real net figure.
        assert rows[0][0] == "Basic"
        assert rows[0][1] == Decimal(salary.salary)


def test_slip_carries_one_signed_adjustment(app):
    """A positive and a negative Adjustment row on opposite sides read as two
    separate amounts; one signed line is unambiguous."""
    from attendance.reports import slip_other_earnings, slip_other_deductions
    with app.app_context():
        result = seed_statutory_employee(pf=True)
        db.session.query(SalaryRecord).filter_by(employee_id="5").update({"adjustment": Decimal("-2500")})
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        earnings = dict(slip_other_earnings(result))
        assert earnings["Adjustment (+/-)"] == Decimal("-2500.00")
        # Deductions no longer carry a second Adjustment line.
        assert "Adjustment" not in dict(slip_other_deductions(result))


def test_slip_deductions_are_ordered_pf_esic_pt_tds(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    text = slip_text(client, app)
    order = [text.index(label) for label in ("P.F", "ESIC", "Professional Tax", "TDS")]
    assert order == sorted(order)


def test_slip_shows_employee_code_before_name(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    text = slip_text(client, app)
    assert text.index("Employee Code") < text.index("Employee Name")


def test_slip_leave_summary_matches_the_payroll_result(client, app):
    with app.app_context():
        result = seed_statutory_employee(pf=True)
        expected = [str(leave_days(getattr(result, field))) for field in
                    ("opening_leave", "leave_earned", "leave_used", "closing_leave")]

    login(client)
    text = slip_text(client, app)
    assert "LEAVE SUMMARY" in text
    for value in expected:
        assert value in text


def test_daily_slip_has_no_leave_summary(app):
    """Daily wage employees do not accrue leave, so the block would be all zeroes."""
    from attendance.reports import salary_slip_story
    from reportlab.lib.styles import getSampleStyleSheet
    with app.app_context():
        seed_two_wage_groups()
        salary = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        story = salary_slip_story("2026-07", salary, result, getSampleStyleSheet(), 190)
        rendered = " ".join(str(item) for item in story)
        assert "LEAVE SUMMARY" not in rendered


# --- Session inactivity timeout ---

def test_session_survives_fourteen_minutes_of_inactivity(client, app):
    """Reviewing a payroll month means long gaps between clicks; 5 minutes was
    logging people out mid-task."""
    client.post("/login", data={"username": "admin", "password": "12345"})
    with client.session_transaction() as user_session:
        user_session["last_activity_at"] = datetime.utcnow().timestamp() - (14 * 60)
    response = client.get("/", follow_redirects=True)
    assert b"session expired" not in response.data
    assert b"Dashboard" in response.data


def test_session_expires_after_fifteen_minutes_of_inactivity(client, app):
    client.post("/login", data={"username": "admin", "password": "12345"})
    with client.session_transaction() as user_session:
        user_session["last_activity_at"] = datetime.utcnow().timestamp() - (15 * 60 + 1)
    response = client.get("/", follow_redirects=True)
    assert b"Your session expired after 15 minutes of inactivity" in response.data


def test_timeout_message_follows_the_configured_value(app):
    """The wording used to hardcode "5 minutes", so changing the setting left the
    message telling people something untrue."""
    from attendance.authentication import inactivity_timeout_label
    for seconds, expected in ((900, "15 minutes"), (300, "5 minutes"), (60, "1 minute"), (90, "90 seconds")):
        app.config["SESSION_INACTIVITY_TIMEOUT_SECONDS"] = seconds
        with app.test_request_context():
            assert inactivity_timeout_label() == expected


def test_slip_page_carries_only_the_footnote(client, app):
    """No brand line and no page number: the masthead already identifies the slip."""
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    text = slip_text(client, app)
    assert "SMARTfill Payroll" not in text
    assert "Page 1" not in text
    assert "system generated payslip" in text


def test_slip_net_pay_carries_the_currency(client, app):
    with app.app_context():
        result = seed_statutory_employee(pf=True)
        expected = f"{Decimal(result.final_salary):,.2f} INR"

    login(client)
    text = slip_text(client, app)
    assert expected in text
    assert "EARNINGS (INR)" in text and "DEDUCTIONS (INR)" in text
    assert "(Rs.)" not in text


def test_slip_shows_the_total_paid_into_the_pf_account(client, app):
    with app.app_context():
        result = seed_statutory_employee(pf=True)
        total = Decimal(result.pf_employee) + Decimal(result.pf_employer)

    login(client)
    text = slip_text(client, app)
    assert "Total contribution to PF account" in text
    assert f"{total:,.2f}" in text


# --- A half day worked draws half a leave for the unworked half ---

def test_half_day_draws_half_a_leave_when_the_balance_allows(app):
    with app.app_context():
        seed_leave_month(opening=Decimal("1"))
        add_july_attendance(set())
        # 22 July worked 5h 21m: over the 3h half-day floor, under the 6h full-day one.
        record = AttendanceRecord.query.filter_by(employee_id="5", date=date(2026, 7, 22)).one()
        record.actual_minutes = 321
        record.raw_working_hours = "5h 21m"
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row for row in result.detail_json}
        day = by_date["2026-07-22"]
        assert day["attendance_status"] == "Half Day Present / Half-Day Leave"
        assert Decimal(day["leave_used"]) == Decimal("0.5")
        # The day is paid in full, so it costs leave rather than pay.
        assert Decimal(day["paid_day_value"]) == Decimal("1")
        assert Decimal(result.lop_days) == 0
        assert Decimal(result.leave_used) == Decimal("0.5")


def test_half_day_is_half_unpaid_when_no_leave_is_left(app):
    """Without this the unworked half was neither charged to leave nor deducted,
    so a half day was quietly paid in full."""
    with app.app_context():
        seed_leave_month(opening=Decimal("0"))
        add_july_attendance(set())
        for day in (6, 7, 8, 9, 10, 13, 14):
            record = AttendanceRecord.query.filter_by(employee_id="5", date=date(2026, 7, day)).one()
            record.actual_minutes = 321
            record.raw_working_hours = "5h 21m"
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        statuses = [row["attendance_status"] for row in result.detail_json]
        assert "Half Day Present" in statuses
        # Half a day of loss of pay for every half day the balance could not cover.
        uncovered = statuses.count("Half Day Present")
        assert Decimal(result.lop_days) == Decimal("0.5") * uncovered
        assert Decimal(result.lop_deduction) > 0


def test_daily_wage_half_days_do_not_touch_leave_or_lop(app):
    """Daily wage has no leave balance, so a half day is simply half paid."""
    with app.app_context():
        result = daily_bonus_result({1: 540, 2: 300, 3: 540, 4: 540})
        assert Decimal(result.half_days) == Decimal("1")
        assert Decimal(result.leave_used or 0) == 0
        assert Decimal(result.lop_days or 0) == 0
        assert Decimal(result.paid_working_days) == Decimal("3.5")


def test_half_day_with_leave_reads_clearly_everywhere(client, app):
    with app.app_context():
        seed_leave_month(opening=Decimal("1"))
        add_july_attendance(set())
        record = AttendanceRecord.query.filter_by(employee_id="5", date=date(2026, 7, 22)).one()
        record.actual_minutes = 321
        db.session.commit()
        calculate_payroll_month("2026-07")

    login(client)
    assert b"Half Day + Half Leave" in client.get("/payroll/2026-07/employee/5").data
    # The label wraps inside the narrow calendar cell, so match the words.
    text = pdf_text(client.get("/reports/2026-07/attendance-summary-monthly.pdf").data)
    assert "Half Day + Half" in text and "Leave" in text


# --- Off-site days: full day, no overtime, no short hours ---

def offsite_before_and_after(day, minutes, status):
    """The same day calculated with punch data, then again as an off-site override."""
    seed_leave_month(opening=Decimal("0"))
    add_july_attendance(set())
    record = AttendanceRecord.query.filter_by(employee_id="5", date=date(2026, 7, day)).one()
    record.actual_minutes = minutes
    record.raw_working_hours = f"{minutes // 60}h {minutes % 60:02d}m"
    db.session.commit()

    def calculate():
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        return result, next(x for x in result.detail_json if x["date"] == f"2026-07-{day:02d}")

    before = calculate()
    db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="5",
                                      date=date(2026, 7, day), manual_status=status))
    db.session.commit()
    return before, calculate()


def test_offsite_day_earns_no_overtime(app):
    with app.app_context():
        # 10 hours would otherwise be 60 payable overtime minutes.
        (_, plain), (result, day) = offsite_before_and_after(6, 600, "Worked On-Site")
        assert plain["raw_ot"] == 60
        assert day["attendance_status"] == "Worked On-Site"
        assert Decimal(day["paid_day_value"]) == Decimal("1")
        assert day["raw_ot"] == 0 and day["payable_ot"] == 0
        assert result.ot_minutes == 0
        assert Decimal(result.ot_amount) == 0


def test_offsite_day_attracts_no_short_hours_deduction(app):
    with app.app_context():
        # 7 hours would otherwise be 120 short-hours minutes.
        (_, plain), (result, day) = offsite_before_and_after(7, 420, "Work From Home")
        assert plain["shortage_minutes"] == 120
        assert day["attendance_status"] == "Work From Home"
        assert Decimal(day["paid_day_value"]) == Decimal("1")
        # Reported as zero too, so no phantom "Less hours" badge appears.
        assert day["shortage_minutes"] == 0
        assert result.less_hours_minutes == 0
        assert Decimal(result.less_hours_deduction) == 0


def test_daily_wage_offsite_day_is_also_exempt(app):
    with app.app_context():
        seed_daily_bonus_month({1: 600, 2: 540, 3: 540, 4: 540})
        record = AttendanceRecord.query.filter_by(employee_id="6", date=date(2026, 7, 1)).one()
        db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="6",
                                          date=record.date, manual_status="Worked On-Site"))
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        day = next(x for x in result.detail_json if x["date"] == "2026-07-01")
        assert day["attendance_status"] == "Worked On-Site"
        assert day["raw_ot"] == 0 and day["shortage_minutes"] == 0
        assert result.ot_minutes == 0
        assert Decimal(result.paid_working_days) == Decimal("4")


# --- TDS entered on the employee master and deducted as-is ---

def test_tds_from_the_master_is_deducted_and_shown_on_the_slip(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)
        db.session.get(Employee, "5").tds = Decimal("2500")
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(result.tds) == Decimal("2500.00")
        # Not derived from anything: whatever is entered is what is deducted.
        assert Decimal(result.final_salary) == Decimal("30000") - Decimal(result.total_deduction)

    login(client)
    text = slip_text(client, app)
    assert "TDS" in text
    assert "2,500.00" in text


def test_tds_is_editable_from_the_employee_page(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    login(client)
    client.post("/master/5", data={
        "employee_id": "5", "name": "Worker", "wage_type": "Monthly", "salary": "30000",
        "basic_salary": "0", "hra": "0", "allowance": "0", "tds": "1750",
        "employment_status": "ACTIVE", "master_controls_present": "1",
    }, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Employee, "5").tds == Decimal("1750.00")
        audit = AuditLog.query.filter_by(action="Employee Master Updated").first()
        assert "TDS" in audit.detail


def test_tds_round_trips_through_import_and_export(client, app):
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000"),
                                tds=Decimal("3200")))
        db.session.commit()

    login(client)
    export = client.get("/master/export.csv").data.decode()
    assert "TDS" in export.splitlines()[0]
    assert "3200.00" in export
    # An untouched export re-imports cleanly.
    assert b"imported" in import_master(client, export.encode()).data
    with app.app_context():
        assert db.session.get(Employee, "5").tds == Decimal("3200.00")

    import_master(client, (
        "Employee ID,Name,Wage Type,Salary,TDS\n"
        "5,Worker,Monthly,30000,4100\n"
    ).encode())
    with app.app_context():
        assert db.session.get(Employee, "5").tds == Decimal("4100.00")


def test_import_rejects_tds_on_a_daily_employee(client, app):
    with app.app_context():
        db.session.add(Employee(id="6", name="Day Worker", salary_type="Daily",
                                normalized_salary_type="DAILY", salary=Decimal("600")))
        db.session.commit()

    login(client)
    page = import_master(client, (
        "Employee ID,Name,Wage Type,Salary,TDS\n"
        "6,Day Worker,Daily,600,2000\n"
    ).encode())
    assert b"only applies to monthly wage employees" in page.data
    with app.app_context():
        assert Decimal(db.session.get(Employee, "6").tds or 0) == 0


def test_switching_an_employee_to_daily_clears_tds(client, app):
    """TDS is a monthly-only field, so a daily record must not keep one."""
    with app.app_context():
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000"),
                                tds=Decimal("2500")))
        db.session.commit()

    login(client)
    # Wage type cannot change on an active employee, so this exercises the daily branch
    # of the form handler directly.
    with app.app_context():
        from attendance.master import save_master_employee
        db.session.query(Employee).filter_by(id="5").update(
            {"salary_type": None, "normalized_salary_type": None})
        db.session.commit()
        save_master_employee({"employee_id": "5", "name": "Worker", "wage_type": "Daily",
                              "salary": "600", "master_controls_present": "1"}, "admin")
        db.session.commit()
        assert Decimal(db.session.get(Employee, "5").tds) == 0


def test_tds_appears_in_the_compliance_panel(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)
        db.session.get(Employee, "5").tds = Decimal("2500")
        db.session.commit()
        calculate_payroll_month("2026-07")

    login(client)
    page = client.get("/payroll/2026-07/employee/5").data.decode()
    figures = dict(re.findall(r"<dt>([^<]+)</dt><dd>([^<]+)</dd>", page))
    assert figures["TDS"] == "2500.00"
    # It is an employee-side deduction, so it belongs in that total.
    assert figures["Deducted from employee"] == "4500.00"


# --- Daily sheet states time and a band, never an amount ---

def test_daily_sheet_states_hours_not_amounts(client, app):
    with app.app_context():
        # 8h 00m worked: an hour short, so both a shortage and a deduction exist.
        result = daily_bonus_result({1: 480, 2: 540, 3: 540, 4: 540})
        assert Decimal(result.less_hours_deduction) > 0

    login(client)
    text = pdf_text(client.get("/reports/2026-07/summary-daily-wage.pdf").data)
    assert "Less Hours" in text and "Over Time" in text
    # The label lost the word "Deduction" and the value is a duration.
    assert "Less Hours Deduction" not in text
    assert "1h 00m" in text
    assert "Absence This Month\n1h 00m\nBonus" in text
    assert pdf_money(result.less_hours_deduction) not in text


def test_absence_this_month_uses_working_days():
    from attendance.utils import minutes_to_working_day_shortage

    assert minutes_to_working_day_shortage(45 * 60 + 48) == "5d 0h 48m short"
    assert minutes_to_working_day_shortage(19 * 60 + 40) == "2d 1h 40m short"
    assert minutes_to_working_day_shortage((9 * 60) + (6 * 60) + 44) == "1d 6h 44m short"


def test_daily_sheet_bonus_is_a_bare_band(client, app):
    from attendance.reports import bonus_band_text
    with app.app_context():
        earned = daily_bonus_result({1: 540, 2: 540, 3: 540, 4: 540})
        assert bonus_band_text(earned) == "10%"

    login(client)
    text = pdf_text(client.get("/reports/2026-07/summary-daily-wage.pdf").data)
    assert "Bonus" in text
    assert "Attendance Bonus" not in text
    assert "10%" in text
    # No amount, and none of the longer wordings.
    assert "of earned wage" not in text
    assert "Not earned this month" not in text
    assert pdf_money(earned.attendance_bonus_amount) not in text


def test_bonus_band_is_nil_when_nothing_is_earned(app):
    from attendance.reports import bonus_band_text
    with app.app_context():
        missed = daily_bonus_result({1: None, 2: None, 3: 540, 4: 540})
        assert bonus_band_text(missed) == "NIL"


def test_slip_pf_rows_are_ordered_employer_employee_charges(client, app):
    with app.app_context():
        seed_statutory_employee(pf=True)

    login(client)
    text = slip_text(client, app)
    order = [text.index(label) for label in
             ("PF contribution by Employer", "PF contribution by Employee", "PF Admin & EDLI Charges")]
    assert order == sorted(order)
    # The charge line says who bears it.
    assert "Paid by" in text
    assert "\nPF Employee\n" not in text


def test_slip_band_shows_days_in_month_between_period_and_payable(client, app):
    with app.app_context():
        result = seed_statutory_employee(pf=True)

    login(client)
    text = slip_text(client, app)
    assert "Days in this Month: 31" in text
    order = [text.index(label) for label in
             ("Pay Slip:", "Days in this Month:", "Payable days:", "Loss of pay days:")]
    assert order == sorted(order)
    # Nothing wraps: each cell is one line in the extracted text.
    assert f"Payable days: {total_paid_days(result)}" in text


# --- Pay documents are released only after the wage group is finalized ---


def test_pay_documents_are_blocked_while_the_month_is_a_draft(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    for path in ("/reports/2026-07/final-report.pdf",
                 "/reports/2026-07/employee/5.pdf",
                 "/reports/2026-07/salary-sheet.pdf",
                 "/reports/2026-07/salary-sheet.xlsx"):
        response = client.get(path)
        assert response.status_code == 302, path
        assert not response.data.startswith(b"%PDF"), path
        followed = client.get(path, follow_redirects=True)
        assert b"once monthly wage payroll is finalized" in followed.data, path


def test_pay_documents_open_once_monthly_payroll_is_finalized(client, app):
    with app.app_context():
        seed_two_wage_groups()
    finalize_group(app, "2026-07", "MONTHLY")

    login(client)
    for path in ("/reports/2026-07/final-report.pdf",
                 "/reports/2026-07/employee/5.pdf",
                 "/reports/2026-07/salary-sheet.pdf"):
        response = client.get(path)
        assert response.status_code == 200, path
        assert response.data.startswith(b"%PDF"), path
    assert client.get("/reports/2026-07/salary-sheet.xlsx").status_code == 200


def test_a_draft_month_still_issues_the_daily_wage_attendance_summary(client, app):
    """Daily wage employees get an attendance summary, not a slip, so no pay figure
    leaves the app early and the document stays available through the draft month."""
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    for path in ("/reports/2026-07/employee/6.pdf",
                 "/reports/2026-07/summary-daily-wage.pdf",
                 "/reports/2026-07/attendance-summary-monthly.pdf",
                 "/reports/2026-07/department-wise.pdf"):
        response = client.get(path)
        assert response.status_code == 200, path
        assert response.data.startswith(b"%PDF"), path


def test_reports_page_locks_the_pay_document_cards_until_finalized(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    draft = client.get("/reports/").data
    # The cards stay listed, so the month's report set reads the same either way.
    assert b"Salary Slips (Monthly)" in draft and b"Salary Sheet (PF &amp; ESIC)" in draft
    assert b"/reports/2026-07/final-report.pdf" not in draft
    assert b"/reports/2026-07/salary-sheet.pdf" not in draft
    assert b"/reports/2026-07/salary-sheet.xlsx" not in draft
    assert draft.count(b"Available once monthly wage payroll is finalized.") == 2
    # Attendance reports are unaffected.
    assert b"/reports/2026-07/summary-daily-wage.pdf" in draft

    finalize_group(app, "2026-07", "MONTHLY")
    final = client.get("/reports/").data
    assert b"/reports/2026-07/final-report.pdf" in final
    assert b"/reports/2026-07/salary-sheet.xlsx" in final
    assert b"Available once monthly wage payroll is finalized." not in final


def test_employee_page_offers_the_slip_only_after_finalization(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    draft = client.get("/payroll/2026-07/employee/5").data
    assert b"Slip after finalize" in draft
    assert b"/reports/2026-07/employee/5.pdf" not in draft
    # The daily wage employee's attendance PDF is never gated.
    assert b"/reports/2026-07/employee/6.pdf" in client.get("/payroll/2026-07/employee/6").data

    finalize_group(app, "2026-07", "MONTHLY")
    final = client.get("/payroll/2026-07/employee/5").data
    assert b"/reports/2026-07/employee/5.pdf" in final
    assert b"Slip after finalize" not in final


# --- Salary slip history: a whole financial year in one file, for ITR filing ---

def seed_slip_history(employee_id="5", months=(("2026-02", True), ("2026-03", False), ("2026-07", True))):
    """One monthly wage employee paid across several months.

    Each month is finalized or left as a draft per the flag, so the history can be
    checked against what was actually issued.
    """
    db.session.add(Employee(id=employee_id, name="Month Worker", salary_type="Monthly",
                            normalized_salary_type="MONTHLY", salary=Decimal("30000"),
                            department="Design", designation="Design Manager"))
    db.session.add(WeekOffRule(employee_id=employee_id, confirmed_at=datetime.utcnow()))
    for month, finalized in months:
        db.session.add(PayrollMonth(month=month, monthly_finalized_at=datetime.utcnow() if finalized else None,
                                    status="FINALIZED" if finalized else "DRAFT"))
        db.session.add(SalaryRecord(payroll_month=month, employee_id=employee_id, name="Month Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        year, month_number = (int(part) for part in month.split("-"))
        for day in (1, 2):
            when = date(year, month_number, day)
            db.session.add(AttendanceRecord(payroll_month=month, employee_id=employee_id, employee_name="Month Worker",
                                            date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                            last_punch="06:30 PM", raw_working_hours="9h 00m",
                                            actual_minutes=540, parse_status="OK"))
    db.session.commit()
    for month, _ in months:
        calculate_payroll_month(month)


def test_slip_history_downloads_every_finalized_month_oldest_first(client, app):
    with app.app_context():
        seed_slip_history()

    login(client)
    response = client.get("/salary-slips/5.pdf")
    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    pages = PdfReader(BytesIO(response.data)).pages
    # One slip to a page, April-to-March reading order, and the draft month is absent.
    assert [page.extract_text().split("Pay Slip: ")[1].split("Days in")[0].strip() for page in pages] == [
        "February 2026", "July 2026",
    ]


def test_slip_history_can_be_narrowed_to_one_financial_year(client, app):
    with app.app_context():
        seed_slip_history(months=(("2026-02", True), ("2026-07", True)))

    login(client)
    # India's tax year runs April to March: February 2026 belongs to FY 2025-26 and
    # July 2026 to FY 2026-27, so the two months split across separate files.
    previous = client.get("/salary-slips/5.pdf?fy=2025")
    current = client.get("/salary-slips/5.pdf?fy=2026")
    assert "fy-2025-26" in previous.headers["Content-Disposition"]
    assert pdf_text(previous.data).count("Pay Slip") == 1
    assert "February 2026" in pdf_text(previous.data)
    assert "July 2026" in pdf_text(current.data)
    assert "February 2026" not in pdf_text(current.data)


def test_slip_history_lists_only_employees_who_were_issued_slips(client, app):
    with app.app_context():
        seed_two_wage_groups()
    finalize_group(app, "2026-07", "MONTHLY")

    login(client)
    page = client.get("/salary-slips").data
    assert b"Month Worker" in page
    # Daily wage employees are never issued a salary slip.
    assert b"Day Worker" not in page
    assert b"/salary-slips/5" in page


def test_slip_history_holds_back_a_month_that_is_still_a_draft(client, app):
    with app.app_context():
        seed_slip_history(months=(("2026-07", False),))

    login(client)
    # Nothing has been issued, so the employee is off the list entirely.
    assert b"/salary-slips/5\"" not in client.get("/salary-slips").data
    page = client.get("/salary-slips/5").data
    assert b"No salary slips have been issued to this employee yet." in page
    assert b"Not yet issued: July 2026" in page
    blocked = client.get("/salary-slips/5.pdf")
    assert blocked.status_code == 302
    assert "/salary-slips/5" in blocked.headers["Location"]


# --- July 2026 calculation audit ---

def test_short_hours_round_up_while_overtime_rounds_down(app):
    """The two are deliberately asymmetric: 48 minutes short costs 60; 9h42 pays to 9h30."""
    from attendance.payroll_rules import calculate_monthly_overtime, calculate_monthly_shortage
    # 8h12 is 48 minutes short of the 9h day.
    assert calculate_monthly_shortage(8 * 60 + 12) == 60
    # 8h47 is inside the 8h50 grace threshold by 3 minutes, but the band starts below
    # it, so 13 minutes short is charged as a full 15.
    assert calculate_monthly_shortage(8 * 60 + 47) == 15
    # A day at or above the grace threshold is charged nothing at all.
    assert calculate_monthly_shortage(8 * 60 + 50) == 0
    # 9h42 reaches the 9h30 trigger and pays 30 minutes over the 9h full day.
    assert calculate_monthly_overtime(9 * 60 + 42)[1] == 30
    assert calculate_monthly_overtime(9 * 60 + 22)[1] == 0


def test_overtime_is_paid_at_the_configured_multiple_of_the_ordinary_rate(app):
    """The company pays overtime at the ordinary rate, so the multiplier is 1."""
    from attendance.payroll_rules import calculate_monthly_overtime
    from attendance.settings import MONTHLY_RULES
    assert MONTHLY_RULES["OVERTIME_MULTIPLIER"] == 1
    # A quarter-hour rate of 100 pays 100 for each 15 minutes of overtime.
    _raw, rounded, amount = calculate_monthly_overtime(9 * 60 + 30, Decimal("100"))
    assert rounded == 30
    assert amount == Decimal("200")
    # And the multiplier is what carries the rate, not a constant in the maths.
    assert calculate_monthly_overtime(9 * 60 + 30, Decimal("100"), multiplier=2)[2] == Decimal("400")


def test_a_day_is_worth_the_month_divided_by_its_own_length(app):
    """The manual sheet pro-rates on calendar days, so July and February differ."""
    from attendance.payroll_rules import salary_days_for_month
    assert salary_days_for_month("2026-07") == 31
    assert salary_days_for_month("2026-02") == 28
    assert salary_days_for_month("2026-04") == 30


def test_esi_coverage_follows_the_wage_rate_not_a_short_month(app):
    """Unpaid leave must not enrol an employee who is above the ceiling."""
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="7", name="Above Ceiling", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("25000"),
                                basic_salary=Decimal("16250"), hra=Decimal("8750"),
                                esic_enabled=True))
        db.session.add(WeekOffRule(employee_id="7", confirmed_at=datetime.utcnow()))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="7", name="Above Ceiling",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("25000")))
        # A single working day: the month's earned wage lands far below the ceiling.
        when = date(2026, 7, 1)
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="7", employee_name="Above Ceiling",
                                        date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                        last_punch="06:30 PM", raw_working_hours="9h 00m",
                                        actual_minutes=540, parse_status="OK"))
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="7").one()
        assert Decimal(result.esi_employee) == Decimal("0")
        assert Decimal(result.esi_employer) == Decimal("0")


def test_professional_tax_is_not_deducted_from_daily_wage_employees(app):
    """PT stays a monthly wage deduction, as it is on the manual wage sheet."""
    with app.app_context():
        seed_two_wage_groups()
        # Well clear of the 12,000 slab, so only the wage type keeps PT away.
        daily = SalaryRecord.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        daily.salary = Decimal("7000")
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="6").one()
        assert Decimal(result.professional_tax or 0) == Decimal("0")
        # The monthly employee in the same month is unaffected by the exemption.
        monthly = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(monthly.professional_tax) == Decimal("200")


# --- The handwritten register applied in bulk ---

def seed_register_month(month="2026-07"):
    """One monthly employee with a punched day and a no-punch working day."""
    db.session.add(PayrollMonth(month=month))
    db.session.add(Employee(id="5", name="Field Staff", salary_type="Monthly",
                            normalized_salary_type="MONTHLY", salary=Decimal("30000")))
    db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
    db.session.add(SalaryRecord(payroll_month=month, employee_id="5", name="Field Staff",
                                salary_type="Monthly", normalized_salary_type="MONTHLY",
                                salary=Decimal("30000")))
    db.session.add(AttendanceRecord(payroll_month=month, employee_id="5", employee_name="Field Staff",
                                    date=date(2026, 7, 1), day="Wednesday", first_punch="09:30 AM",
                                    last_punch="06:30 PM", raw_working_hours="9h 00m",
                                    actual_minutes=540, parse_status="OK"))
    db.session.add(AttendanceRecord(payroll_month=month, employee_id="5", employee_name="Field Staff",
                                    date=date(2026, 7, 2), day="Thursday", parse_status="NEEDS_REVIEW",
                                    warning="Missing punch and working hours"))
    db.session.commit()


def test_register_export_can_be_narrowed_to_the_no_punch_days(client, app):
    with app.app_context():
        seed_register_month()

    login(client)
    everything = client.get("/attendance/2026-07/register.csv")
    assert everything.status_code == 200
    assert everything.data.decode().count("Field Staff") == 2

    missing = client.get("/attendance/2026-07/register.csv?scope=missing")
    body = missing.data.decode()
    # Only the day the punch machine could not see, which is what the register is for.
    assert "02-07-2026" in body
    assert "01-07-2026" not in body
    # The row says why the day needs the register, so the sheet can be filled in
    # without cross-checking the punch report.
    assert "Missing punch and working hours" in body


def test_missing_punch_register_groups_rows_by_issue_priority(client, app):
    with app.app_context():
        seed_register_month()
        db.session.add(Employee(id="6", name="Odd Staff", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(Employee(id="7", name="Other Staff", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.add(WeekOffRule(employee_id="6", confirmed_at=datetime.utcnow()))
        db.session.add(WeekOffRule(employee_id="7", confirmed_at=datetime.utcnow()))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="6", employee_name="Odd Staff",
                                        date=date(2026, 7, 3), day="Friday", first_punch="09:30 AM",
                                        raw_working_hours="", actual_minutes=None,
                                        parse_status="NEEDS_REVIEW", warning="Odd punch count"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="6", employee_name="Odd Staff",
                                        date=date(2026, 7, 6), day="Monday", first_punch="06:30 PM",
                                        raw_working_hours="", actual_minutes=None,
                                        parse_status="NEEDS_REVIEW", warning="Odd punch count"))
        db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="7", employee_name="Other Staff",
                                        date=date(2026, 7, 4), day="Saturday", first_punch="09:30 AM",
                                        last_punch="03:30 AM", raw_working_hours="", actual_minutes=None,
                                        parse_status="NEEDS_REVIEW",
                                        warning="Punch out before punch in (18h 00m session)"))
        db.session.commit()

    login(client)
    body = client.get("/attendance/2026-07/register.csv?scope=missing").data.decode()
    lines = body.splitlines()
    assert "Issue,Issue Count,System Status" in lines[0]
    headers = lines[0].split(",")
    issue_index = headers.index("Issue")
    count_index = headers.index("Issue Count")
    issues = [line.split(",")[issue_index] for line in lines[1:]]
    counts = [line.split(",")[count_index] for line in lines[1:]]
    assert issues == [
        "Odd punch count",
        "Odd punch count",
        "Missing punch and working hours",
        "Punch out before punch in (18h 00m session)",
    ]
    assert counts == ["2", "2", "1", "1"]


def test_register_import_applies_day_statuses_in_bulk(client, app):
    with app.app_context():
        seed_register_month()

    login(client)
    csv_body = (
        "Employee ID,Date,Register Status,Notes\n"
        "5,02-07-2026,Worked On-Site,Customer site\n"
    )
    response = client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(csv_body.encode()), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"1 day status(es) applied" in response.data

    with app.app_context():
        override = AttendanceOverride.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert override.manual_status == "Worked On-Site"
        assert override.notes == "Customer site"
        # And it reaches the payroll: the day is paid rather than absent.
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(result.paid_working_days) == Decimal("2")
        assert Decimal(result.lop_days) == Decimal("0")


def test_a_bad_register_row_is_reported_by_line_and_nothing_is_applied(client, app):
    with app.app_context():
        seed_register_month()

    login(client)
    csv_body = (
        "Employee ID,Date,Register Status,Notes\n"
        "5,01-07-2026,Paid Leave,\n"          # valid
        "5,02-07-2026,Working Hard,\n"        # not a status
        "9,02-07-2026,Paid Leave,\n"          # no such attendance row
    )
    response = client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(csv_body.encode()), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"2 row(s) could not be applied" in response.data
    assert b"Row 3" in response.data and b"Working Hard" in response.data
    assert b"Row 4" in response.data
    with app.app_context():
        # The valid row on line 2 is not applied either: a half-applied register
        # would leave the month in a state nobody can reason about.
        assert AttendanceOverride.query.count() == 0


def test_a_blank_register_status_leaves_the_day_alone_and_auto_clears_it(client, app):
    with app.app_context():
        seed_register_month()
        db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="5",
                                          date=date(2026, 7, 2), manual_status="Paid Leave"))
        db.session.commit()

    login(client)
    blank = "Employee ID,Date,Register Status\n5,02-07-2026,\n"
    client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(blank.encode()), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    with app.app_context():
        assert AttendanceOverride.query.count() == 1

    auto = "Employee ID,Date,Register Status\n5,02-07-2026,Auto\n"
    response = client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(auto.encode()), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"1 cleared" in response.data
    with app.app_context():
        assert AttendanceOverride.query.count() == 0


def test_a_half_covered_absence_is_topped_up_by_the_leave_earned_this_month(app):
    """Leave in hand must not sit unused while the employee loses pay.

    The opening balance runs out mid-day, leaving the absence half covered. Nothing
    used to look at that day again, so half a day went unpaid even though the month's
    own accrual could have covered it.
    """
    with app.app_context():
        by_date, result = seed_boundary_month("2026-11", 2026, 11, absent_days={9}, opening=Decimal("0.5"))

        assert by_date["2026-11-09"] == "Paid Leave"
        assert Decimal(result.lop_days) == Decimal("0")
        assert Decimal(result.leave_used) == Decimal("1")
        # The second half came out of this month's accrual, not out of thin air.
        assert Decimal(result.closing_leave) == (
            Decimal(result.opening_leave) + Decimal(result.leave_earned) - Decimal(result.leave_used))


def test_leave_top_up_does_not_conjure_a_balance_that_is_not_there(app):
    """With nothing accrued to spare, the half day stays unpaid."""
    with app.app_context():
        # Absent nearly the whole month: the accrual is tiny because it is earned on
        # paid days, so there is nothing to finish covering the split day with.
        by_date, result = seed_boundary_month(
            "2026-11", 2026, 11, absent_days=set(range(2, 28)), opening=Decimal("0.5"))
        assert Decimal(result.lop_days) > 0
        assert Decimal(result.leave_used) <= Decimal(result.opening_leave) + Decimal(result.leave_earned)


def test_an_older_export_still_reimports_after_the_sample_rows_were_renamed(client, app):
    """Files exported before the rename carry the old rows and must still round-trip."""
    with app.app_context():
        db.session.add(Employee(id="1", name="Manish C Hirani", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("68200")))
        db.session.commit()

    login(client)
    legacy = (b"Employee ID,Name,Department,Designation,Wage Type,Salary\n"
              b"1,John C Smith,Accounts,Accounts Executive,Monthly,50000\n"
              b"2,Elvis D Grey,Mechanical Production,Helper,Daily,5000\n"
              b"1,Manish C Hirani,Design,Design Manager,Monthly,68200\n")
    response = client.post("/master/import", data={
        "employee_master_csv": (BytesIO(legacy), "employee_master.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"already named" not in response.data
    with app.app_context():
        # The old sample rows are skipped, so employee 1 keeps their real name.
        assert db.session.get(Employee, "1").name == "Manish C Hirani"


def test_leave_covers_a_day_worked_below_the_half_day_minimum(app):
    """Turning up for an hour should not cost a day's pay when leave is in hand."""
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Short Day", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("31000")))
        db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
        db.session.add(PayrollResult(payroll_month="2026-06", employee_id="5", payroll_rule_type="MONTHLY",
                                     calculation_status="Calculated", closing_leave=Decimal("2"),
                                     final_salary=Decimal("31000")))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Short Day",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("31000")))
        for day, minutes in ((1, 540), (2, 75)):
            when = date(2026, 7, day)
            db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Short Day",
                                            date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                            last_punch="10:45 AM" if minutes == 75 else "06:30 PM",
                                            raw_working_hours="1h 15m" if minutes == 75 else "9h 00m",
                                            actual_minutes=minutes, parse_status="OK"))
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row for row in result.detail_json}
        assert by_date["2026-07-02"]["attendance_status"] == "Paid Leave"
        assert "under the half-day minimum" in by_date["2026-07-02"]["explanation"]
        assert Decimal(result.lop_days) == Decimal("0")
        assert Decimal(result.leave_used) == Decimal("1")


def test_a_manual_unpaid_leave_override_is_never_covered_by_leave(app):
    """An explicit instruction to dock the day must stand, balance or no balance."""
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Docked", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("31000")))
        db.session.add(WeekOffRule(employee_id="5", confirmed_at=datetime.utcnow()))
        db.session.add(PayrollResult(payroll_month="2026-06", employee_id="5", payroll_rule_type="MONTHLY",
                                     calculation_status="Calculated", closing_leave=Decimal("5"),
                                     final_salary=Decimal("31000")))
        db.session.add(SalaryRecord(payroll_month="2026-07", employee_id="5", name="Docked",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("31000")))
        for day in (1, 2):
            when = date(2026, 7, day)
            db.session.add(AttendanceRecord(payroll_month="2026-07", employee_id="5", employee_name="Docked",
                                            date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                            last_punch="06:30 PM", raw_working_hours="9h 00m",
                                            actual_minutes=540, parse_status="OK"))
        db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="5", date=date(2026, 7, 2),
                                          manual_status="Unpaid Leave / LOP"))
        db.session.commit()
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        by_date = {row["date"]: row["attendance_status"] for row in result.detail_json}
        assert by_date["2026-07-02"] == "Unpaid Leave / LOP"
        assert Decimal(result.leave_used) == Decimal("0")


def test_leave_accrual_matches_the_documented_formula(app):
    """Accrual is 2 days x paid days / days in month, on the FINAL paid-day count.

    Settling it in one pass computed the figure from a paid-day count that predated
    the leave it granted, so it read low for anyone who spent leave and high for
    anyone whose leave was later withdrawn for want of a balance.
    """
    with app.app_context():
        seed_leave_month(opening=Decimal("2"))
        add_july_attendance({3, 10, 14, 22})
        calculate_payroll_month("2026-07")
        result = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()

        paid = (Decimal(result.paid_working_days) + Decimal(result.week_offs)
                + Decimal(result.paid_leaves) + Decimal(result.holidays or 0))
        expected = (paid / Decimal(31) * Decimal(2)).quantize(Decimal("0.01"), rounding=ROUND_DOWN)
        assert Decimal(result.leave_earned) == expected
        # And the accrual is settled, not a snapshot: recalculating changes nothing.
        calculate_payroll_month("2026-07")
        again = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(again.leave_earned) == Decimal(result.leave_earned)
        assert Decimal(again.leave_used) == Decimal(result.leave_used)


def test_accrual_is_not_credited_for_leave_that_was_withdrawn(app):
    """A sandwich day with no balance behind it is loss of pay, so it earns nothing."""
    with app.app_context():
        by_date, result = seed_boundary_month("2026-11", 2026, 11, absent_days={7, 9})
        paid = (Decimal(result.paid_working_days) + Decimal(result.week_offs)
                + Decimal(result.paid_leaves) + Decimal(result.holidays or 0))
        expected = (paid / Decimal(30) * Decimal(2)).quantize(Decimal("0.01"), rounding=ROUND_DOWN)
        assert Decimal(result.leave_earned) == expected
        assert Decimal(result.lop_days) > 0


# --- V0.08 audit ---

def test_a_register_import_clears_the_figures_it_invalidates(client, app):
    """Stale pay must not survive a change to which days are payable.

    Editing punches clears the month's results; importing the register writes the
    same day statuses, so it has to do the same. Otherwise the payroll page keeps
    showing figures that no longer follow from the attendance, and the month can be
    finalized on them.
    """
    with app.app_context():
        seed_register_month()
        calculate_payroll_month("2026-07")
        month = db.session.get(PayrollMonth, "2026-07")
        month.attendance_submitted = True
        db.session.commit()
        before = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(before.paid_working_days) == Decimal("1")

    login(client)
    csv_body = "Employee ID,Date,Register Status\n5,02-07-2026,Worked On-Site\n"
    response = client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(csv_body.encode()), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"calculated result(s) cleared" in response.data

    with app.app_context():
        assert PayrollResult.query.filter_by(payroll_month="2026-07").count() == 0
        assert db.session.get(PayrollMonth, "2026-07").attendance_submitted is False
        # Recalculating picks the day up, so nothing was lost by clearing it.
        calculate_payroll_month("2026-07")
        after = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert Decimal(after.paid_working_days) == Decimal("2")


def test_a_register_import_that_changes_nothing_leaves_the_figures_alone(client, app):
    """A blank register must not throw away a calculated month for no reason."""
    with app.app_context():
        seed_register_month()
        calculate_payroll_month("2026-07")

    login(client)
    csv_body = "Employee ID,Date,Register Status\n5,02-07-2026,\n"
    response = client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(csv_body.encode()), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"calculated result(s) cleared" not in response.data
    with app.app_context():
        assert PayrollResult.query.filter_by(payroll_month="2026-07").count() == 1


def test_a_leaver_is_still_paid_for_the_month_they_worked(client, app):
    """Marking someone Left must not delete the pay they are owed for that month."""
    with app.app_context():
        seed_two_wage_groups()
        calculate_payroll_month("2026-07")
        assert PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").count() == 1

    login(client)
    client.post("/master", data={
        "action": "disable", "employee_id": "5", "employment_status": "LEFT",
        "inactive_reason": "resigned", "disable_confirmation": "confirm",
        "left_on": "31-07-2026",
    }, follow_redirects=True)

    with app.app_context():
        calculate_payroll_month("2026-07")
        july = PayrollResult.query.filter_by(payroll_month="2026-07", employee_id="5").first()
        assert july is not None, "the month they left in must still be paid"
        assert Decimal(july.final_salary) > 0
        # And they are gone from the month after, without anyone having to remember.
        db.session.add(PayrollMonth(month="2026-08"))
        db.session.add(SalaryRecord(payroll_month="2026-08", employee_id="5", name="Month Worker",
                                    salary_type="Monthly", normalized_salary_type="MONTHLY",
                                    salary=Decimal("30000")))
        when = date(2026, 8, 3)
        db.session.add(AttendanceRecord(payroll_month="2026-08", employee_id="5", employee_name="Month Worker",
                                        date=when, day=when.strftime("%A"), first_punch="09:30 AM",
                                        last_punch="06:30 PM", raw_working_hours="9h 00m",
                                        actual_minutes=540, parse_status="OK"))
        db.session.commit()
        calculate_payroll_month("2026-08")
        assert PayrollResult.query.filter_by(payroll_month="2026-08", employee_id="5").count() == 0


def test_an_employee_left_out_of_a_month_is_named_on_the_payroll_page(client, app):
    """A person can never drop out of a run unnoticed."""
    with app.app_context():
        seed_two_wage_groups()
        employee = db.session.get(Employee, "5")
        employee.employment_status = "TERMINATED"   # no last working day, as on older data
        db.session.commit()

    login(client)
    page = client.get("/payroll/2026-07").data
    assert b"not included in payroll" in page
    assert b"no last working day recorded" in page
    assert b"Month Worker" in page


def test_the_last_working_day_round_trips_through_the_master_export(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    client.post("/master", data={
        "action": "disable", "employee_id": "5", "employment_status": "LEFT",
        "inactive_reason": "resigned", "disable_confirmation": "confirm",
        "left_on": "15-07-2026",
    }, follow_redirects=True)
    export = client.get("/master/export.csv").data
    assert b"LEFT,15-07-2026" in export

    response = import_master(client, export)
    assert b"imported" in response.data
    with app.app_context():
        employee = db.session.get(Employee, "5")
        assert employee.employment_status == "LEFT"
        assert employee.left_on == date(2026, 7, 15)


def test_marking_a_leaver_by_import_requires_the_last_working_day(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    csv_body = ("Employee ID,Name,Wage Type,Salary,Status,Last Working Day\n"
                "5,Month Worker,Monthly,30000,LEFT,\n").encode()
    response = import_master(client, csv_body)
    assert b"Last Working Day is required" in response.data
    with app.app_context():
        assert db.session.get(Employee, "5").employment_status == "ACTIVE"


def test_a_new_month_cannot_start_while_an_earlier_one_is_open(client, app):
    """A month opens on the previous month's closing balance, so that month is settled first."""
    with app.app_context():
        seed_two_wage_groups()          # creates 2026-07, left as a draft
        calculate_payroll_month("2026-07")

    login(client)
    blocked = client.post("/payroll/new", data={"month": "2026-08"}, follow_redirects=True)
    assert b"Finalize July 2026 before starting August 2026" in blocked.data
    with app.app_context():
        assert db.session.get(PayrollMonth, "2026-08") is None
    # And the page says so before anyone picks a month.
    assert b"Finish July 2026 first" in client.get("/payroll/new").data

    finalize_group(app, "2026-07", "MONTHLY")
    finalize_group(app, "2026-07", "DAILY")
    allowed = client.post("/payroll/new", data={"month": "2026-08"}, follow_redirects=True)
    assert b"before starting August 2026" not in allowed.data
    with app.app_context():
        assert db.session.get(PayrollMonth, "2026-08") is not None


def test_the_first_month_in_the_system_starts_freely(client, app):
    """There is no previous balance to settle, so nothing is in the way."""
    login(client)
    response = client.post("/payroll/new", data={"month": "2026-07"}, follow_redirects=True)
    assert b"before starting July 2026" not in response.data
    with app.app_context():
        assert db.session.get(PayrollMonth, "2026-07") is not None


def test_reopening_an_existing_month_is_never_blocked(client, app):
    """The gate is about starting a new month, not about revisiting one."""
    with app.app_context():
        seed_two_wage_groups()
        db.session.add(PayrollMonth(month="2026-08"))
        db.session.commit()

    login(client)
    response = client.post("/payroll/new", data={"month": "2026-08"}, follow_redirects=True)
    assert b"Finalize July 2026" not in response.data


def test_a_master_export_carrying_a_leaver_can_be_imported_back(client, app):
    """The export is the backup path, so it has to survive a round trip.

    Anyone marked as having left before the last working day existed carries a blank
    date, and demanding one on the way back in rejected the whole file.
    """
    with app.app_context():
        seed_two_wage_groups()
        employee = db.session.get(Employee, "5")
        employee.employment_status = "TERMINATED"     # no date, as on data from before
        db.session.commit()

    login(client)
    export = client.get("/master/export.csv").data
    assert b"TERMINATED," in export
    response = import_master(client, export)
    assert b"Last Working Day is required" not in response.data
    assert b"imported" in response.data
    with app.app_context():
        assert Employee.query.count() == 2
        assert db.session.get(Employee, "5").employment_status == "TERMINATED"


def test_a_punch_report_saved_by_a_spreadsheet_tool_still_imports(client, app):
    """Both legal ways of naming the worksheet have to work.

    A punch machine writes the sheet path relative to the workbook; Excel and most
    libraries write it absolute from the package root. A file that has been opened
    and re-saved used to fail with a raw archive error.
    """
    from openpyxl import Workbook
    with app.app_context():
        db.session.add(PayrollMonth(month="2026-07"))
        db.session.add(Employee(id="5", name="Worker", salary_type="Monthly",
                                normalized_salary_type="MONTHLY", salary=Decimal("30000")))
        db.session.commit()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Daily Punch Report"])
    sheet.append(["Employee ID", "Employee Name", "Department", "Designation", "01-07-2026\nWednesday"])
    sheet.append(["5", "Worker", "Design", "Manager", "09:30 AM\n06:30 PM"])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    # openpyxl writes Target="/xl/worksheets/sheet1.xml", the absolute form.
    from zipfile import ZipFile
    with ZipFile(BytesIO(buffer.getvalue())) as archive:
        assert b'Target="/xl/worksheets/sheet1.xml"' in archive.read("xl/_rels/workbook.xml.rels")

    login(client)
    response = client.post("/attendance/2026-07", data={
        "action": "import_attendance",
        "attendance_csv": (BytesIO(buffer.getvalue()), "punch.xlsx"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"Imported attendance: 1 rows" in response.data
    with app.app_context():
        assert AttendanceRecord.query.filter_by(payroll_month="2026-07", employee_id="5").count() == 1


def test_a_new_employees_week_off_waits_to_be_confirmed(client, app):
    """The app asks for week offs to be checked, so the system must not answer for it."""
    login(client)
    csv_body = ("Employee ID,Name,Wage Type,Salary\n"
                "5,Month Worker,Monthly,30000\n"
                "6,Day Worker,Daily,600\n").encode()
    assert b"imported" in import_master(client, csv_body).data
    with app.app_context():
        rules = WeekOffRule.query.all()
        assert len(rules) == 2 and all(rule.confirmed_at is None for rule in rules)
        assert all(rule.sunday == "WEEK_OFF_ALL" for rule in rules)

    page = client.get("/weekoffs").data.decode()
    assert '5_present' in page, "each row must mark itself as carried by the form"
    client.post("/weekoffs", data={"5_present": "1", "5_sunday": "WEEK_OFF_ALL"}, follow_redirects=True)
    with app.app_context():
        assert WeekOffRule.query.filter_by(employee_id="5").one().confirmed_at is not None
        # The one nobody looked at is still waiting.
        assert WeekOffRule.query.filter_by(employee_id="6").one().confirmed_at is None


def test_the_week_off_pattern_round_trips_through_the_master_export(client, app):
    """A rebuild from the export has to restore who works which days."""
    with app.app_context():
        seed_two_wage_groups()
        rule = WeekOffRule.query.filter_by(employee_id="5").one()
        rule.saturday = "WEEK_OFF_2,WEEK_OFF_4"
        rule.sunday = "WEEK_OFF_ALL"
        db.session.commit()

    login(client)
    export = client.get("/master/export.csv").data
    assert b"Saturday=2,4; Sunday=All" in export

    with app.app_context():
        # Wipe the pattern, as a restore into an empty system would find it.
        rule = WeekOffRule.query.filter_by(employee_id="5").one()
        rule.saturday = rule.sunday = "WORKING"
        db.session.commit()

    assert b"imported" in import_master(client, export).data
    with app.app_context():
        rule = WeekOffRule.query.filter_by(employee_id="5").one()
        assert rule.saturday == "WEEK_OFF_2,WEEK_OFF_4"
        assert rule.sunday == "WEEK_OFF_ALL"
        # A pattern someone put in the file is a decision, so it counts as confirmed.
        assert rule.confirmed_at is not None


def test_a_nonsense_week_off_pattern_is_rejected_by_line(client, app):
    with app.app_context():
        seed_two_wage_groups()

    login(client)
    csv_body = ("Employee ID,Name,Wage Type,Salary,Week Off Pattern\n"
                "5,Month Worker,Monthly,30000,Funday=All\n").encode()
    response = import_master(client, csv_body)
    assert b"Row 2: Week Off Pattern" in response.data
    with app.app_context():
        assert WeekOffRule.query.filter_by(employee_id="5").one().sunday == "WEEK_OFF_ALL"


def test_manual_day_overrides_travel_in_the_register_export(client, app):
    """The register file is how a day status survives a rebuild."""
    with app.app_context():
        seed_register_month()
        db.session.add(AttendanceOverride(payroll_month="2026-07", employee_id="5",
                                          date=date(2026, 7, 2), manual_status="Worked On-Site",
                                          notes="Customer site"))
        db.session.commit()

    login(client)
    export = client.get("/attendance/2026-07/register.csv").data
    assert b"Worked On-Site,Customer site" in export

    with app.app_context():
        AttendanceOverride.query.delete()
        db.session.commit()

    response = client.post("/attendance/2026-07", data={
        "action": "import_register",
        "register_csv": (BytesIO(export), "register.csv"),
    }, content_type="multipart/form-data", follow_redirects=True)
    assert b"day status(es) applied" in response.data
    with app.app_context():
        restored = AttendanceOverride.query.filter_by(payroll_month="2026-07", employee_id="5").one()
        assert restored.manual_status == "Worked On-Site"
        assert restored.notes == "Customer site"
