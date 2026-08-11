import calendar
import csv
import re
from collections import defaultdict
from datetime import date
from io import BytesIO, StringIO
from decimal import Decimal
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from attendance import db
from attendance.loans import loan_installment_for_loan, loan_paid_before_month, loan_pending_after_month, loan_remaining_before_month, loan_repayment_schedule
from attendance.models import AttendanceRecord, Employee, Loan, PayrollResult, SalaryRecord
from attendance.settings import COMPANY_ADDRESS
from attendance.statutory import PROFESSIONAL_TAX_SLABS, STATUTORY_RULES
from attendance.utils import format_percent, leave_days, minutes_to_duration, money

ONES = [
    "zero", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
    "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen",
    "seventeen", "eighteen", "nineteen",
]
TENS = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]

# Print palette mirrors the iOS system colours used by static/css/app.css, using the
# text-safe variants so the paper output keeps the same contrast as the screen.
INK = colors.HexColor("#1C1C1E")           # label
MUTED = colors.HexColor("#6B6B70")         # secondaryLabel, flattened for print
FAINT = colors.HexColor("#8E8E93")         # tertiaryLabel
SEPARATOR = colors.HexColor("#D8D8DC")     # hairline
SURFACE = colors.white
SURFACE_SOFT = colors.HexColor("#F2F2F7")  # secondarySystemBackground
ZEBRA = colors.HexColor("#FAFAFC")
TINT = colors.HexColor("#0069DE")          # system blue (fill)
TINT_TEXT = colors.HexColor("#0050C8")     # system blue (text-safe)
TINT_WASH = colors.HexColor("#EBF2FD")
GREEN_TEXT = colors.HexColor("#1B7032")
RED_TEXT = colors.HexColor("#B3000C")
ORANGE_TEXT = colors.HexColor("#9A4A00")
GREEN_WASH = colors.HexColor("#E8F7EC")
ORANGE_WASH = colors.HexColor("#FDF0E3")
RED_WASH = colors.HexColor("#FCEBEC")
TEAL_TEXT = colors.HexColor("#0B6B7A")
TEAL_WASH = colors.HexColor("#E4F5F8")

# Kept for the brand lockup only; the report chrome itself is system-coloured.
BRAND_BLUE = colors.HexColor("#0C306A")
BRAND_GREEN = colors.HexColor("#A6CE15")
BRAND_MUTED = MUTED
CARD_RADIUS = 7
LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "img" / "smartfill-logo.png"


def _status_colours(status):
    """Wash and text colour for a calculation or attendance status."""
    text = str(status or "").strip().lower()
    if text in {"calculated", "full day present", "full day", "week off worked", "paid"}:
        return GREEN_WASH, GREEN_TEXT
    if text in {"needs review", "punch error", "half day present", "half day", "pending"}:
        return ORANGE_WASH, ORANGE_TEXT
    if "lop" in text or text in {"absent / attendance missing", "attendance missing", "skipped"}:
        return RED_WASH, RED_TEXT
    return SURFACE_SOFT, MUTED


def payroll_month_days(month):
    year, month_number = (int(part) for part in month.split("-"))
    return calendar.monthrange(year, month_number)[1]


def total_paid_days(result):
    if not result:
        return ""
    if result.payroll_rule_type == "DAILY":
        return Decimal(result.paid_working_days or 0) + Decimal(result.holidays or 0)
    return Decimal(result.paid_working_days or 0) + Decimal(result.week_offs or 0) + Decimal(result.paid_leaves or 0)


def result_has_loan(result):
    if not result:
        return False
    return bool(
        Decimal(getattr(result, "loan_deduction", 0) or 0) != 0
        or Decimal(getattr(result, "loan_pending_amount", 0) or 0) != 0
    )


def result_has_advance(result):
    if not result:
        return False
    return Decimal(getattr(result, "advance_deduction", 0) or 0) != 0


def daily_result_calculated(result):
    return bool(result and result.final_salary is not None and result.payroll_rule_type == "DAILY")


def payroll_summary_csv(month):
    out = StringIO()
    writer = csv.writer(out)
    month_days = payroll_month_days(month)
    writer.writerow([
        "Employee ID", "Name", "Department", "Designation", "Wage Type", "Payroll Rule Status", "Salary",
        "Month Days", "Paid Working Days", "Week Offs", "Total Paid Days", "Full Days", "Half Days", "Paid Leave", "LOP", "Opening Leave",
        "Leave Earned", "Leave Used", "Closing Leave", "Working Hour Deduction",
        "LOP Deduction", "Overtime", "Absence Minutes", "Attendance Bonus %", "Attendance Bonus",
        "Adjustment", "Leave Encashment Days", "Leave Encashment Amount", "Loan Deduction", "Advance Salary Deduction", "Final Salary", "Calculation Status",
    ])
    salaries = {s.employee_id: s for s in SalaryRecord.query.filter_by(payroll_month=month).all()}
    results = {r.employee_id: r for r in PayrollResult.query.filter_by(payroll_month=month).all()}
    # The Department column existed but was always blank; employee master now holds it.
    employees = {e.id: e for e in Employee.query.all()}
    for employee_id, salary in salaries.items():
        result = results.get(employee_id)
        employee = employees.get(employee_id)
        writer.writerow([
            employee_id,
            salary.name,
            (employee.department if employee else "") or "",
            (employee.designation if employee else "") or "",
            salary.salary_type,
            result.calculation_status if result else "Not Calculated",
            salary.salary,
            month_days if result and result.final_salary is not None else "",
            result.paid_working_days if result and result.final_salary is not None else "",
            result.week_offs if result and result.final_salary is not None else "",
            total_paid_days(result) if result and result.final_salary is not None else "",
            result.full_days if result and result.final_salary is not None else "",
            result.half_days if result and result.final_salary is not None else "",
            result.paid_leaves if result and result.final_salary is not None else "",
            result.lop_days if result and result.final_salary is not None else "",
            result.opening_leave if result and result.final_salary is not None else "",
            result.leave_earned if result and result.final_salary is not None else "",
            result.leave_used if result and result.final_salary is not None else "",
            result.closing_leave if result and result.final_salary is not None else "",
            result.less_hours_deduction if result and result.final_salary is not None else "",
            result.lop_deduction if result and result.final_salary is not None else "",
            result.ot_amount if result and result.final_salary is not None else "",
            # The attendance bonus only exists for daily wage, so it stays blank elsewhere.
            getattr(result, "absence_minutes", 0) if daily_result_calculated(result) else "",
            getattr(result, "attendance_bonus_percent", 0) if daily_result_calculated(result) else "",
            getattr(result, "attendance_bonus_amount", 0) if daily_result_calculated(result) else "",
            salary.adjustment,
            result.leave_encashment_days if result and result.final_salary is not None else (salary.leave_encashment_days if getattr(salary, "leave_encashment_enabled", False) else ""),
            result.leave_encashment_amount if result and result.final_salary is not None else (salary.leave_encashment_amount if getattr(salary, "leave_encashment_enabled", False) else ""),
            result.loan_deduction if result and result.final_salary is not None else getattr(salary, "loan", ""),
            result.advance_deduction if result and result.final_salary is not None else "",
            result.final_salary if result and result.final_salary is not None else "",
            result.calculation_status if result else "Payroll Rules Not Configured",
        ])
    return out.getvalue()


def attendance_detail_csv(month):
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["Employee ID", "Date", "Day", "First Punch", "Last Punch", "Raw Working Hours", "Actual Duration", "Status", "Warning"])
    for rec in AttendanceRecord.query.filter_by(payroll_month=month).order_by(AttendanceRecord.employee_id, AttendanceRecord.date):
        writer.writerow([rec.employee_id, rec.date, rec.day, rec.first_punch, rec.last_punch, rec.raw_working_hours, minutes_to_duration(rec.actual_minutes), rec.parse_status, rec.warning])
    return out.getvalue()


def employee_name_map(month):
    salary_names = {salary.employee_id: salary.name for salary in SalaryRecord.query.filter_by(payroll_month=month).all()}
    employee_names = {employee.id: employee.name for employee in Employee.query.all()}
    employee_names.update(salary_names)
    return employee_names


def overtime_report_csv(month):
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["Employee ID", "Employee Name", "Date", "In Time", "Out Time", "Working Hours", "OT Paid Minutes", "OT Paid Amount"])
    names = employee_name_map(month)
    for result in PayrollResult.query.filter_by(payroll_month=month).order_by(PayrollResult.employee_id):
        for item in result.detail_json or []:
            payable_ot = int(item.get("payable_ot") or 0)
            ot_amount = Decimal(str(item.get("ot_amount") or "0"))
            if payable_ot <= 0 and ot_amount <= 0:
                continue
            writer.writerow([
                result.employee_id,
                names.get(result.employee_id, ""),
                item.get("date", ""),
                item.get("first_punch", ""),
                item.get("last_punch", ""),
                item.get("actual_duration") or item.get("raw_working_hours", ""),
                payable_ot,
                ot_amount,
            ])
    return out.getvalue()


def less_hours_report_csv(month):
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["Employee ID", "Employee Name", "Date", "In Time", "Out Time", "Working Hours", "Less Hours Minutes", "Less Hours Deduction"])
    names = employee_name_map(month)
    for result in PayrollResult.query.filter_by(payroll_month=month).order_by(PayrollResult.employee_id):
        for item in result.detail_json or []:
            shortage = int(item.get("shortage_minutes") or 0)
            deduction = Decimal(str(item.get("shortage_deduction") or "0"))
            if shortage <= 0 and deduction <= 0:
                continue
            writer.writerow([
                result.employee_id,
                names.get(result.employee_id, ""),
                item.get("date", ""),
                item.get("first_punch", ""),
                item.get("last_punch", ""),
                item.get("rounded_duration") or item.get("actual_duration") or item.get("raw_working_hours", ""),
                shortage,
                deduction,
            ])
    return out.getvalue()


def error_report_csv(month):
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["Area", "Employee ID", "Issue"])
    for salary in SalaryRecord.query.filter_by(payroll_month=month):
        if salary.warning:
            writer.writerow(["Salary", salary.employee_id, salary.warning])
    for rec in AttendanceRecord.query.filter_by(payroll_month=month):
        if rec.warning:
            writer.writerow(["Attendance", rec.employee_id, f"{rec.date}: {rec.warning}"])
    for result in PayrollResult.query.filter_by(payroll_month=month):
        if result.calculation_status != "Calculated":
            writer.writerow(["Payroll", result.employee_id, result.message or result.calculation_status])
    return out.getvalue()


def display_month(month):
    try:
        year, month_number = (int(part) for part in month.split("-"))
    except (AttributeError, ValueError):
        return month
    return f"{calendar.month_name[month_number]} {year}"


def calculated_results_for_month(month):
    return [
        result for result in PayrollResult.query.filter_by(payroll_month=month).order_by(PayrollResult.employee_id)
        if result.final_salary is not None
    ]


def report_pdf(title, subtitle, headers, rows, col_widths=None, font_size=7, landscape_page=True, kpis=None, status_column=None):
    buffer = BytesIO()
    pagesize = landscape(A4) if landscape_page else A4
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=11 * mm, rightMargin=11 * mm, topMargin=11 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("ReportCell", parent=styles["Normal"], fontSize=font_size, leading=font_size + 2.2, textColor=INK)
    header_style = ParagraphStyle(
        "ReportHeaderCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        fontSize=font_size - 0.4,
        textColor=MUTED,
    )
    empty_style = ParagraphStyle("ReportEmpty", parent=cell_style, textColor=FAINT)
    available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
    if not col_widths:
        col_widths = [available_width / len(headers)] * len(headers)

    table_rows = [[Paragraph(str(value).upper(), header_style) for value in headers]]
    if rows:
        for row in rows:
            table_rows.append([Paragraph(str(value if value not in (None, "") else "—"), cell_style) for value in row])
    else:
        table_rows.append([Paragraph("No records found", empty_style)] + [Paragraph("", cell_style)] * (len(headers) - 1))

    story = [_report_brand_header(title, subtitle, styles, available_width), Spacer(1, 9)]
    if kpis:
        story.extend([_kpi_row(kpis, available_width), Spacer(1, 9)])
    story.append(_table(table_rows, col_widths=col_widths, font_size=font_size, status_column=status_column if rows else None))
    page_callback = _titled_page_callback(title)
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


PAYROLL_SUMMARY_HEADERS = [
    "ID", "Employee", "Designation", "Wage Type", "Base", "Days", "Working", "Week Off",
    "Total Paid", "Leave", "LOP", "Deduction", "Addition", "Payable", "Status",
]
PAYROLL_SUMMARY_WIDTHS = [
    11 * mm, 30 * mm, 28 * mm, 16 * mm, 18 * mm, 11 * mm, 15 * mm, 15 * mm,
    16 * mm, 12 * mm, 11 * mm, 20 * mm, 18 * mm, 21 * mm, 21 * mm,
]


def payroll_summary_rows(month, wage_group):
    """Summary rows for one wage group, with the group's payable and deduction totals."""
    names = employee_name_map(month)
    designations = {employee.id: employee.designation or "" for employee in Employee.query.all()}
    salaries = {s.employee_id: s for s in SalaryRecord.query.filter_by(payroll_month=month).all()}
    rows = []
    total_payable = Decimal("0")
    total_deduction = Decimal("0")
    # Sorted numerically: a plain string sort listed 11 and 13 before 2.
    for result in sorted(calculated_results_for_month(month), key=lambda r: employee_id_sort_key(r.employee_id)):
        salary = salaries.get(result.employee_id)
        if not salary or salary.normalized_salary_type != wage_group:
            continue
        total_payable += Decimal(result.final_salary or 0)
        total_deduction += Decimal(result.total_deduction or 0)
        rows.append([
            result.employee_id,
            names.get(result.employee_id, result.employee_id),
            designations.get(result.employee_id, ""),
            salary.salary_type or "",
            pdf_money(salary.salary),
            payroll_month_days(month),
            result.paid_working_days,
            result.week_offs,
            total_paid_days(result),
            result.paid_leaves,
            result.lop_days,
            pdf_money(result.total_deduction),
            pdf_money(result.total_addition),
            pdf_money(result.final_salary),
            result.calculation_status,
        ])
    return rows, total_payable, total_deduction


def build_payroll_summary_pdf(month):
    """Payroll summary in two parts in one file: monthly first, then daily.

    Each wage group starts on its own page so a section can be printed or handed
    over on its own without carrying rows from the other group.
    """
    buffer = BytesIO()
    pagesize = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=11 * mm, rightMargin=11 * mm,
                            topMargin=11 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
    cell_style = ParagraphStyle("SummaryCell", parent=styles["Normal"], fontSize=6.2, leading=8.4, textColor=INK)
    header_style = ParagraphStyle("SummaryHead", parent=cell_style, fontName="Helvetica-Bold", fontSize=5.8, textColor=MUTED)
    empty_style = ParagraphStyle("SummaryEmpty", parent=cell_style, textColor=FAINT)

    story = []
    for index, (wage_group, label) in enumerate((("MONTHLY", "Monthly Wage"), ("DAILY", "Daily Wage"))):
        rows, total_payable, total_deduction = payroll_summary_rows(month, wage_group)
        if index:
            story.append(PageBreak())
        story.extend([
            _report_brand_header(f"Payroll Summary - {label}", display_month(month), styles, available_width),
            Spacer(1, 9),
            _kpi_row([
                ("Employees", len(rows)),
                ("Total payable", pdf_money(total_payable)),
                ("Total deductions", pdf_money(total_deduction)),
                ("Payroll month", display_month(month)),
            ], available_width),
            Spacer(1, 9),
        ])
        table_rows = [[Paragraph(str(value).upper(), header_style) for value in PAYROLL_SUMMARY_HEADERS]]
        for row in rows:
            table_rows.append([Paragraph(str(value if value not in (None, "") else "—"), cell_style) for value in row])
        if not rows:
            table_rows.append([Paragraph(f"No {label.lower()} employees found", empty_style)]
                              + [Paragraph("", cell_style)] * (len(PAYROLL_SUMMARY_HEADERS) - 1))
        story.append(_table(table_rows, col_widths=PAYROLL_SUMMARY_WIDTHS, font_size=6.2,
                            status_column=14 if rows else None))
    page_callback = _titled_page_callback(f"Payroll Summary - {display_month(month)}")
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


def build_attendance_detail_pdf(month):
    rows = []
    names = employee_name_map(month)
    for rec in AttendanceRecord.query.filter_by(payroll_month=month).order_by(AttendanceRecord.employee_id, AttendanceRecord.date):
        rows.append([
            rec.employee_id,
            names.get(rec.employee_id, rec.employee_name or rec.employee_id),
            rec.date,
            rec.day,
            rec.first_punch or "-",
            rec.last_punch or "-",
            minutes_to_duration(rec.actual_minutes),
            rec.parse_status,
            rec.warning or "",
        ])
    review_rows = sum(1 for row in rows if row[7] != "OK")
    return report_pdf(
        "Detailed Attendance",
        f"{display_month(month)} · Every imported punch day",
        ["ID", "Employee", "Date", "Day", "In", "Out", "Working Hours", "Status", "Warning"],
        rows,
        col_widths=[13 * mm, 42 * mm, 22 * mm, 22 * mm, 20 * mm, 20 * mm, 25 * mm, 24 * mm, 86 * mm],
        font_size=6.2,
        kpis=[
            ("Attendance rows", f"{len(rows):,}"),
            ("Employees", len({row[0] for row in rows})),
            ("Needs review", f"{review_rows:,}"),
            ("Payroll month", display_month(month)),
        ],
    )


def build_overtime_report_pdf(month):
    rows = []
    names = employee_name_map(month)
    # Totals come from the stored payroll figures, which are rounded once. Adding up
    # the per-day amounts shown in the table would drift from what payroll actually
    # paid, because each of those is rounded for display.
    total_minutes = 0
    total_amount = Decimal("0")
    for result in PayrollResult.query.filter_by(payroll_month=month).order_by(PayrollResult.employee_id):
        contributed = False
        for item in result.detail_json or []:
            payable_ot = int(item.get("payable_ot") or 0)
            ot_amount = Decimal(str(item.get("ot_amount") or "0"))
            if payable_ot <= 0 and ot_amount <= 0:
                continue
            rows.append([
                result.employee_id,
                names.get(result.employee_id, ""),
                item.get("date", ""),
                item.get("first_punch", "") or "-",
                item.get("last_punch", "") or "-",
                item.get("actual_duration") or item.get("raw_working_hours", ""),
                payable_ot,
                pdf_money(ot_amount),
            ])
            contributed = True
        if contributed:
            total_minutes += int(result.payable_ot_minutes or 0)
            total_amount += Decimal(result.ot_amount or 0)
    return report_pdf(
        "Overtime Report",
        f"{display_month(month)} · Only days with payable overtime",
        ["ID", "Employee", "Date", "In Time", "Out Time", "Working Hours", "OT Paid Minutes", "OT Amount"],
        rows,
        col_widths=[16 * mm, 58 * mm, 26 * mm, 26 * mm, 26 * mm, 32 * mm, 36 * mm, 32 * mm],
        font_size=7,
        kpis=[
            ("Overtime days", len(rows)),
            ("Employees", len({row[0] for row in rows})),
            ("Payable minutes", f"{total_minutes:,}"),
            ("Overtime amount", pdf_money(total_amount)),
        ],
    )


def build_less_hours_report_pdf(month):
    rows = []
    names = employee_name_map(month)
    total_minutes = 0
    total_deduction = Decimal("0")
    for result in PayrollResult.query.filter_by(payroll_month=month).order_by(PayrollResult.employee_id):
        contributed = False
        for item in result.detail_json or []:
            shortage = int(item.get("shortage_minutes") or 0)
            deduction = Decimal(str(item.get("shortage_deduction") or "0"))
            if shortage <= 0 and deduction <= 0:
                continue
            rows.append([
                result.employee_id,
                names.get(result.employee_id, ""),
                item.get("date", ""),
                item.get("first_punch", "") or "-",
                item.get("last_punch", "") or "-",
                item.get("rounded_duration") or item.get("actual_duration") or item.get("raw_working_hours", ""),
                shortage,
                pdf_money(deduction),
            ])
            contributed = True
        if contributed:
            total_minutes += int(result.less_hours_minutes or 0)
            total_deduction += Decimal(result.less_hours_deduction or 0)
    return report_pdf(
        "Less Hours Report",
        f"{display_month(month)} · Only days with a short-hours deduction",
        ["ID", "Employee", "Date", "In Time", "Out Time", "Working Hours", "Less Minutes", "Deduction"],
        rows,
        col_widths=[16 * mm, 58 * mm, 26 * mm, 26 * mm, 26 * mm, 32 * mm, 32 * mm, 36 * mm],
        font_size=7,
        kpis=[
            ("Short days", len(rows)),
            ("Employees", len({row[0] for row in rows})),
            ("Short minutes", f"{total_minutes:,}"),
            ("Total deduction", pdf_money(total_deduction)),
        ],
    )


def build_error_report_pdf(month):
    rows = []
    for salary in SalaryRecord.query.filter_by(payroll_month=month).order_by(SalaryRecord.employee_id):
        if salary.warning:
            rows.append(["Salary", salary.employee_id, salary.name, salary.warning])
    names = employee_name_map(month)
    for rec in AttendanceRecord.query.filter_by(payroll_month=month).order_by(AttendanceRecord.employee_id, AttendanceRecord.date):
        if rec.warning:
            rows.append(["Attendance", rec.employee_id, names.get(rec.employee_id, rec.employee_name or ""), f"{rec.date}: {rec.warning}"])
    for result in PayrollResult.query.filter_by(payroll_month=month).order_by(PayrollResult.employee_id):
        if result.calculation_status != "Calculated":
            rows.append(["Payroll", result.employee_id, names.get(result.employee_id, ""), result.message or result.calculation_status])
    area_counts = {area: sum(1 for row in rows if row[0] == area) for area in ("Salary", "Attendance", "Payroll")}
    return report_pdf(
        "Error Report",
        f"{display_month(month)} · Items to resolve before finalizing",
        ["Area", "Employee ID", "Employee", "Issue"],
        rows,
        col_widths=[28 * mm, 26 * mm, 56 * mm, 160 * mm],
        font_size=7,
        kpis=[
            ("Total issues", len(rows)),
            ("Attendance", area_counts["Attendance"]),
            ("Payroll", area_counts["Payroll"]),
            ("Wage data", area_counts["Salary"]),
        ],
    )


def display_attendance_status(status):
    mapping = {
        "Full Day Present": "Full Day",
        "Half Day Present": "Half Day",
        "Paid Leave": "Paid Leave",
        "Half-Day Paid Leave": "Half-Day Paid Leave",
        "Full Day LOP": "Full Day LOP",
        "Half Day LOP": "Half Day LOP",
        "Unpaid Leave / LOP": "LOP",
        "Holiday": "Holiday",
        "Week Off": "Week Off",
        "Week Off Worked": "Week Off Worked",
        "Sandwich Leave": "Sandwich Leave",
        "Needs Review": "Needs Review",
        "Punch Error": "Punch Error",
        "Absent / Attendance Missing": "Absent",
        # Long enough to wrap a calendar cell, so shorten it the same way the web
        # view does.
        "Half-Day Paid Leave / Half-Day LOP": "Half Leave + Half LOP",
        "Half Day Present / Half-Day Leave": "Half Day + Half Leave",
        "Worked On-Site": "Worked On-Site",
        "Work From Home": "Work From Home",
        "Ignore": "Ignore",
    }
    return mapping.get(status or "", status or "N/A")


def pdf_money(value):
    if value is None:
        return "N/A"
    return f"{Decimal(value):,.2f}"


def words_below_thousand(number):
    number = int(number)
    parts = []
    if number >= 100:
        parts.append(ONES[number // 100] + " hundred")
        number %= 100
    if number >= 20:
        tens = TENS[number // 10]
        ones = number % 10
        parts.append(tens + (f" {ONES[ones]}" if ones else ""))
    elif number:
        parts.append(ONES[number])
    return " ".join(parts)


def integer_to_indian_words(number):
    number = int(number)
    if number == 0:
        return "zero"
    groups = [
        (10000000, "crore"),
        (100000, "lakh"),
        (1000, "thousand"),
        (1, ""),
    ]
    parts = []
    for value, label in groups:
        chunk = number // value
        if chunk:
            parts.append(words_below_thousand(chunk) + (f" {label}" if label else ""))
            number %= value
    return " ".join(parts)


def money_in_words(value):
    if value is None:
        return "Not calculated"
    amount = Decimal(value).quantize(Decimal("0.01"))
    rupees = int(amount)
    paise = int((amount - Decimal(rupees)) * 100)
    words = f"Rupees {integer_to_indian_words(rupees)}"
    if paise:
        words += f" and {integer_to_indian_words(paise)} paise"
    return words.capitalize() + " only"


def result_for_employee(month, employee_id):
    salary = SalaryRecord.query.filter_by(payroll_month=month, employee_id=employee_id).first()
    result = PayrollResult.query.filter_by(payroll_month=month, employee_id=employee_id).first()
    employee = db.session.get(Employee, employee_id)
    return employee, salary, result


def employee_salary_summary_rows(salary, result):
    if not result:
        return [["Calculation Status", "Not Calculated"]]
    base_salary = pdf_money(salary.salary) if salary else "N/A"
    final_salary = pdf_money(result.final_salary) if result.final_salary is not None else "Not Calculated"
    rows = [
        ["Calculation Status", result.calculation_status],
        ["Wage Type", salary.salary_type if salary else "N/A"],
        ["Base Salary", base_salary],
        ["Days in Month", payroll_month_days(result.payroll_month)],
        ["Paid Working Days", result.paid_working_days],
        ["Week Offs", result.week_offs],
        ["Total Paid Days", total_paid_days(result)],
        ["Full Days", result.full_days],
        ["Half Days", result.half_days],
        ["Paid Leave", result.paid_leaves],
        ["LOP Days", result.lop_days],
        ["Leave Balance", result.opening_leave],
        ["Leave Earned This Month", result.leave_earned],
        ["Leave Used This Month", result.leave_used],
        ["Leave Carry Forwarded", result.closing_leave],
        ["Less-Hours Deduction", pdf_money(result.less_hours_deduction)],
        ["PF Employee Contribution", pdf_money(getattr(result, "pf_employee", 0))],
        ["ESIC Employee Contribution", pdf_money(getattr(result, "esi_employee", 0))],
        ["LOP Deduction", pdf_money(result.lop_deduction)],
        ["Over Time Amount", pdf_money(result.ot_amount)],
        ["Adjustment", pdf_money(result.manual_adjustment)],
        ["Leave Encashment Days", getattr(result, "leave_encashment_days", 0)],
        ["Leave Encashment Amount", pdf_money(getattr(result, "leave_encashment_amount", 0))],
    ]
    if result.payroll_rule_type == "DAILY":
        rows.extend([
            ["Absence This Month", bonus_absence_text(result)],
            ["Attendance Bonus", bonus_percent_text(result)],
            ["Attendance Bonus Amount", pdf_money(getattr(result, "attendance_bonus_amount", 0))],
        ])
    if result_has_loan(result):
        rows.extend([
            ["Loan Deduction", pdf_money(getattr(result, "loan_deduction", 0))],
            ["Pending Loan Amount", pdf_money(getattr(result, "loan_pending_amount", 0))],
        ])
    if result_has_advance(result):
        rows.append(["Advance Salary Deduction", pdf_money(getattr(result, "advance_deduction", 0))])
    rows.extend([
        ["Total Deduction", pdf_money(result.total_deduction)],
        ["Total Addition", pdf_money(result.total_addition)],
        ["Final Payable Salary", final_salary],
    ])
    return rows


def bonus_absence_text(result):
    return minutes_to_duration(int(getattr(result, "absence_minutes", 0) or 0))


def bonus_percent_text(result):
    """Bonus band as it should read on a slip, e.g. "10% of earned wage"."""
    percent = Decimal(getattr(result, "attendance_bonus_percent", 0) or 0)
    if percent == 0:
        # A zero bonus reads very differently depending on why, so say which it is.
        employee = db.session.get(Employee, result.employee_id)
        if employee and employee.bonus_ignored:
            return "Excluded in Employee Master"
        return "Not earned this month"
    return f"{format_percent(percent)}%"


def bonus_short_text(result):
    """Bonus for a narrow table column, where the full sentence would not fit."""
    percent = Decimal(getattr(result, "attendance_bonus_percent", 0) or 0)
    if percent > 0:
        return f"{format_percent(percent)}%"
    employee = db.session.get(Employee, result.employee_id)
    return "Excluded" if employee and employee.bonus_ignored else "—"


def employee_compact_summary_rows(salary, result):
    if not result:
        return [["Status", "Not Calculated", "Wage Type", salary.salary_type if salary else "N/A"]]
    final_value = result.final_salary if result.final_salary is not None else None
    if result.payroll_rule_type == "DAILY":
        rows = [
            ["Payable Salary", pdf_money(final_value) if final_value is not None else "Not Calculated", "In Words", money_in_words(final_value)],
            ["Daily Wage", pdf_money(salary.salary) if salary else "N/A", "Wage Type", salary.salary_type if salary else "N/A"],
            ["Status", result.calculation_status, "Days in Month", payroll_month_days(result.payroll_month)],
            ["Paid Working Days", result.paid_working_days, "Paid Holidays", result.holidays],
            ["Payable Days", total_paid_days(result), "Week Offs", result.week_offs],
            ["Less Hours Deduction", pdf_money(result.less_hours_deduction), "Over Time", pdf_money(result.ot_amount)],
            ["Adjustment", pdf_money(result.manual_adjustment), "Absence This Month", bonus_absence_text(result)],
            ["Attendance Bonus", bonus_percent_text(result), "Attendance Bonus Amount", pdf_money(getattr(result, "attendance_bonus_amount", 0))],
        ]
        if result_has_loan(result):
            rows.append(["Loan Deduction", pdf_money(getattr(result, "loan_deduction", 0)), "Pending Loan Amount", pdf_money(getattr(result, "loan_pending_amount", 0))])
        if result_has_advance(result):
            rows.append(["Advance Salary Deduction", pdf_money(getattr(result, "advance_deduction", 0)), "", ""])
        return rows
    rows = [
        ["Payable Salary", pdf_money(final_value) if final_value is not None else "Not Calculated", "In Words", money_in_words(final_value)],
        ["Base Salary", pdf_money(salary.salary) if salary else "N/A", "Wage Type", salary.salary_type if salary else "N/A"],
        ["Status", result.calculation_status, "Days in Month", payroll_month_days(result.payroll_month)],
        ["Paid Working Days", result.paid_working_days, "Week Offs", result.week_offs],
        ["Total Paid Days", total_paid_days(result), "LOP Days", result.lop_days],
        ["Leave Balance", result.opening_leave, "Leave Earned This Month", result.leave_earned],
        ["Leave Used This Month", result.leave_used, "Leave Carry Forwarded", result.closing_leave],
        ["Less Hours Deduction", pdf_money(result.less_hours_deduction), "Over Time", pdf_money(result.ot_amount)],
        ["Adjustment", pdf_money(result.manual_adjustment), "Leave Encashed", f"{getattr(result, 'leave_encashment_days', 0)}d / {pdf_money(getattr(result, 'leave_encashment_amount', 0))}"],
    ]
    # Statutory rows appear only where the employee is actually covered, so a slip
    # for someone outside PF or ESI is not padded with zeroes.
    if Decimal(getattr(result, "pf_employee", 0) or 0):
        rows.append(["PF Employee Contribution", pdf_money(result.pf_employee),
                     "PF Employer Contribution", pdf_money(result.pf_employer)])
    if Decimal(getattr(result, "esi_employee", 0) or 0):
        rows.append(["ESIC Employee Contribution", pdf_money(result.esi_employee),
                     "ESIC Employer Contribution", pdf_money(result.esi_employer)])
    if result_has_loan(result):
        rows.append(["Loan Deduction", pdf_money(getattr(result, "loan_deduction", 0)), "Pending Loan Amount", pdf_money(getattr(result, "loan_pending_amount", 0))])
    if result_has_advance(result):
        rows.append(["Advance Salary Deduction", pdf_money(getattr(result, "advance_deduction", 0)), "", ""])
    return rows


def monthly_attendance_summary_rows(salary, result):
    """The monthly slip with every pay figure removed.

    Payable salary, the amount in words and the base salary come off; the attendance
    and leave picture stays. This is an attendance document, not a payslip.
    """
    if not result:
        return [["Status", "Not Calculated", "Wage Type", salary.salary_type if salary else "N/A"]]
    # Grouped by what a reader is looking for: the day counts together, then the
    # leave position, then the money. Overtime and adjustment share the last row.
    rows = [
        ["Status", result.calculation_status, "Days in Month", payroll_month_days(result.payroll_month)],
        ["Paid Working Days", result.paid_working_days, "Total Paid Days", total_paid_days(result)],
        ["Week Offs", result.week_offs, "LOP Days", result.lop_days],
        ["Leave Balance", result.opening_leave, "Leave Earned This Month", result.leave_earned],
        ["Leave Used This Month", result.leave_used, "Leave Carry Forwarded", result.closing_leave],
        ["Leave Encashed", f"{getattr(result, 'leave_encashment_days', 0)}d / {pdf_money(getattr(result, 'leave_encashment_amount', 0))}",
         "Less Hours Deduction", pdf_money(result.less_hours_deduction)],
        ["Over Time", pdf_money(result.ot_amount), "Adjustment", pdf_money(result.manual_adjustment)],
    ]
    if result_has_loan(result):
        rows.append(["Loan Deduction", pdf_money(getattr(result, "loan_deduction", 0)),
                     "Pending Loan Amount", pdf_money(getattr(result, "loan_pending_amount", 0))])
    if result_has_advance(result):
        rows.append(["Advance Salary Deduction", pdf_money(getattr(result, "advance_deduction", 0)), "", ""])
    return rows


def bonus_band_text(result):
    """The bonus as a bare band: NIL, 5% or 10%.

    The daily sheet carries no money at all, so the band is stated without the amount
    it works out to.
    """
    percent = Decimal(getattr(result, "attendance_bonus_percent", 0) or 0)
    return f"{format_percent(percent)}%" if percent > 0 else "NIL"


def daily_attendance_summary_rows(salary, result):
    """The daily slip stripped back to attendance only.

    No rupee figure appears anywhere: short hours and overtime are stated as time,
    and the bonus as its band. What is left says whether the worker turned up, how
    short they were, and what that earned.
    """
    if not result:
        return [["Status", "Not Calculated", "", ""]]
    return [
        ["Status", result.calculation_status, "Payable Days", total_paid_days(result)],
        ["Less Hours", minutes_to_duration(result.less_hours_minutes or 0),
         "Over Time", minutes_to_duration(result.payable_ot_minutes or 0)],
        ["Absence This Month", bonus_absence_text(result), "Bonus", bonus_band_text(result)],
    ]


# Cell fill and text colour per attendance status, so a month reads at a glance.
CALENDAR_TONES = {
    "Full Day": (GREEN_WASH, GREEN_TEXT),
    "Half Day": (ORANGE_WASH, ORANGE_TEXT),
    "Week Off Worked": (GREEN_WASH, GREEN_TEXT),
    # Marked present by override rather than by punches, so it reads distinctly.
    "Worked On-Site": (TEAL_WASH, TEAL_TEXT),
    "Work From Home": (TEAL_WASH, TEAL_TEXT),
    "Week Off": (TINT_WASH, TINT_TEXT),
    "Holiday": (TINT_WASH, TINT_TEXT),
    "Paid Leave": (ORANGE_WASH, ORANGE_TEXT),
    "Half-Day Paid Leave": (ORANGE_WASH, ORANGE_TEXT),
    "Sandwich Leave": (ORANGE_WASH, ORANGE_TEXT),
    "Half Leave + Half LOP": (ORANGE_WASH, ORANGE_TEXT),
    "Half Day + Half Leave": (ORANGE_WASH, ORANGE_TEXT),
    "Full Day LOP": (RED_WASH, RED_TEXT),
    "Half Day LOP": (RED_WASH, RED_TEXT),
    "LOP": (RED_WASH, RED_TEXT),
    "Absent": (RED_WASH, RED_TEXT),
    "Needs Review": (RED_WASH, RED_TEXT),
    "Punch Error": (RED_WASH, RED_TEXT),
}
CALENDAR_WEEKDAYS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


def punch_sessions(punches, first_punch="", last_punch=""):
    """Format punches as in-out pairs, e.g. ["09:34 AM - 10:55 AM", "05:43 PM - 06:30 PM"].

    Showing only the first and last punch made a split day look continuous, so a
    2h 08m day could read as 09:34 to 18:30.
    """
    punches = [p for p in (punches or []) if p]
    if not punches:
        pair = [value for value in (first_punch, last_punch) if value]
        return [" - ".join(pair)] if pair else []
    sessions = []
    for index in range(0, len(punches) - 1, 2):
        sessions.append(f"{punches[index]} - {punches[index + 1]}")
    if len(punches) % 2:
        sessions.append(f"{punches[-1]} - ?")
    return sessions


def calendar_tone(status):
    return CALENDAR_TONES.get(status, (SURFACE_SOFT, MUTED))


def attendance_calendar_grid(month, result):
    """Attendance laid out as calendar weeks for the PDF, Sunday first.

    Returns (weeks, statuses) where each cell is either None for padding or a dict
    with the day number, status and hours.
    """
    year, month_number = (int(part) for part in month.split("-"))
    days_in_month = calendar.monthrange(year, month_number)[1]
    by_day = {}
    for item in (result.detail_json or []) if result else []:
        try:
            day = int(str(item.get("date", ""))[-2:])
        except ValueError:
            continue
        by_day[day] = {
            "day": day,
            "status": display_attendance_status(item.get("attendance_status")),
            "hours": item.get("rounded_duration") or item.get("actual_duration") or "",
            "sessions": punch_sessions(item.get("punches"), item.get("first_punch") or "", item.get("last_punch") or ""),
            "shortage": int(item.get("shortage_minutes") or 0),
            "overtime": int(item.get("payable_ot") or 0),
        }
    first_weekday = (date(year, month_number, 1).weekday() + 1) % 7
    cells = [None] * first_weekday
    for day in range(1, days_in_month + 1):
        cells.append(by_day.get(day, {
            "day": day, "status": "", "hours": "", "sessions": [], "shortage": 0, "overtime": 0,
        }))
    while len(cells) % 7:
        cells.append(None)
    return [cells[index:index + 7] for index in range(0, len(cells), 7)]


def employee_detail_rows(result, compact=False):
    rows = [["Date", "Working Hours", "Status", "Paid Day", "Shortage", "Over Time"]]
    if not result or not result.detail_json:
        rows.append(["N/A", "N/A", "Not Calculated", "N/A", "N/A", "N/A"])
        return rows
    for item in result.detail_json:
        rows.append([
            item.get("date", ""),
            item.get("rounded_duration", ""),
            display_attendance_status(item.get("attendance_status")),
            item.get("paid_day_value", ""),
            item.get("shortage_minutes", 0),
            item.get("payable_ot", 0),
        ])
    return rows


def employee_detail_compact_rows(result):
    detail_rows = employee_detail_rows(result)[1:]
    split = (len(detail_rows) + 1) // 2
    numbered_rows = list(enumerate(detail_rows, start=1))
    left = numbered_rows[:split]
    right = numbered_rows[split:]
    rows = [["Sr No", "Date", "Working Hours", "Status", "Shortage", "Over Time", "", "Sr No", "Date", "Working Hours", "Status", "Shortage", "Over Time"]]
    max_len = max(len(left), len(right), 1)
    for index in range(max_len):
        l_index, l = left[index] if index < len(left) else ("", ["", "", "", "", "", ""])
        r_index, r = right[index] if index < len(right) else ("", ["", "", "", "", "", ""])
        rows.append([l_index, l[0], l[1], l[2], l[4], l[5], "", r_index, r[0], r[1], r[2], r[4], r[5]])
    return rows


class SlipVariant:
    """What a per-employee sheet shows.

    The salary slip is the full document. The two summaries drop the pay figures and
    keep the attendance picture. The daily summary additionally carries no SMARTfill
    branding anywhere on the page, which is a regulatory requirement for cash-wage
    workers, so it must not gain a logo, a footer brand line or a PDF author tag.
    """

    def __init__(self, title, branded=True, footer=True, show_employee_id=True, show_role=True):
        self.title = title
        self.branded = branded
        self.footer = footer
        self.show_employee_id = show_employee_id
        self.show_role = show_role


SLIP = SlipVariant("Salary Slip")
MONTHLY_SUMMARY = SlipVariant("Attendance Summary", footer=False)
DAILY_SUMMARY = SlipVariant(
    "Summary", branded=False, footer=False, show_employee_id=False, show_role=False,
)


def _brand_logo(width=30 * mm, height=10.5 * mm):
    if LOGO_PATH.exists():
        image = Image(str(LOGO_PATH), width=width, height=height)
        image.hAlign = "LEFT"
        return image
    return Paragraph("SMARTfill", ParagraphStyle("LogoFallback", fontName="Helvetica-Bold", fontSize=13, textColor=BRAND_BLUE))


def _salary_slip_header(month, salary, result, styles, compact=False, variant=SLIP):
    employee_id = salary.employee_id if salary else (result.employee_id if result else "")
    employee_name = salary.name if salary else employee_id
    title_size = 14 if compact else 30
    text_size = 7 if compact else 11
    if variant is not SLIP:
        # The summaries drop the salary figures, which frees the room these sizes
        # needed. Everything here is roughly twice the compact slip. The title is
        # sized to stay on one line in its column rather than wrapping.
        title_size = 16
        text_size = 12

    # The month and the employee identity are what a reader actually looks for, so
    # they get their own larger sizes. Status and the brand line stay as meta text.
    month_size = 9 if compact else 17
    name_size = 9.5 if compact else 18
    role_size = 7.5 if compact else 12.5
    if variant is not SLIP:
        month_size, name_size, role_size = 15, 16, 12
    title_style = ParagraphStyle(
        "SlipTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=title_size,
        leading=title_size + 1,
        textColor=BRAND_BLUE,
        alignment=TA_LEFT,
    )
    meta_style = ParagraphStyle(
        "SlipMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=text_size,
        leading=text_size + 1,
        textColor=BRAND_MUTED,
    )
    month_style = ParagraphStyle(
        "SlipMonth",
        parent=meta_style,
        fontSize=month_size,
        leading=month_size + 2,
        textColor=INK,
    )
    employee_style = ParagraphStyle(
        "SlipEmployee",
        parent=meta_style,
        fontName="Helvetica-Bold",
        fontSize=name_size,
        leading=name_size + 2,
        textColor=colors.HexColor("#172033"),
    )
    role_style = ParagraphStyle(
        "SlipRole",
        parent=meta_style,
        fontSize=role_size,
        leading=role_size + 2,
    )
    employee = db.session.get(Employee, employee_id) if employee_id else None
    role_bits = [bit for bit in [(employee.designation if employee else ""), (employee.department if employee else "")] if bit]
    status = result.calculation_status if result else "Not Calculated"
    _wash, status_colour = _status_colours(status)
    status_style = ParagraphStyle("SlipStatus", parent=meta_style, fontName="Helvetica-Bold", textColor=status_colour)

    left = [
        Paragraph(variant.title, title_style),
        Paragraph(display_month(month), month_style),
    ]
    # The daily summary carries no employee number and no designation: it identifies
    # the worker by name alone.
    identity = f"{employee_id} &middot; {employee_name}" if variant.show_employee_id else employee_name
    middle = [Paragraph(identity, employee_style)]
    if variant.show_role:
        middle.append(Paragraph(" &middot; ".join(role_bits) if role_bits else "Designation not set", role_style))
    # Status and the brand line belong to the salary slip only; the summaries carry
    # the status inside the table instead.
    right = [Paragraph(status, status_style), Paragraph("SMARTfill Payroll", meta_style)] if variant is SLIP else []
    cells = [left, middle, right]
    if variant is SLIP:
        widths = [40 * mm, 76 * mm, 44 * mm] if compact else [44 * mm, 74 * mm, 40 * mm]
        logo_width, logo_height, logo_column = (
            (26 * mm, 9 * mm, 34 * mm) if compact else (30 * mm, 10.5 * mm, 38 * mm)
        )
    else:
        # The summary sheets are one to a page and total 190mm of usable width.
        # The logo is sized up from the slip's 26mm now that there is room for it.
        widths = [66 * mm, 80 * mm, 0]
        logo_width, logo_height, logo_column = 40 * mm, 14 * mm, 44 * mm
    if variant.branded:
        cells.append(_brand_logo(width=logo_width, height=logo_height))
        widths = widths + [logo_column]
    else:
        # No logo column at all, so the name is not left floating in a narrow cell.
        widths = [widths[0], sum(widths[1:]) + logo_column, 0]
    table = Table([cells], colWidths=widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), SURFACE_SOFT),
        ("BOX", (0, 0), (-1, -1), 0.5, SEPARATOR),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("ALIGN", (3, 0), (3, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5 if compact else 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5 if compact else 8),
    ]))
    table.cornerRadii = [CARD_RADIUS] * 4
    return table


def _report_brand_header(title, subtitle, styles, available_width, compact=False):
    """iOS large-title masthead: heavy tight title, muted subtitle, logo to the right."""
    title_size = 12 if compact else 17
    subtitle_size = 6.5 if compact else 8.5
    title_style = ParagraphStyle(
        "BrandedReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=title_size,
        leading=title_size + 2,
        textColor=INK,
        spaceAfter=0,
    )
    subtitle_style = ParagraphStyle(
        "BrandedReportSubtitle",
        parent=styles["Normal"],
        fontSize=subtitle_size,
        leading=subtitle_size + 3,
        textColor=MUTED,
    )
    left = [Paragraph(title, title_style), Paragraph(subtitle, subtitle_style)]
    logo_width = 30 * mm if compact else 34 * mm
    table = Table(
        [[left, _brand_logo(width=logo_width, height=10 * mm if compact else 11.5 * mm)]],
        colWidths=[available_width - logo_width - 8 * mm, logo_width + 8 * mm],
    )
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, SEPARATOR),
    ]))
    return table


def _kpi_row(items, available_width, compact=False):
    """A row of iOS metric tiles: muted caption above a large tight value."""
    if not items:
        return Spacer(1, 0)
    label_size = 5.6 if compact else 7
    value_size = 9 if compact else 12.5
    label_style = ParagraphStyle("KpiLabel", fontName="Helvetica", fontSize=label_size, leading=label_size + 2, textColor=MUTED)
    value_style = ParagraphStyle("KpiValue", fontName="Helvetica-Bold", fontSize=value_size, leading=value_size + 2, textColor=INK)
    cells = [[Paragraph(label, label_style), Paragraph(str(value), value_style)] for label, value in items]
    width = available_width / len(items)
    table = Table([[Table([[c[0]], [c[1]]], colWidths=[width - 6]) for c in cells]], colWidths=[width] * len(items))
    inner = TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, 0), (-1, -1), SURFACE_SOFT),
        ("BOX", (0, 0), (-1, -1), 0.5, SEPARATOR),
    ])
    for cell_table in table._cellvalues[0]:
        cell_table.setStyle(inner)
        cell_table.cornerRadii = [CARD_RADIUS] * 4
    table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("RIGHTPADDING", (-1, 0), (-1, 0), 0),
        ("LEFTPADDING", (1, 0), (-1, 0), 3),
        ("RIGHTPADDING", (0, 0), (-2, 0), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return table


def _pdf_header_footer(canvas, doc):
    """Hairline rule with muted footer text, matching the app's separator treatment."""
    canvas.saveState()
    baseline = 7.5 * mm
    canvas.setStrokeColor(SEPARATOR)
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, baseline + 3.5 * mm, doc.pagesize[0] - doc.rightMargin, baseline + 3.5 * mm)
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(MUTED)
    canvas.drawString(doc.leftMargin, baseline, "SMARTfill Payroll")
    canvas.setFillColor(FAINT)
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, baseline, f"Page {doc.page}")
    canvas.restoreState()


SLIP_FOOTER_NOTE = "*** This is a system generated payslip, hence a signature is not required."


def _salary_slip_page_callback(title):
    """Slip pages carry only the footnote, pinned to the bottom of the sheet.

    Drawing it here rather than flowing it after the last table keeps it at the foot
    whatever the slip's height, and leaves off the brand line and page number that
    the other reports carry.
    """
    def callback(canvas, doc):
        canvas.setTitle(title)
        canvas.setAuthor("SMARTfill Attendance & Payroll Management")
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(FAINT)
        canvas.drawCentredString(doc.pagesize[0] / 2, 9 * mm, SLIP_FOOTER_NOTE)
        canvas.restoreState()

    return callback


def _titled_page_callback(title, variant=SLIP):
    def callback(canvas, doc):
        canvas.setTitle(title)
        # The author tag is branding too, and it survives in the file's metadata long
        # after the page is printed, so an unbranded document must not carry it.
        if variant.branded:
            canvas.setAuthor("SMARTfill Attendance & Payroll Management")
        if variant.footer:
            _pdf_header_footer(canvas, doc)

    return callback


def _status_column_style(data, column, header=True):
    """Colour a status column per row, mirroring the status badges in the web UI."""
    style = []
    start = 1 if header else 0
    for index in range(start, len(data)):
        cell = data[index][column] if column < len(data[index]) else ""
        text = getattr(cell, "text", cell)
        # Paragraph.text keeps the source markup, so strip any tags before matching.
        text = re.sub(r"<[^>]+>", "", str(text)).strip()
        if not text or text == "—":
            continue
        _wash, colour = _status_colours(text)
        if colour is not MUTED:
            style.append(("TEXTCOLOR", (column, index), (column, index), colour))
    return style


def _table(data, col_widths=None, font_size=8, header=True, highlight_rows=None, blank_columns=None, status_column=None, accent_rows=None, center=False, center_from=None):
    """An iOS-style list: rounded card, hairline row separators, no vertical rules.

    The previous look was a full grid with a tinted header band. Apple's tables
    separate rows with a single hairline and let whitespace do the column work, so
    the grid is dropped in favour of zebra banding and generous padding.
    """
    compact = font_size <= 5.8
    table = Table(
        data,
        colWidths=col_widths,
        repeatRows=1 if header else 0,
        cornerRadii=[CARD_RADIUS] * 4,
    )
    pad_y = 1.5 if compact else 4.5
    pad_x = 3 if compact else 6
    highlight_rows = highlight_rows or []
    blank_columns = blank_columns or []
    style = [
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("TEXTCOLOR", (0, 0), (-1, -1), INK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), pad_x),
        ("RIGHTPADDING", (0, 0), (-1, -1), pad_x),
        ("TOPPADDING", (0, 0), (-1, -1), pad_y),
        ("BOTTOMPADDING", (0, 0), (-1, -1), pad_y),
        ("BACKGROUND", (0, 0), (-1, -1), SURFACE),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, SEPARATOR),
        ("BOX", (0, 0), (-1, -1), 0.5, SEPARATOR),
    ]
    if center:
        style.append(("ALIGN", (0, 0), (-1, -1), "CENTER"))
    if center_from is not None:
        # Narrow numeric columns read as one run of text when every header is flush
        # left against its neighbour; centring them puts whitespace on both sides.
        style.append(("ALIGN", (center_from, 0), (-1, -1), "CENTER"))
    if header:
        style.extend([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), SURFACE_SOFT),
            ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, SEPARATOR),
            ("TOPPADDING", (0, 0), (-1, 0), pad_y + 1),
            ("BOTTOMPADDING", (0, 0), (-1, 0), pad_y + 1),
        ])
        # Zebra banding, which reads better than rules on wide landscape tables.
        for index in range(2, len(data), 2):
            style.append(("BACKGROUND", (0, index), (-1, index), ZEBRA))
    else:
        style.append(("FONTNAME", (0, 0), (-1, -1), "Helvetica"))
    for row in highlight_rows:
        style.extend([
            ("BACKGROUND", (0, row), (-1, row), TINT_WASH),
            ("TEXTCOLOR", (0, row), (-1, row), TINT_TEXT),
            ("FONTNAME", (0, row), (-1, row), "Helvetica-Bold"),
        ])
    # Rows that need their own colour rather than the blue highlight, so a figure
    # like the attendance bonus reads as earned or not earned at a glance.
    for row, wash, text_colour in accent_rows or []:
        style.extend([
            ("BACKGROUND", (0, row), (-1, row), wash),
            ("TEXTCOLOR", (0, row), (-1, row), text_colour),
            ("FONTNAME", (0, row), (-1, row), "Helvetica-Bold"),
        ])
    if status_column is not None:
        style.extend(_status_column_style(data, status_column, header=header))
    # Spacer columns used by the two-up compact slip; keep them invisible.
    for column in blank_columns:
        style.extend([
            ("BACKGROUND", (column, 0), (column, -1), SURFACE),
            ("TEXTCOLOR", (column, 0), (column, -1), SURFACE),
            ("LINEBELOW", (column, 0), (column, -1), 0, SURFACE),
        ])
    table.setStyle(TableStyle(style))
    return table


def attendance_calendar_table(month, result, styles, available_width, compact=False, variant=SLIP, sizes=None):
    """The month's attendance as a colour-coded calendar, matching the web view."""
    weeks = attendance_calendar_grid(month, result)
    day_size = 5.6 if compact else 7.2
    meta_size = 4.4 if compact else 5.6
    punch_size = 4.2 if compact else 5.3
    if variant is not SLIP:
        # Roughly double the compact slip. One sheet per page pays for the room.
        day_size, meta_size, punch_size = 11, 8.6, 8.6
    if sizes:
        day_size, meta_size, punch_size = sizes
    # The summaries centre every cell; the slip keeps its left-aligned column.
    cell_align = TA_CENTER if variant is not SLIP else TA_LEFT
    day_style = ParagraphStyle("CalDay", fontName="Helvetica-Bold", fontSize=day_size, leading=day_size + 1.5, textColor=INK, alignment=cell_align)
    punch_style = ParagraphStyle("CalPunch", fontName="Helvetica", fontSize=punch_size, leading=punch_size + 1.2, textColor=INK, alignment=cell_align)
    head_style = ParagraphStyle("CalHead", fontName="Helvetica-Bold", fontSize=meta_size, leading=meta_size + 1.5, textColor=MUTED, alignment=1)

    data = [[Paragraph(name.upper(), head_style) for name in CALENDAR_WEEKDAYS]]
    style = [
        ("GRID", (0, 0), (-1, -1), 0.4, SEPARATOR),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("BACKGROUND", (0, 0), (-1, 0), SURFACE_SOFT),
    ]
    for week_index, week in enumerate(weeks, start=1):
        row = []
        for column, cell in enumerate(week):
            if cell is None:
                row.append("")
                style.append(("BACKGROUND", (column, week_index), (column, week_index), SURFACE))
                continue
            wash, text_colour = calendar_tone(cell["status"])
            status_style = ParagraphStyle(
                f"CalStatus{week_index}{column}", fontName="Helvetica-Bold",
                fontSize=meta_size, leading=meta_size + 1.4, textColor=text_colour,
                alignment=cell_align,
            )
            meta_style_local = ParagraphStyle(
                f"CalMeta{week_index}{column}", fontName="Helvetica",
                fontSize=meta_size, leading=meta_size + 1.4, textColor=MUTED,
                alignment=cell_align,
            )
            parts = [Paragraph(str(cell["day"]), day_style)]
            if cell["status"]:
                parts.append(Paragraph(cell["status"], status_style))
            for session in cell["sessions"]:
                parts.append(Paragraph(session, punch_style))
            if cell["hours"]:
                parts.append(Paragraph(cell["hours"], meta_style_local))
            extras = []
            if cell["shortage"]:
                extras.append(f"-{cell['shortage']}m")
            if cell["overtime"]:
                extras.append(f"+{cell['overtime']}m OT")
            if extras:
                parts.append(Paragraph(" ".join(extras), meta_style_local))
            row.append(parts)
            style.append(("BACKGROUND", (column, week_index), (column, week_index), wash))
        data.append(row)

    column_width = available_width / 7
    table = Table(data, colWidths=[column_width] * 7, repeatRows=1)
    table.setStyle(TableStyle(style))
    return table


def attendance_bonus_accent_rows(summary_rows, result):
    """Colour the attendance bonus row green when earned, red when it was not."""
    if not result or result.payroll_rule_type != "DAILY":
        return []
    earned = Decimal(getattr(result, "attendance_bonus_percent", 0) or 0) > 0
    wash, text_colour = (GREEN_WASH, GREEN_TEXT) if earned else (RED_WASH, RED_TEXT)
    # The label sits in the first column on the slip and in the third on the daily
    # summary, so match either rather than pinning it to one position.
    return [
        (index, wash, text_colour)
        for index, row in enumerate(summary_rows)
        if row and {"Attendance Bonus", "Bonus"} & {row[0], row[2] if len(row) > 2 else None}
    ]


def attendance_summary_block(month, salary, result, styles, variant, available_width):
    """One employee's attendance sheet: header, figures, then the month calendar."""
    header = _salary_slip_header(month, salary, result, styles, compact=True, variant=variant)
    rows = (daily_attendance_summary_rows if variant is DAILY_SUMMARY else monthly_attendance_summary_rows)(salary, result)
    label_width = available_width * 0.26
    value_width = available_width * 0.24
    summary = _table(
        rows,
        col_widths=[label_width, value_width, label_width, available_width - (2 * label_width) - value_width],
        font_size=9, header=False, center=True,
        accent_rows=attendance_bonus_accent_rows(rows, result),
    )
    calendar = attendance_calendar_table(month, result, styles, available_width, compact=True, variant=variant)
    return KeepTogether([header, Spacer(1, 6), summary, Spacer(1, 6), calendar])


# The salary slip's earning components. Only these three appear; anything else that
# adds to pay is listed under Others so the slip still reconciles to the net figure.
SLIP_EARNING_COMPONENTS = (
    ("basic_salary", "Basic"),
    ("hra", "House Rent Allowance"),
    ("allowance", "Conveyance Allowance"),
)


def slip_paid_ratio(salary_record, result):
    """How much of the contracted salary was actually earned this month.

    Driven by loss of pay alone. Short hours, loans and statutory dues are separate
    deduction lines, so folding them in here would count them twice.
    """
    contracted = Decimal(salary_record.salary or 0) if salary_record else Decimal("0")
    if not contracted or not result:
        return Decimal("1")
    return (contracted - Decimal(result.lop_deduction or 0)) / contracted


def slip_earning_rows(employee, salary_record, result):
    """(component, actual, paid) for the three slip components.

    An employee with no salary breakup captured would otherwise show a gross of zero
    against a real net figure, so the whole salary is shown as Basic in that case.
    """
    ratio = slip_paid_ratio(salary_record, result)
    actuals = [(label, Decimal(getattr(employee, field, 0) or 0) if employee else Decimal("0"))
               for field, label in SLIP_EARNING_COMPONENTS]
    if sum((amount for _label, amount in actuals), Decimal("0")) == 0:
        actuals = [("Basic", Decimal(salary_record.salary or 0) if salary_record else Decimal("0")),
                   ("House Rent Allowance", Decimal("0")), ("Conveyance Allowance", Decimal("0"))]
    return [(label, amount, amount * ratio) for label, amount in actuals]


def slip_other_earnings(result):
    # One Adjustment line carrying its own sign, rather than a positive one here and
    # a negative one under deductions.
    return [
        ("Overtime", Decimal(result.ot_amount or 0)),
        ("Leave Encashment", Decimal(getattr(result, "leave_encashment_amount", 0) or 0)),
        ("Adjustment (+/-)", Decimal(result.manual_adjustment or 0)),
    ]


def slip_other_deductions(result):
    return [
        ("Short Hours", Decimal(result.less_hours_deduction or 0)),
        ("Loan", Decimal(getattr(result, "loan_deduction", 0) or 0)),
        ("Advance Salary", Decimal(getattr(result, "advance_deduction", 0) or 0)),
    ]


def salary_slip_story(month, salary_record, result, styles, available_width):
    """One salary slip in the classic Indian payslip layout.

    Earnings and deductions sit side by side in a single grid so the rules line up,
    with an Others block beneath for anything outside the three fixed components.
    Employer contributions follow, since they are a company cost and must not be
    mistaken for a deduction from the employee.
    """
    employee = db.session.get(Employee, salary_record.employee_id) if salary_record else None
    name = salary_record.name if salary_record else ""
    label_style = ParagraphStyle("SlipLabel", parent=styles["Normal"], fontName="Helvetica",
                                 fontSize=8, leading=10.5, textColor=MUTED)
    value_style = ParagraphStyle("SlipValue", parent=styles["Normal"], fontName="Helvetica-Bold",
                                 fontSize=8, leading=10.5, textColor=INK)
    address_style = ParagraphStyle("SlipAddress", parent=styles["Normal"], fontName="Helvetica",
                                   fontSize=7.4, leading=9.5, textColor=MUTED, alignment=TA_CENTER)
    band_style = ParagraphStyle("SlipBand", parent=styles["Normal"], fontName="Helvetica-Bold",
                                fontSize=9, leading=11.5, textColor=INK)
    words_style = ParagraphStyle("SlipWords", parent=styles["Normal"], fontName="Helvetica-Bold",
                                 fontSize=8, leading=10.5, textColor=TINT_TEXT)

    hairline = [
        ("GRID", (0, 0), (-1, -1), 0.5, SEPARATOR),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    # Masthead: logo, then the registered address.
    masthead = Table([[_brand_logo(width=46 * mm, height=16 * mm)], [Paragraph(COMPANY_ADDRESS, address_style)]],
                     colWidths=[available_width])
    masthead.setStyle(TableStyle(hairline + [
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (0, 0), 8),
        ("BOTTOMPADDING", (0, 0), (0, 0), 8),
    ]))

    lop_days = Decimal(result.lop_days or 0) if result else Decimal("0")
    # Widths sized against the longest month name so no cell wraps.
    band = Table([[
        Paragraph(f"Pay Slip: {display_month(month)}", band_style),
        Paragraph(f"Days in this Month: {payroll_month_days(month)}", band_style),
        Paragraph(f"Payable days: {total_paid_days(result) if result else '—'}", band_style),
        Paragraph(f"Loss of pay days: {leave_days(lop_days)}", band_style),
    ]], colWidths=[available_width * 0.26, available_width * 0.24,
                   available_width * 0.24, available_width * 0.26])
    band.setStyle(TableStyle(hairline + [("BACKGROUND", (0, 0), (-1, -1), SURFACE_SOFT)]))

    # Bank, PF, UAN and PAN are deliberately absent: the slip identifies the employee
    # and nothing more.
    details = Table([
        [Paragraph("Employee Code", label_style), Paragraph(str(salary_record.employee_id if salary_record else ""), value_style),
         Paragraph("Department", label_style), Paragraph((employee.department if employee else "") or "—", value_style)],
        [Paragraph("Employee Name", label_style), Paragraph(name, value_style),
         Paragraph("Designation", label_style), Paragraph((employee.designation if employee else "") or "—", value_style)],
    ], colWidths=[available_width * 0.17, available_width * 0.33, available_width * 0.17, available_width * 0.33])
    details.setStyle(TableStyle(hairline))

    earnings = slip_earning_rows(employee, salary_record, result)
    deductions = [
        ("P.F", Decimal(getattr(result, "pf_employee", 0) or 0)),
        ("ESIC", Decimal(getattr(result, "esi_employee", 0) or 0)),
        ("Professional Tax", Decimal(getattr(result, "professional_tax", 0) or 0)),
        ("TDS", Decimal(getattr(result, "tds", 0) or 0)),
    ] if result else []
    other_earnings = slip_other_earnings(result) if result else []
    other_deductions = slip_other_deductions(result) if result else []

    def money_cell(amount, bold=False):
        return Paragraph(pdf_money(amount), value_style if bold else label_style)

    rows = [[Paragraph("EARNINGS (INR)", band_style), "", "", Paragraph("DEDUCTIONS (INR)", band_style), ""]]
    rows.append([Paragraph(text, label_style) for text in
                 ("Component", "Actual Amount", "Paid Amount", "Component", "Paid Amount")])
    span = max(len(earnings), len(deductions), 1)
    for index in range(span):
        left = earnings[index] if index < len(earnings) else ("", None, None)
        right = deductions[index] if index < len(deductions) else ("", None)
        rows.append([
            Paragraph(left[0], label_style),
            money_cell(left[1]) if left[1] is not None else "",
            money_cell(left[2]) if left[2] is not None else "",
            Paragraph(right[0], label_style),
            money_cell(right[1]) if right[1] is not None else "",
        ])
    earn_actual = sum((row[1] for row in earnings), Decimal("0"))
    earn_paid = sum((row[2] for row in earnings), Decimal("0"))
    deduct_total = sum((amount for _label, amount in deductions), Decimal("0"))
    subtotal_row = len(rows)
    rows.append([Paragraph("Sub Total", value_style), money_cell(earn_actual, True), money_cell(earn_paid, True),
                 Paragraph("Sub Total", value_style), money_cell(deduct_total, True)])

    others_row = len(rows)
    rows.append([Paragraph("Others", value_style), "", "", Paragraph("Others", value_style), ""])
    other_span = max(len(other_earnings), len(other_deductions), 1)
    for index in range(other_span):
        left = other_earnings[index] if index < len(other_earnings) else ("", None)
        right = other_deductions[index] if index < len(other_deductions) else ("", None)
        rows.append([
            Paragraph(left[0], label_style), "",
            money_cell(left[1]) if left[1] is not None else "",
            Paragraph(right[0], label_style),
            money_cell(right[1]) if right[1] is not None else "",
        ])
    other_earn_total = sum((amount for _label, amount in other_earnings), Decimal("0"))
    other_deduct_total = sum((amount for _label, amount in other_deductions), Decimal("0"))
    other_subtotal_row = len(rows)
    rows.append([Paragraph("Sub Total", value_style), "", money_cell(other_earn_total, True),
                 Paragraph("Sub Total", value_style), money_cell(other_deduct_total, True)])

    gross_row = len(rows)
    rows.append([Paragraph("Gross Pay", value_style), money_cell(earn_actual, True),
                 money_cell(earn_paid + other_earn_total, True),
                 Paragraph("Gross Deductions", value_style), money_cell(deduct_total + other_deduct_total, True)])

    net = Decimal(result.final_salary or 0) if result and result.final_salary is not None else Decimal("0")
    net_row = len(rows)
    rows.append([Paragraph("Net Pay", band_style), "", "", "",
                 Paragraph(f"{pdf_money(net)} INR", value_style)])
    words_row = len(rows)
    rows.append([Paragraph(f"In words: {money_in_words(net)}", words_style), "", "", "", ""])

    column_widths = [available_width * 0.22, available_width * 0.16, available_width * 0.16,
                     available_width * 0.28, available_width * 0.18]
    grid = Table(rows, colWidths=column_widths)
    grid.setStyle(TableStyle(hairline + [
        ("SPAN", (0, 0), (2, 0)), ("SPAN", (3, 0), (4, 0)),
        ("SPAN", (0, others_row), (2, others_row)), ("SPAN", (3, others_row), (4, others_row)),
        ("SPAN", (0, net_row), (3, net_row)),
        ("SPAN", (0, words_row), (4, words_row)),
        ("ALIGN", (1, 1), (2, -1), "RIGHT"), ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("BACKGROUND", (0, 0), (-1, 1), SURFACE_SOFT),
        ("BACKGROUND", (0, subtotal_row), (-1, subtotal_row), SURFACE_SOFT),
        ("BACKGROUND", (0, other_subtotal_row), (-1, other_subtotal_row), SURFACE_SOFT),
        ("BACKGROUND", (0, gross_row), (-1, gross_row), SURFACE_SOFT),
        ("BACKGROUND", (0, net_row), (-1, net_row), TINT_WASH),
    ]))

    # Employer side. Kept apart from the deduction grid so it can never read as money
    # taken from the employee.
    pf_employee = Decimal(getattr(result, "pf_employee", 0) or 0) if result else Decimal("0")
    pf_employer = Decimal(getattr(result, "pf_employer", 0) or 0) if result else Decimal("0")
    esi_employee = Decimal(getattr(result, "esi_employee", 0) or 0) if result else Decimal("0")
    esi_employer = Decimal(getattr(result, "esi_employer", 0) or 0) if result else Decimal("0")
    edli = Decimal(getattr(result, "pf_edli", 0) or 0) if result else Decimal("0")
    admin = Decimal(getattr(result, "pf_admin", 0) or 0) if result else Decimal("0")
    # EDLI and the administration charge are both employer-borne PF costs, so they
    # are shown as one line rather than two.
    contributions = Table([
        [Paragraph("PF &amp; ESIC Contributions", band_style), "", "", ""],
        [Paragraph("PF contribution by Employer", label_style), money_cell(pf_employer),
         Paragraph("ESIC contribution by Employer", label_style), money_cell(esi_employer)],
        [Paragraph("PF contribution by Employee", label_style), money_cell(pf_employee),
         Paragraph("ESIC contribution by Employee", label_style), money_cell(esi_employee)],
        [Paragraph("PF Admin &amp; EDLI Charges (Paid by Employer)", label_style),
         money_cell(edli + admin), "", ""],
        [Paragraph("Total contribution to PF account", value_style),
         money_cell(pf_employee + pf_employer, True), "", ""],
    ], colWidths=[available_width * 0.32, available_width * 0.18, available_width * 0.32, available_width * 0.18])
    contributions.setStyle(TableStyle(hairline + [
        ("SPAN", (0, 0), (3, 0)),
        ("SPAN", (2, 3), (3, 3)), ("SPAN", (2, 4), (3, 4)),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"), ("ALIGN", (3, 1), (3, 2), "RIGHT"),
        ("BACKGROUND", (0, 0), (-1, 0), SURFACE_SOFT),
        ("BACKGROUND", (0, 4), (-1, 4), SURFACE_SOFT),
    ]))

    opening = Decimal(result.opening_leave or 0) if result else Decimal("0")
    earned = Decimal(result.leave_earned or 0) if result else Decimal("0")
    used = Decimal(result.leave_used or 0) if result else Decimal("0")
    closing = Decimal(result.closing_leave or 0) if result else Decimal("0")
    leave_block = Table([
        [Paragraph("LEAVE SUMMARY", band_style), "", "", ""],
        [Paragraph("Balance from last month", label_style), Paragraph("Earned this month", label_style),
         Paragraph("Used this month", label_style), Paragraph("Carry forward", label_style)],
        [Paragraph(str(leave_days(opening)), value_style), Paragraph(str(leave_days(earned)), value_style),
         Paragraph(str(leave_days(used)), value_style), Paragraph(str(leave_days(closing)), value_style)],
    ], colWidths=[available_width * 0.25] * 4)
    leave_block.setStyle(TableStyle(hairline + [
        ("SPAN", (0, 0), (3, 0)),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, 1), SURFACE_SOFT),
    ]))

    story = [masthead, band, details, grid, Spacer(1, 6), contributions]
    # Daily wage employees have no leave, so the block is monthly only.
    if result and result.payroll_rule_type != "DAILY":
        story.extend([Spacer(1, 6), leave_block])
    return story


def employee_report_block(month, salary, result, styles, compact=False):
    header = _salary_slip_header(month, salary, result, styles, compact=compact)
    if compact:
        words_style = ParagraphStyle("PayWords", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=5.4, leading=6.2, textColor=colors.HexColor("#0C306A"), alignment=TA_CENTER)
        summary_rows = employee_compact_summary_rows(salary, result)
        if result:
            summary_rows[0][3] = Paragraph(summary_rows[0][3], words_style)
        summary = _table(
            summary_rows, col_widths=[34 * mm, 28 * mm, 38 * mm, 94 * mm], font_size=5.0,
            header=False, highlight_rows=[0] if result else [],
            accent_rows=attendance_bonus_accent_rows(summary_rows, result),
            center=True,
        )
        detail = attendance_calendar_table(month, result, styles, 194 * mm, compact=True)
        return KeepTogether([header, Spacer(1, 2), summary, Spacer(1, 2), detail])
    return KeepTogether([
        header,
        Spacer(1, 6),
        _table(employee_salary_summary_rows(salary, result), col_widths=[55 * mm, 65 * mm], header=False, center=True),
        Spacer(1, 8),
        attendance_calendar_table(month, result, styles, 190 * mm),
    ])


def build_employee_pdf(month, employee_id):
    """One employee's payroll document.

    Monthly wage gets a salary slip. Daily wage does not receive one, so the request
    is answered with that employee's attendance summary instead.
    """
    employee, salary, result = result_for_employee(month, employee_id)
    if salary and salary.normalized_salary_type == "DAILY":
        return build_attendance_summary_pdf(month, "DAILY", employee_id=employee_id)
    employee_name = salary.name if salary else (employee.name if employee else employee_id)
    title = f"{employee_name} Salary Slip - {display_month(month)}"
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=8 * mm, rightMargin=8 * mm, topMargin=8 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    available_width = A4[0] - doc.leftMargin - doc.rightMargin
    story = salary_slip_story(month, salary, result, styles, available_width) if salary else []
    page_callback = _salary_slip_page_callback(title)
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


def salaries_for_wage_group(month, wage_group):
    return (
        SalaryRecord.query.filter_by(payroll_month=month, normalized_salary_type=wage_group)
        .all()
    )


def build_attendance_summary_pdf(month, wage_group, employee_id=None):
    """Attendance sheets for a wage group, one employee to a page.

    MONTHLY keeps the SMARTfill logo. DAILY carries no branding at all. Passing
    `employee_id` narrows the file to that one employee.
    """
    variant = DAILY_SUMMARY if wage_group == "DAILY" else MONTHLY_SUMMARY
    title = "Summary for Daily Wage Group" if wage_group == "DAILY" else "Attendance Summary for Monthly"
    salaries = sorted(salaries_for_wage_group(month, wage_group), key=lambda s: employee_id_sort_key(s.employee_id))
    if employee_id is not None:
        salaries = [item for item in salaries if item.employee_id == employee_id]
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    available_width = A4[0] - doc.leftMargin - doc.rightMargin
    results = {r.employee_id: r for r in PayrollResult.query.filter_by(payroll_month=month).all()}
    story = []
    for index, salary in enumerate(salaries):
        if index:
            story.append(PageBreak())
        story.append(attendance_summary_block(month, salary, results.get(salary.employee_id), styles, variant, available_width))
    if not story:
        story.append(Paragraph(
            f"No {wage_group.lower()} wage employees found for {display_month(month)}.",
            ParagraphStyle("Empty", parent=styles["Normal"], fontSize=11, textColor=FAINT),
        ))
    page_callback = _titled_page_callback(f"{title} - {display_month(month)}", variant=variant)
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


def employee_id_sort_key(value):
    return (0, int(value), "") if str(value).isdigit() else (1, 0, str(value).lower())


def build_all_employees_pdf(month):
    """Salary slips for the monthly wage group.

    Daily wage employees do not receive a salary slip; they get the unbranded
    attendance summary instead, so they are excluded here rather than being given a
    document the company does not issue to them.
    """
    salaries = sorted(salaries_for_wage_group(month, "MONTHLY"), key=lambda s: employee_id_sort_key(s.employee_id))
    title = f"Salary Slips - {display_month(month)}"
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=8 * mm, rightMargin=8 * mm, topMargin=8 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    available_width = A4[0] - doc.leftMargin - doc.rightMargin
    story = []
    results = {r.employee_id: r for r in PayrollResult.query.filter_by(payroll_month=month).all()}
    for index, salary in enumerate(salaries):
        # One slip to a page: each is handed to a different person.
        if index:
            story.append(PageBreak())
        story.extend(salary_slip_story(month, salary, results.get(salary.employee_id), styles, available_width))
    page_callback = _salary_slip_page_callback(title)
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


# This report goes to department managers to track attendance, so it carries no
# salary, deduction or payable figures anywhere. Monthly employees are measured on
# leave; daily wage employees have no leave, so they are measured on absence and the
# attendance bonus instead.
# Column widths are sized so every header clears its neighbour at 6.6pt; both rows
# total the 190mm of usable width on portrait A4.
DEPARTMENT_MONTHLY_HEADERS = [
    "ID", "Employee", "Designation", "Days", "Working", "Week Off",
    "Total Paid", "LOP", "Leave Used", "Leave Earned", "Leave CF", "Status",
]
DEPARTMENT_MONTHLY_WIDTHS = [
    9 * mm, 27 * mm, 23 * mm, 11 * mm, 15 * mm, 14 * mm,
    15 * mm, 11 * mm, 16 * mm, 19 * mm, 13 * mm, 17 * mm,
]
DEPARTMENT_DAILY_HEADERS = [
    "ID", "Employee", "Designation", "Days", "Working", "Week Off",
    "Total Paid", "Holidays", "Absence", "Bonus", "Status",
]
DEPARTMENT_DAILY_WIDTHS = [
    9 * mm, 30 * mm, 26 * mm, 11 * mm, 15 * mm, 14 * mm,
    15 * mm, 14 * mm, 22 * mm, 13 * mm, 21 * mm,
]
UNASSIGNED_DEPARTMENT = "Not Assigned"


def department_wise_groups(month):
    """Employees grouped by wage type then department, monthly first.

    Returns [(wage label, [(department, [(result, salary, employee)], subtotal)])].
    Departments sort alphabetically with unassigned last, so a missing department
    never hides at the top of the report.
    """
    employees = {employee.id: employee for employee in Employee.query.all()}
    salaries = {s.employee_id: s for s in SalaryRecord.query.filter_by(payroll_month=month).all()}
    groups = []
    for wage_group, label in (("MONTHLY", "Monthly Wage"), ("DAILY", "Daily Wage")):
        by_department = defaultdict(list)
        for result in calculated_results_for_month(month):
            salary = salaries.get(result.employee_id)
            if not salary or salary.normalized_salary_type != wage_group:
                continue
            employee = employees.get(result.employee_id)
            department = (employee.department if employee else "") or UNASSIGNED_DEPARTMENT
            by_department[department].append((result, salary, employee))
        departments = []
        for department in sorted(by_department, key=lambda name: (name == UNASSIGNED_DEPARTMENT, name.lower())):
            members = sorted(by_department[department], key=lambda item: employee_id_sort_key(item[0].employee_id))
            departments.append((department, members))
        groups.append((wage_group, label, departments))
    return groups


def department_attendance_row(result, employee, month, wage_group):
    """One employee's attendance line. No salary figure appears in either variant."""
    common = [
        result.employee_id,
        employee.name if employee else result.employee_id,
        (employee.designation if employee else "") or "",
        payroll_month_days(month),
        result.paid_working_days,
        result.week_offs,
        total_paid_days(result),
    ]
    if wage_group == "DAILY":
        return common + [result.holidays, bonus_absence_text(result), bonus_short_text(result), result.calculation_status]
    return common + [
        result.lop_days, result.leave_used, result.leave_earned, result.closing_leave,
        result.calculation_status,
    ]


def department_totals_row(members, wage_group):
    """Department footer: headcount and the attendance totals a manager adds up."""
    def day_total(getter):
        return leave_days(sum((Decimal(getter(r) or 0) for r, _s, _e in members), Decimal("0")))

    paid = day_total(lambda r: r.paid_working_days)
    total_paid = day_total(total_paid_days)
    week_offs = sum((int(r.week_offs or 0) for r, _s, _e in members), 0)
    label = f"{len(members)} employee(s)"
    if wage_group == "DAILY":
        absence = sum((int(getattr(r, "absence_minutes", 0) or 0) for r, _s, _e in members), 0)
        holidays = sum((int(r.holidays or 0) for r, _s, _e in members), 0)
        return ["", label, "Total", "", paid, week_offs, total_paid, holidays,
                minutes_to_duration(absence), "", ""]
    return ["", label, "Total", "", paid, week_offs, total_paid,
            day_total(lambda r: r.lop_days), day_total(lambda r: r.leave_used),
            day_total(lambda r: r.leave_earned), day_total(lambda r: r.closing_leave), ""]


def build_department_wise_pdf(month):
    """Attendance by department for department managers.

    Every monthly department first, then every daily one. Each department gets a
    table of its employees followed by each employee's month calendar. Deliberately
    carries no salary, deduction or payable figure: this sheet leaves the payroll
    office and goes to a manager who has no business seeing pay.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    available_width = A4[0] - doc.leftMargin - doc.rightMargin
    groups = department_wise_groups(month)
    wage_style = ParagraphStyle("WageGroup", parent=styles["Heading2"], fontName="Helvetica-Bold",
                                fontSize=13, leading=15, textColor=INK, spaceAfter=0)
    dept_style = ParagraphStyle("DeptHead", parent=styles["Normal"], fontName="Helvetica-Bold",
                                fontSize=10, leading=12, textColor=TINT_TEXT)
    person_style = ParagraphStyle("DeptPerson", parent=styles["Normal"], fontName="Helvetica-Bold",
                                  fontSize=9, leading=11, textColor=INK)
    empty_style = ParagraphStyle("DeptEmpty", parent=styles["Normal"], fontSize=8.5, textColor=FAINT)
    # Name and designation are the only free-text columns, and a long designation
    # such as "Head of Business Development" overruns its column as a plain string.
    # Only a Paragraph wraps, so those two cells are wrapped and the rest stay plain,
    # which keeps the status column readable by _status_column_style.
    text_cell = ParagraphStyle("DeptCell", parent=styles["Normal"], fontName="Helvetica",
                               fontSize=7, leading=8.4, textColor=INK)
    total_cell = ParagraphStyle("DeptCellTotal", parent=text_cell, fontName="Helvetica-Bold",
                                textColor=TINT_TEXT)
    head_left = ParagraphStyle("DeptHeadCell", parent=styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=6.6, leading=8, textColor=MUTED)
    head_centre = ParagraphStyle("DeptHeadCellC", parent=head_left, alignment=TA_CENTER)

    def wrap_text_columns(row, style):
        row[1] = Paragraph(str(row[1] or ""), style)
        row[2] = Paragraph(str(row[2] or ""), style)
        return row

    def header_row(names):
        # Wrapped so a long heading stacks inside its column instead of running into
        # the next one. Cells 3 onwards are centred to match the numbers below them.
        return [Paragraph(name, head_left if index < 3 else head_centre)
                for index, name in enumerate(names)]

    headcount = sum(len(members) for _g, _l, departments in groups for _d, members in departments)
    story = [
        _report_brand_header("Department Wise Attendance Summary", display_month(month), styles, available_width),
        Spacer(1, 9),
        _kpi_row([
            ("Employees", headcount),
            ("Departments", len({d for _g, _l, deps in groups for d, _m in deps})),
            ("Days in month", payroll_month_days(month)),
            ("Payroll month", display_month(month)),
        ], available_width),
        Spacer(1, 10),
    ]
    for group_index, (wage_group, label, departments) in enumerate(groups):
        if group_index:
            story.append(PageBreak())
        story.extend([Paragraph(f"{label} Employees", wage_style), Spacer(1, 6)])
        if not departments:
            story.extend([Paragraph(f"No {label.lower()} employees for this month.", empty_style), Spacer(1, 9)])
            continue
        daily = wage_group == "DAILY"
        headers = DEPARTMENT_DAILY_HEADERS if daily else DEPARTMENT_MONTHLY_HEADERS
        widths = DEPARTMENT_DAILY_WIDTHS if daily else DEPARTMENT_MONTHLY_WIDTHS
        for department_index, (department, members) in enumerate(departments):
            # Each department starts a fresh page so a single department's sheet can
            # be torn off and handed to its manager without another one's rows on it.
            # The first department of a group already sits on a fresh page.
            if department_index:
                story.append(PageBreak())
            rows = [header_row(headers)]
            for result, _salary, employee in members:
                rows.append(wrap_text_columns(
                    department_attendance_row(result, employee, month, wage_group), text_cell))
            rows.append(wrap_text_columns(department_totals_row(members, wage_group), total_cell))
            table = _table(rows, col_widths=widths, font_size=7,
                           status_column=len(headers) - 1, center_from=3,
                           accent_rows=[(len(rows) - 1, TINT_WASH, TINT_TEXT)])
            story.extend([
                KeepTogether([
                    Paragraph(f"{department} &mdash; {len(members)} employee(s)", dept_style),
                    Spacer(1, 4), table,
                ]),
                Spacer(1, 8),
            ])
            # Each employee's month, so a manager can see the pattern behind the totals.
            for result, _salary, employee in members:
                name = employee.name if employee else result.employee_id
                role = (employee.designation if employee else "") or "Designation not set"
                calendar = attendance_calendar_table(
                    month, result, styles, available_width,
                    variant=MONTHLY_SUMMARY, sizes=(7.6, 6, 6),
                )
                story.extend([
                    KeepTogether([
                        Paragraph(f"{result.employee_id} &middot; {name} &mdash; {role}", person_style),
                        Spacer(1, 3), calendar,
                    ]),
                    Spacer(1, 7),
                ])
    page_callback = _titled_page_callback(f"Department Wise Attendance Summary - {display_month(month)}")
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


# The salary register, laid out like the sheet the payroll office already keeps.
# Column names follow that sheet so the two can be read side by side.
SALARY_REGISTER_HEADERS = [
    "Sr.NO", "ID", "NAME", "Attendance", "WO", "Occasional Leave", "Paid Leave",
    "BASIC", "HRA", "ALLOWANCE", "PAID BASIC", "PAID HRA", "PAID ALLOWANCE",
    "OA", "LATE REPORTING", "SHORT LEAVE", "LOAN", "ESI", "PF",
    "PROFESSIONAL TAX", "TDS", "NET SALARY",
]
# Columns holding money, used to right-align and total them in both outputs.
SALARY_REGISTER_MONEY_COLUMNS = tuple(range(7, 22))
SALARY_REGISTER_TOTAL_COLUMNS = (13, 14, 15, 16, 17, 18, 19, 20, 21)


def salary_register_rows(month):
    """One row per calculated monthly-wage employee, in the payroll sheet's order.

    `PAID BASIC/HRA/ALLOWANCE` are the contracted components scaled by what the
    employee actually earned, which is how the existing sheet derives them.
    """
    employees = {employee.id: employee for employee in Employee.query.all()}
    salaries = {s.employee_id: s for s in SalaryRecord.query.filter_by(payroll_month=month).all()}
    rows = []
    for result in sorted(calculated_results_for_month(month), key=lambda r: employee_id_sort_key(r.employee_id)):
        salary = salaries.get(result.employee_id)
        if not salary or salary.normalized_salary_type != "MONTHLY":
            continue
        employee = employees.get(result.employee_id)
        # Paid components reflect loss of pay only. Short hours has its own column,
        # so folding it in here would report the reduction twice.
        ratio = slip_paid_ratio(salary, result)
        basic = Decimal(employee.basic_salary or 0) if employee else Decimal("0")
        hra = Decimal(employee.hra or 0) if employee else Decimal("0")
        allowance = Decimal(employee.allowance or 0) if employee else Decimal("0")
        rows.append([
            len(rows) + 1,
            result.employee_id,
            salary.name,
            result.paid_working_days,
            result.week_offs,
            result.lop_days,
            result.paid_leaves,
            money(basic), money(hra), money(allowance),
            money(basic * ratio), money(hra * ratio), money(allowance * ratio),
            money(result.ot_amount),
            money(result.less_hours_deduction),
            money(result.lop_deduction),
            money(Decimal(result.loan_deduction or 0) + Decimal(result.advance_deduction or 0)),
            money(result.esi_employee),
            money(result.pf_employee),
            money(getattr(result, "professional_tax", 0)),
            money(getattr(result, "tds", 0)),
            money(result.final_salary),
        ])
    return rows


def salary_register_totals(rows):
    totals = {index: Decimal("0") for index in SALARY_REGISTER_TOTAL_COLUMNS}
    for row in rows:
        for index in totals:
            totals[index] += Decimal(row[index] or 0)
    return totals


def build_salary_register_xlsx(month):
    """The register as a spreadsheet, mirroring the sheet's own header block."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = salary_register_rows(month)
    totals = salary_register_totals(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = month

    last_column = len(SALARY_REGISTER_HEADERS)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title = sheet.cell(1, 1, f"Salary Sheet - {display_month(month)}")
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="center")

    sheet.cell(2, 2, "Days of Month").font = Font(bold=True)
    sheet.cell(2, 3, payroll_month_days(month))
    # The rate sits directly above the column it drives, as it does on the existing
    # sheet, so the header row below reads as its label.
    esi_column = SALARY_REGISTER_HEADERS.index("ESI") + 1
    pf_column = SALARY_REGISTER_HEADERS.index("PF") + 1
    pt_column = SALARY_REGISTER_HEADERS.index("PROFESSIONAL TAX") + 1
    esi_rate = sheet.cell(2, esi_column, float(STATUTORY_RULES["ESI_EMPLOYEE_PERCENT"] / 100))
    pf_rate = sheet.cell(2, pf_column, float(STATUTORY_RULES["PF_EMPLOYEE_PERCENT"] / 100))
    for cell in (esi_rate, pf_rate):
        cell.number_format = "0.00%"
    _threshold, amount = PROFESSIONAL_TAX_SLABS[0]
    sheet.cell(2, pt_column, float(amount)).number_format = "#,##0"

    head_fill = PatternFill("solid", fgColor="F2F2F7")
    thin = Side(style="thin", color="D8D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, name in enumerate(SALARY_REGISTER_HEADERS, start=1):
        cell = sheet.cell(3, column, name)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for offset, row in enumerate(rows):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(4 + offset, column)
            # Numbers go in as numbers so the sheet stays usable for further work.
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.border = border
            if column - 1 in SALARY_REGISTER_MONEY_COLUMNS:
                cell.number_format = "#,##0.00"

    total_row = 4 + len(rows)
    label = sheet.cell(total_row, 3, f"Total - {len(rows)} employee(s)")
    label.font = Font(bold=True)
    for index in SALARY_REGISTER_TOTAL_COLUMNS:
        cell = sheet.cell(total_row, index + 1, float(totals[index]))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"
        cell.border = border

    widths = [7, 7, 26] + [13] * (last_column - 3)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "D4"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# Sized so no heading wraps mid-word; A3 landscape leaves 400mm of usable width.
SALARY_REGISTER_WIDTHS = [
    12 * mm, 10 * mm, 38 * mm, 18 * mm, 10 * mm, 18 * mm, 14 * mm,
    16 * mm, 15 * mm, 18 * mm, 18 * mm, 16 * mm, 20 * mm,
    13 * mm, 18 * mm, 15 * mm, 14 * mm, 13 * mm, 14 * mm, 18 * mm, 14 * mm, 21 * mm,
]


def build_salary_register_pdf(month):
    """The same register as a landscape sheet for printing and signature."""
    rows = salary_register_rows(month)
    totals = salary_register_totals(rows)
    buffer = BytesIO()
    pagesize = landscape(A3)
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
    head_style = ParagraphStyle("RegHead", parent=styles["Normal"], fontName="Helvetica-Bold",
                                fontSize=6, leading=7.4, textColor=MUTED, alignment=TA_CENTER)
    name_style = ParagraphStyle("RegName", parent=styles["Normal"], fontName="Helvetica",
                                fontSize=6.6, leading=8, textColor=INK)
    empty_style = ParagraphStyle("RegEmpty", parent=styles["Normal"], fontSize=8, textColor=FAINT)

    table_rows = [[Paragraph(name, head_style) for name in SALARY_REGISTER_HEADERS]]
    for row in rows:
        cells = [pdf_money(value) if index in SALARY_REGISTER_MONEY_COLUMNS else value
                 for index, value in enumerate(row)]
        cells[2] = Paragraph(str(cells[2]), name_style)
        table_rows.append(cells)
    if rows:
        total_cells = [""] * len(SALARY_REGISTER_HEADERS)
        total_cells[2] = Paragraph(f"Total &mdash; {len(rows)} employee(s)", name_style)
        for index in SALARY_REGISTER_TOTAL_COLUMNS:
            total_cells[index] = pdf_money(totals[index])
        table_rows.append(total_cells)

    story = [
        _report_brand_header("Salary Sheet", display_month(month), styles, available_width),
        Spacer(1, 8),
    ]
    if rows:
        story.append(_table(table_rows, col_widths=SALARY_REGISTER_WIDTHS, font_size=6.6, center_from=3,
                            accent_rows=[(len(table_rows) - 1, TINT_WASH, TINT_TEXT)]))
    else:
        story.append(Paragraph(f"No monthly wage employees calculated for {display_month(month)}.", empty_style))
    page_callback = _titled_page_callback(f"Salary Sheet - {display_month(month)}")
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()


def loan_summary_rows(loan, employee, month, schedule):
    return [
        ["Employee", f"{loan.employee_id} - {employee.name if employee else 'Unknown'}", "Status", "Active" if loan.is_active else "Inactive"],
        ["Loan Start Date", loan.start_date, "Expected End Date", schedule[-1]["date"] if schedule else "N/A"],
        ["Loan Amount", pdf_money(loan.amount), "Monthly Deduction", pdf_money(loan.monthly_deduction)],
        ["Planned Tenure", f"{loan.tenure_months} month(s)", "Report Month", month],
        ["Paid Before Report Month", pdf_money(loan_paid_before_month(loan, month)), "Remaining Before Report Month", pdf_money(loan_remaining_before_month(loan, month))],
        ["This Month Installment", pdf_money(loan_installment_for_loan(loan, month)), "Pending Loan Amount", pdf_money(loan_pending_after_month(loan, month))],
        ["Notes", loan.notes or "No notes", "", ""],
    ]


def loan_schedule_rows(schedule):
    rows = [["Sr No", "Installment Month", "Installment Date", "Amount", "Status", "Pending After", "Notes"]]
    for index, item in enumerate(schedule, start=1):
        rows.append([
            index,
            item["month"],
            item["date"],
            pdf_money(item["installment"]),
            item["status"],
            pdf_money(item["remaining_after"]),
            item["notes"],
        ])
    return rows


def build_loan_pdf(loan_id, month):
    loan = db.session.get(Loan, loan_id)
    if not loan:
        raise ValueError("Loan was not found.")
    employee = db.session.get(Employee, loan.employee_id)
    title = f"Loan #{loan.id} Summary - {employee.name if employee else loan.employee_id}"
    schedule = loan_repayment_schedule(loan, month)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=12 * mm, rightMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    available_width = A4[0] - doc.leftMargin - doc.rightMargin
    story = [
        _report_brand_header(f"Loan #{loan.id} Summary", f"Employee: {employee.name if employee else loan.employee_id}", styles, available_width),
        Spacer(1, 8),
        _table(loan_summary_rows(loan, employee, month, schedule), col_widths=[38 * mm, 58 * mm, 42 * mm, 58 * mm], font_size=8, header=False, highlight_rows=[5]),
        Spacer(1, 8),
        Paragraph("Installment Schedule", ParagraphStyle("LoanSection", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=10, textColor=colors.HexColor("#0C306A"), spaceAfter=4)),
        _table(loan_schedule_rows(schedule), col_widths=[12 * mm, 28 * mm, 28 * mm, 28 * mm, 25 * mm, 30 * mm, 35 * mm], font_size=7),
    ]
    page_callback = _titled_page_callback(title)
    doc.build(story, onFirstPage=page_callback, onLaterPages=page_callback)
    buffer.seek(0)
    return buffer.getvalue()
